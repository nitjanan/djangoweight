from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect, HttpResponseBadRequest, StreamingHttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.views.decorators.cache import cache_page
from weightapp.models import Weight, Production, BaseLossType, ProductionLossItem, BaseMill, BaseLineType, ProductionGoal, StoneEstimate, StoneEstimateItem, BaseStoneType, BaseTimeEstimate, BaseCustomer, BaseSite, WeightHistory, BaseTransport, BaseCar, BaseScoop, BaseCarTeam, BaseCar, BaseDriver, BaseCarRegistration, BaseJobType, BaseCustomerSite, UserScale, BaseMachineType, BaseCompany, UserProfile, BaseSEC, SetWeightOY, SetCompStone, SetPatternCode, Stock, StockStone, StockStoneItem, BaseStockSource, ApproveWeight, SetLineMessaging, GasPrice, BaseSiteStore, PortStock, PortStockStone, PortStockStoneItem, ProductionMachineItem, BaseWeightRange, LoadingRate, LoadingRateLoc, LoadingRateItem
from django.db.models import Sum, Q, Max, Value
from decimal import Decimal, InvalidOperation
from django.views.decorators.cache import cache_control
from django.contrib.auth.forms import AuthenticationForm
from django.core.paginator import Paginator
from .filters import WeightFilter, ProductionFilter, StoneEstimateFilter, BaseMillFilter, BaseStoneTypeFilter, BaseScoopFilter, BaseCarTeamFilter, BaseCarFilter, BaseSiteFilter, BaseCustomerFilter, BaseDriverFilter, BaseCarRegistrationFilter, BaseJobTypeFilter, BaseCustomerSiteFilter, StockFilter, GasPriceFilter, PortStockFilter, LoadingRateFilter
from .forms import ProductionForm, ProductionLossItemForm, ProductionModelForm, ProductionLossItemFormset, ProductionLossItemInlineFormset, ProductionGoalForm, StoneEstimateForm, StoneEstimateItemInlineFormset, WeightForm, WeightStockForm, BaseMillForm, BaseStoneTypeForm ,BaseScoopForm, BaseCarTeamForm, BaseCarForm, BaseSiteForm, BaseCustomerForm, BaseDriverForm, BaseCarRegistrationForm, BaseJobTypeForm, BaseCustomerSiteForm, StockForm, StockStoneForm, StockStoneItemForm, StockStoneItemInlineFormset, GasPriceForm, WeightPortForm, PortStockForm, PortStockStoneForm, PortStockStoneItemInlineFormset, ProductionMachineItemInlineFormset, LoadingRateForm, LoadingRateLocForm, LoadingRateItemInlineFormset
import xlwt
from django.db.models import Count, Avg
import stripe, logging, datetime
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Color, NamedStyle, Side, Border
from openpyxl.utils import get_column_letter
from datetime import date, timedelta, datetime, time
from django.views import generic
from django.forms import formset_factory, modelformset_factory, inlineformset_factory, Select
from django import forms
from django.db.models import Sum, Subquery
import random
from django.db.models.functions import Coalesce, ExtractMonth, ExtractYear, TruncMonth, TruncYear
from django.db.models import F, ExpressionWrapper, Case, When, OuterRef, Exists
from django.db import models
import pandas as pd
import calendar
from collections import defaultdict, OrderedDict
from re import escape as reescape
from django.db.models import Value as V
from django.db.models.functions import Cast, Concat
from django.contrib.auth.decorators import login_required
from django.urls import reverse

from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework import generics, viewsets, permissions, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.request import Request
from rest_framework.renderers import TemplateHTMLRenderer
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination

from weightapp.serializers import BaseScoopSerializer, BaseMillSerializer, WeightSerializer, BaseCustomerSerializer, BaseStoneTypeSerializer, BaseCarTeamSerializer, BaseDriverSerializer, BaseCarRegistrationSerializer, BaseCarRegistrationSerializer, BaseCarSerializer, BaseSiteSerializer, BaseCarSerializer, BaseStoneTypeTestSerializer, BaseJobTypeSerializer, SignUpSerializer, BaseCustomerSiteSerializer, CarPartnerSerializer
from rest_framework.decorators import api_view
from django.contrib.auth.models import User
from django.db import IntegrityError
from .tokens import create_jwt_pair_for_user
import csv
from io import StringIO
from decimal import Decimal
import ast
import json
from django.conf import settings # calls the object written in settings.py
from django.views.decorators.csrf import csrf_exempt
from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import MessageEvent, TextMessage, TextSendMessage
from apscheduler.schedulers.background import BackgroundScheduler
from zoneinfo import ZoneInfo
from apscheduler.triggers.cron import CronTrigger
import requests
from itertools import groupby
from operator import itemgetter
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from io import BytesIO
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.styles import Protection
import time
import re

line_bot_api = LineBotApi(settings.LINE_CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(settings.LINE_CHANNEL_SECRET)

''' สำหรับเช็ค UserId, groupId, roomId โดยส่งข้อความไปทาง line ที่มี BOT อยู่
@csrf_exempt
def callback(request):
    if request.method == 'POST':
        body = request.body.decode('utf-8')
        req = json.loads(body)
        print('Full Payload:', json.dumps(req, indent=2))

        intent = req["queryResult"]["intent"]["displayName"]
        # Extract common data
        source = req['originalDetectIntentRequest']['payload']['data']['source']
        source_type = source.get('type')
        user_id = source.get('userId')
        group_id = source.get('groupId') if source_type == 'group' else None
        room_id = source.get('roomId') if source_type == 'room' else None

        # Log context
        print('Source Type:', source_type)
        print('User ID:', user_id)
        print('Group ID:', group_id)
        print('Room ID:', room_id)

        # Handle display name if userId is present
        display_name = None
        if user_id:
            profile = line_bot_api.get_profile(user_id)
            display_name = profile.display_name
            print('Display Name:', display_name)

        # Call reply function
        reply_token = req['originalDetectIntentRequest']['payload']['data']['replyToken']
        text = req['originalDetectIntentRequest']['payload']['data']['message']['text']

        # Call reply function
        reply(intent, text, reply_token, user_id, display_name)

        return JsonResponse({'status': 'OK'})
    return JsonResponse({'error': 'Invalid request method'}, status=405)

def reply(intent, text, reply_token, user_id, display_name):
        weight = (
            Weight.objects.filter(bws__weight_type = 1)
            .values('customer_name')
            .annotate(sum_weight_total=Sum('weight_total'))
            .order_by('-sum_weight_total')[:3]
        )
 
        # Prepare the message text
        messages = []
        for i in weight:
            tmp_text = f"ลูกค้า {i['customer_name']} : {i['sum_weight_total']} ตัน"
            messages.append(tmp_text)

        # Combine messages into a single text
        final_message = "\n".join(messages)

        text_message = TextSendMessage(text = final_message)
        line_bot_api.reply_message(reply_token, text_message)
'''
def split_message(message, max_length=4999):
    return [message[i:i+max_length] for i in range(0, len(message), max_length)]

def send_weight_edit(start_time, end_time, target_user_id):
    today = datetime.today().strftime("วันที่ %d/%m/%Y")
    messages = []
    final_message = ''
    error = ''

    try:
        weight = WeightHistory.objects.filter(
                bws__weight_type=1,
                user_update__isnull=False,
                v_stamp__gte=start_time,
                v_stamp__lt=end_time
        ).values('weight_id', 'doc_id', 'bws__company__name', 'date').annotate(last_v_stamp=Max('v_stamp'))
        
        # Group the data by company name
        grouped_weights = defaultdict(list)
        for i in weight:
            grouped_weights[i['bws__company__name']].append(i)

        for company_name, weights in grouped_weights.items():
            company_message = f"========== {company_name} =========="
            messages.append(company_message)
                
            for idx, i in enumerate(weights, start=1):
                tmp_time = i['date'].strftime("@%d/%m/%Y")
                tmp_text = f"{idx}) เลขที่ {i['doc_id']} {tmp_time}"
                messages.append(tmp_text)

            messages.append("\n")

    except Exception as e:
        error = 'start_time = ' + str(start_time) + ', end_time = ' + str(end_time) + ', error : ' + str(e)

    # Combine messages into a single text
    if messages:
        final_message = "🚨 "+ today + " มีการแก้ไขรายการชั่ง"+ "\n" + "\n".join(messages)
        split_messages = split_message(final_message)
        for msg in split_messages:
            # Send the message
            text_message = TextSendMessage(text=msg)
            line_bot_api.push_message(target_user_id, text_message)
    elif error:
        # Send the message
        text_message = TextSendMessage(text=error)
        line_bot_api.push_message(target_user_id, text_message)
    else:
        final_message = "✅ ไม่มีข้อมูลการแก้ไขรายการชั่งของ " + today
        # Send the message
        text_message = TextSendMessage(text=final_message)
        line_bot_api.push_message(target_user_id, text_message)

#old send_1pm_summary อันใหม่เปลี่ยนเป็น url
def send_line_1pm(request):
    try:
        lm = SetLineMessaging.objects.get(id=1)
        end_time = datetime.now().replace(hour=13, minute=0, second=0, microsecond=0)# Time range: previous day 3:00 PM to today 11:00 AM
        start_time = end_time - timedelta(hours=24)

        if lm.target_id:
            target_user_id = lm.target_id #user/group ID (Line id)
            send_weight_edit(start_time, end_time, target_user_id)

    except SetLineMessaging.DoesNotExist:
        pass

    return HttpResponse("1PM summary sent (or skipped).")

'''
# Schedule the tasks
scheduler = BackgroundScheduler()
scheduler.add_job(
    send_1pm_summary,
    trigger=CronTrigger(hour=13, minute=0),
    id="1pm_summary",
    replace_existing=True,
)
scheduler.start()
'''

#generate Code Base
def generateCodeId(model_name, type, wt, middle):
    missing_id = None

    try:
        if wt:
            spc = SetPatternCode.objects.get(m_name = model_name, wt_id = wt)
        else:
            spc = SetPatternCode.objects.get(m_name = model_name)

        if type == 1:
            id_pt = spc.pattern
            start_id = spc.start + id_pt
            end_id = spc.end + id_pt
        elif type == 2:
            id_pt = spc.pattern
            start_id = id_pt + spc.start
            end_id = id_pt + spc.end
        elif type == 3:
            id_pt = middle + spc.pattern
            start_id = id_pt + spc.start
            end_id = id_pt + spc.end

        number_start = int(spc.start)
        number_end = int(spc.end)

        model_class = spc.get_model()
        if model_class:
            pk_field = model_class._meta.pk.name

            ids_in_range = model_class.objects.filter(
                Q(**{f"{pk_field}__gte": start_id}) & Q(**{f"{pk_field}__lte": end_id})
            ).values_list(pk_field, flat=True)

        # Convert to a set for faster lookup
        ids_set = set(ids_in_range)

        # Helper function to generate IDs in the given format
        def generate_id(number):
            tmp_id = None
            if type == 1:
                tmp_id = f"{str(number).zfill(len(spc.start)) + id_pt}"
            elif type == 2:
                tmp_id = f"{id_pt + str(number).zfill(len(spc.start))}"
            elif type == 3:
                tmp_id = f"{id_pt + str(number).zfill(len(spc.start))}"
            return tmp_id

        # Iterate from number_start to number_end to find the missing ID
        for i in range(number_start, number_end + 1):
            candidate_id = generate_id(i)
            if candidate_id not in ids_set:
                missing_id = candidate_id
                break
    except:
        pass

    return missing_id

def generateOilCustomerId(car_team_id):
    run_number = car_team_id[1:]
    return f"{'90-V-' + str(run_number).zfill(3)}"

def findCompanyIn(request):
    code = request.session['company_code']

    #หาหน้าต่างการมองเห็นบริษัททั้งหมดของ user
    user_profile = UserProfile.objects.get(user = request.user.id)
    company_all = BaseCompany.objects.filter(userprofile = user_profile).values('code')

    if code == "ALL":
        company_in = company_all
    else:
        company_in = BaseCompany.objects.filter(code = code).values('code')
    return company_in

def format_duration(duration):
    result = None
    if duration:
        hours = duration // timedelta(hours=1)
        minutes = (duration % timedelta(hours=1)) // timedelta(minutes=1)
        result = f"{hours:02d}:{minutes:02d}"
    return result

def generate_pastel_color():
    # Generate random pastel colors by restricting the RGB channels within a specific range
    red = random.randint(150, 255)
    green = random.randint(150, 255)
    blue = random.randint(150, 255)
    return f"{red:02x}{green:02x}{blue:02x}"
    
def set_border(ws, side=None, blank=True):
    wb = ws._parent
    side = side if side else Side(border_style='thin', color='000000')
    for cell in ws._cells.values():
        cell.border = Border(top=side, bottom=side, left=side, right=side)
    if blank:
        white = Side(border_style='thin', color='FFFFFF')
        wb._borders.append(Border(top=white, bottom=white, left=white, right=white))
        wb._cell_styles[0].borderId = len(wb._borders) - 1

def getSumByStone(request, mode, stoneType, type, company_in):

    ''' เปลี่ยนเป็นเลือกระหว่างวันที่ 2024-04-10
    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    start_date = datetime.strptime(startDateInMonth(str(previous_date_time.strftime('%Y-%m-%d'))), "%Y-%m-%d")
    end_date = datetime.strptime(endDateInMonth(str(previous_date_time.strftime('%Y-%m-%d'))), "%Y-%m-%d")
    '''

    start_date = request.session['db_start_date']
    end_date = request.session['db_end_date']

    start_year = datetime.strptime(start_date, '%Y-%m-%d').year

    #type 1 = sell, 2 = stock, 3 = produce, 4 = purchase 
    if type == 1:#เครื่องขาย ต้นทางจากโรงโม่
        w = Weight.objects.filter(~Q(site = '200PL') & ~Q(site = '300PL'), mill__mill_source = 1, bws__company__code__in = company_in, bws__weight_type = mode, stone_type = stoneType, date__range=(start_date, end_date)).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
    elif type == 2:
        w = StockStone.objects.filter(stk__company__code__in = company_in, stone = stoneType, stk__created__range=(start_date, end_date)).values_list('total', flat=True).order_by('-stk__created').first() or Decimal('0.0') #ดึงข้อมูลสต็อกที่เหลือจากวันที่คีย์ล่าสุด ระหว่าง start_date, end_date
        #อันเก่าดึงข้อมูลจากกองสต็อค 09-09-2024
        #w = Weight.objects.filter(Q(site__base_site_name__contains ='สต็อค') | Q(site__base_site_name__contains ='สต๊อก'), bws__company__code__in = company_in, bws__weight_type = mode, stone_type = stoneType, date__range=(start_date, end_date)).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0') 
    elif type == 3:
        if start_year > 2024:#แบบใหม่ 19-12-2024
            w = StoneEstimateItem.objects.filter(se__company__code__in = company_in, se__created__range = (start_date, end_date), stone_type = stoneType).aggregate(s=Sum("total"))["s"] or Decimal('0.0')
        else:#แบบเก่าไม่มีข้อมูล total เลยต้องคำนวณ
            w = Decimal('0.0')
            se_item = StoneEstimateItem.objects.filter(se__created__range = (start_date, end_date), stone_type = stoneType).values('se__created','percent','se__site')
            for i in se_item:
                crush = Weight.objects.filter(bws__company__code__in = company_in, site = i['se__site'], bws__weight_type = mode , date = i['se__created']).aggregate(s = Sum("weight_total"))["s"] or Decimal('0.0')
                w += calculateEstimate(i['percent'], crush)
    elif type == 4:#เครื่องขาย ต้นทางจากสต๊อก
        w = Weight.objects.filter(~Q(site = '200PL') & ~Q(site = '300PL'), mill__mill_source = 2, bws__company__code__in = company_in, bws__weight_type = mode, stone_type = stoneType, date__range=(start_date, end_date)).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
    elif type == 5:#เครื่องขาย ต้นทางจากการซื้อจากที่อื่น
        w = Weight.objects.filter(~Q(site = '200PL') & ~Q(site = '300PL'), mill__mill_source = 3, bws__company__code__in = company_in, bws__weight_type = mode, stone_type = stoneType, date__range=(start_date, end_date)).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
    elif type == 6:#รวม เครื่องขาย
        w = Weight.objects.filter(~Q(site = '200PL') & ~Q(site = '300PL'), mill__mill_source__in = [1,2,3], bws__company__code__in = company_in, bws__weight_type = mode, stone_type = stoneType, date__range=(start_date, end_date)).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')

    elif type == 10:#port stock
        w = PortStockStone.objects.filter(ps__company__code__in = company_in, stone = stoneType, ps__created__range=(start_date, end_date)).values_list('total', flat=True).order_by('-ps__created').first() or Decimal('0.0') #ดึงข้อมูลสต็อกที่เหลือจากวันที่คีย์ล่าสุด ระหว่าง start_date, end_date
    return  float(w)

def getSumOther(request, mode, list_sum_stone, type, company_in):
    ''' เปลี่ยนเป็นเลือกระหว่างวันที่ 2024-04-10
    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    start_date = datetime.strptime(startDateInMonth(str(previous_date_time.strftime('%Y-%m-%d'))), "%Y-%m-%d")
    end_date = datetime.strptime(endDateInMonth(str(previous_date_time.strftime('%Y-%m-%d'))), "%Y-%m-%d")    
    '''

    start_date = request.session['db_start_date']
    end_date = request.session['db_end_date']

    start_year = datetime.strptime(start_date, '%Y-%m-%d').year

    query_filters = Q()
    ss_query_filters = Q()
    for item_number_prefix in list_sum_stone:
        query_filters |= Q(stone_type = item_number_prefix)
        ss_query_filters |= Q(stone = item_number_prefix)

    #type 1 = sell, 2 = stock, 3 = produce, 4 = purchase 
    if type == 1:#เครื่องขาย ต้นทางจากโรงโม่
        w = Weight.objects.filter(~Q(site = '200PL') & ~Q(site = '300PL'), mill__mill_source = 1, bws__company__code__in = company_in, bws__weight_type = mode, date__range=(start_date, end_date)).exclude(query_filters).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
    elif type == 2:
        #อันเก่าดึงข้อมูลจากกองสต็อค 09-09-2024
        #w = Weight.objects.filter(bws__company__code__in = company_in, site__base_site_name__contains='สต็อค', bws__weight_type = mode, date__range=(start_date, end_date)).exclude(query_filters).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
        w = Decimal('0.0')
        try:
            qr = StockStone.objects.filter(stk__company__code__in = company_in, stk__created__range=(start_date, end_date)).exclude(ss_query_filters).values('stk__created', 'stone', 'total').order_by('-stk__created').distinct() #ดึงข้อมูลสต็อกที่เหลือจากวันที่คีย์ล่าสุด ระหว่าง start_date, end_date และรวมกัน
        except StockStone.DoesNotExist:
            qr = None

        if qr:
            qr_list = list(qr)

            # Group by 'stone', keeping the most recent 'stk__created'
            filtered_results = [
                max(group, key=itemgetter('stk__created'))
                for _, group in groupby(sorted(qr_list, key=itemgetter('stone')), key=itemgetter('stone'))
            ]
            w = sum(item['total'] for item in filtered_results)

    elif type == 3:
        if start_year > 2024:#แบบใหม่ 19-12-2024
            w = StoneEstimateItem.objects.filter(se__company__code__in = company_in, se__created__range = (start_date, end_date)).exclude(query_filters).aggregate(s=Sum("total"))["s"] or Decimal('0.0')
        else:#แบบเก่าไม่มีข้อมูล total เลยต้องคำนวณ
            w = Decimal('0.0')
            se_item = StoneEstimateItem.objects.filter(se__created__range = (start_date, end_date)).exclude(query_filters).values('se__created','percent','se__site')
            for i in se_item:
                crush = Weight.objects.filter(bws__company__code__in = company_in, site = i['se__site'], bws__weight_type = mode , date = i['se__created']).aggregate(s = Sum("weight_total"))["s"] or Decimal('0.0')
                w += calculateEstimate(i['percent'], crush)
    elif type == 4:#เครื่องขาย ต้นทางจากสต๊อก
        w = Weight.objects.filter(~Q(site = '200PL') & ~Q(site = '300PL'), mill__mill_source = 2, bws__company__code__in = company_in, bws__weight_type = mode, date__range=(start_date, end_date)).exclude(query_filters).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
    elif type == 5:#เครื่องขาย ต้นทางจากการซื้อจากที่อื่น
        w = Weight.objects.filter(~Q(site = '200PL') & ~Q(site = '300PL'), mill__mill_source = 3, bws__company__code__in = company_in, bws__weight_type = mode, date__range=(start_date, end_date)).exclude(query_filters).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
    elif type == 6:#รวมเครื่องขาย
        w = Weight.objects.filter(~Q(site = '200PL') & ~Q(site = '300PL'), mill__mill_source__in = [1,2,3], bws__company__code__in = company_in, bws__weight_type = mode, date__range=(start_date, end_date)).exclude(query_filters).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
    
    elif type == 10:#port stock
        w = Decimal('0.0')
        try:
            qr = PortStockStone.objects.filter(ps__company__code__in = company_in, ps__created__range=(start_date, end_date)).exclude(ss_query_filters).values('ps__created', 'stone', 'total').order_by('-ps__created').distinct() #ดึงข้อมูลสต็อกที่เหลือจากวันที่คีย์ล่าสุด ระหว่าง start_date, end_date และรวมกัน
        except PortStockStone.DoesNotExist:
            qr = None

        if qr:
            qr_list = list(qr)

            # Group by 'stone', keeping the most recent 'stk__created'
            filtered_results = [
                max(group, key=itemgetter('ps__created'))
                for _, group in groupby(sorted(qr_list, key=itemgetter('stone')), key=itemgetter('stone'))
            ]
            w = sum(item['total'] for item in filtered_results)

    return  float(w)

def getNumListStoneWeightChart(request, mode, stone_list_id, type, company_in):
    #sell
    list_sum_stone = []
    for stone_id in stone_list_id:
        list_sum_stone.append(getSumByStone(request, mode, stone_id, type, company_in))

    list_sum_stone.append(getSumOther(request, mode, stone_list_id, type, company_in))
    return list_sum_stone

######################################
############## Port ##################
######################################

def getPortSumByStone(request, mode, stoneType, type, company_in):
    start_date = request.session['db_start_date']
    end_date = request.session['db_end_date']

    start_year = datetime.strptime(start_date, '%Y-%m-%d').year

    #type 1 = sell, 2 = stock, 3 = produce, 4 = purchase 
    if type == 1:#เครื่องขาย ปลายทาง รับเข้า
        w = Weight.objects.filter(site__store = 1, bws__company__code__in = company_in, bws__weight_type = mode, stone_type = stoneType, date__range=(start_date, end_date)).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
    elif type == 2:
        w = StockStone.objects.filter(stk__company__code__in = company_in, stone = stoneType, stk__created__range=(start_date, end_date)).values_list('total', flat=True).order_by('-stk__created').first() or Decimal('0.0') #ดึงข้อมูลสต็อกที่เหลือจากวันที่คีย์ล่าสุด ระหว่าง start_date, end_date
    elif type == 4:#เครื่องขาย ปลายทาง จ่ายภายนอก
        w = Weight.objects.filter(site__store = 2, bws__company__code__in = company_in, bws__weight_type = mode, stone_type = stoneType, date__range=(start_date, end_date)).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
    elif type == 5:#เครื่องขาย ปลายทาง จ่ายลงเรือ
        w = Weight.objects.filter(site__store = 3, bws__company__code__in = company_in, bws__weight_type = mode, stone_type = stoneType, date__range=(start_date, end_date)).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
    elif type == 6:#รวม เครื่องขาย ปลายทางจ่ายออก
        w = Weight.objects.filter(site__store__in = [2,3], bws__company__code__in = company_in, bws__weight_type = mode, stone_type = stoneType, date__range=(start_date, end_date)).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')    
    return  float(w)

def getNumListStonePortChart(request, mode, stone_list_id, type, company_in):
    #sell
    list_sum_stone = []
    for stone_id in stone_list_id:
        list_sum_stone.append(getPortSumByStone(request, mode, stone_id, type, company_in))

    list_sum_stone.append(getPortSumOther(request, mode, stone_list_id, type, company_in))
    return list_sum_stone

def getPortSumOther(request, mode, list_sum_stone, type, company_in):
    start_date = request.session['db_start_date']
    end_date = request.session['db_end_date']

    start_year = datetime.strptime(start_date, '%Y-%m-%d').year

    query_filters = Q()
    ss_query_filters = Q()
    for item_number_prefix in list_sum_stone:
        query_filters |= Q(stone_type = item_number_prefix)
        ss_query_filters |= Q(stone = item_number_prefix)

    #type 1 = sell, 2 = stock, 3 = produce, 4 = purchase 
    if type == 1:#เครื่องขาย ปลายทาง รับเข้า
        w = Weight.objects.filter(site__store = 1, bws__company__code__in = company_in, bws__weight_type = mode, date__range=(start_date, end_date)).exclude(query_filters).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
    elif type == 2:
        w = Decimal('0.0')
        try:
            qr = StockStone.objects.filter(stk__company__code__in = company_in, stk__created__range=(start_date, end_date)).exclude(ss_query_filters).values('stk__created', 'stone', 'total').order_by('-stk__created').distinct() #ดึงข้อมูลสต็อกที่เหลือจากวันที่คีย์ล่าสุด ระหว่าง start_date, end_date และรวมกัน
        except StockStone.DoesNotExist:
            qr = None

        if qr:
            qr_list = list(qr)

            # Group by 'stone', keeping the most recent 'stk__created'
            filtered_results = [
                max(group, key=itemgetter('stk__created'))
                for _, group in groupby(sorted(qr_list, key=itemgetter('stone')), key=itemgetter('stone'))
            ]
            w = sum(item['total'] for item in filtered_results)

    elif type == 4:#เครื่องขาย ปลายทาง จ่ายภายนอก
        w = Weight.objects.filter(site__store = 2, bws__company__code__in = company_in, bws__weight_type = mode, date__range=(start_date, end_date)).exclude(query_filters).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
    elif type == 5:#เครื่องขาย ปลายทาง จ่ายภายลงเรือ
        w = Weight.objects.filter(site__store = 3, bws__company__code__in = company_in, bws__weight_type = mode, date__range=(start_date, end_date)).exclude(query_filters).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
    elif type == 6:#รวม เครื่องขาย ปลายทางจ่ายออก
        w = Weight.objects.filter(site__store__in = [2,3], bws__company__code__in = company_in, bws__weight_type = mode, date__range=(start_date, end_date)).exclude(query_filters).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
    return  float(w)

# Create your views here.
@login_required(login_url='login')
def index(request):
    #loade_st = time.time()  # Start loade time

    try:
        #active : active คือแท็ปบริษัท active
        active = request.session['company_code']
        company_in = findCompanyIn(request)

        comp = BaseCompany.objects.get(code = active)

        start_date = request.session['db_start_date']
        end_date = request.session['db_end_date']
        now_date = datetime.strptime(start_date, "%Y-%m-%d")

        start_day = datetime.strptime(start_date, "%Y-%m-%d")
        end_day = datetime.strptime(end_date, "%Y-%m-%d")
    except:
        return redirect('logout')

    # today date
    current_date = datetime.now()
    previous_day = current_date - timedelta(days=1)

    s_comp = BaseSite.objects.filter(s_comp__code = active).values('base_site_id', 'base_site_name').order_by('base_site_id')
    s_comp_id = BaseSite.objects.filter(s_comp__code = active).values_list('base_site_id').order_by('base_site_id')
    s_comp_name = BaseSite.objects.filter(s_comp__code = active).values('base_site_name').order_by('base_site_id')

    ####################################
    ########### chart mill #############
    ####################################

    #สร้าง list ระหว่าง start_date และ end_date
    list_date_between = pd.date_range(start_date, end_date).tolist()
    list_date = [date.strftime("%Y-%m-%d") for date in list_date_between]

    # Define lists to store cumulative totals and goal percentages for each mill
    list_goal_mills = [[] for _ in range(len(s_comp_id))]
    cumulative_totals = [0] * len(s_comp_id)

    # Fetch sum goals for each mill
    sum_goals = {}
    for mill_id in s_comp_id:
        sum_goals[mill_id] = ProductionGoal.objects.filter(
            date__year=now_date.year,
            date__month=now_date.month,
            site=mill_id
        ).aggregate(s=Sum('accumulated_goal'))['s']

    # Fetch weights for each mill within the date range
    weights = {}
    for i, mill_id in enumerate(s_comp_id):
        weights[mill_id] = Weight.objects.filter(
            date__range=(start_date, end_date),
            site=mill_id
        ).values('date').annotate(
            cumulative_total=Sum('weight_total')
        ).order_by('date')

    # Iterate through list_date
    for date in list_date:
        for i, mill_id in enumerate(s_comp_id):
            # Iterate through weights for the current mill and update cumulative total
            for w in weights[mill_id]:
                if str(date) == str(w['date']):
                    cumulative_totals[i] += w['cumulative_total']
            # Append goal percentage to the corresponding list
            list_goal_mills[i].append(calculatePersent(cumulative_totals[i], sum_goals[mill_id]))

    list_goal_mill = []
    for i, mill in enumerate(s_comp):
        list_goal_mill.append((mill['base_site_name'], list_goal_mills[i]))

    ####################################
    ##chart loss weight เวลาที่เสีย (ผลิต)##
    ####################################
    actual_working_time = {}
    total_loss_time = {}
    persent_loss_weight = {}
    list_persent_loss_weight = []

    # Iterate over the mill site IDs
    for site_id in s_comp_id:
        # Filter Production objects for the current mill site and calculate actual working time
        actual_working_time[site_id] = Production.objects.filter(
            created__year=f'{now_date.year}',
            created__month=f'{now_date.month}',
            site=site_id
        ).annotate(
            working_time=ExpressionWrapper(F('actual_time') - F('total_loss_time'), output_field=models.DurationField())
        ).aggregate(total_working_time=Sum('working_time'))['total_working_time']

        # Filter Production objects for the current mill site and calculate total loss time
        total_loss_time[site_id] = Production.objects.filter(
            created__range=(start_date, end_date),
            site=site_id
        ).aggregate(s=Sum('total_loss_time'))["s"]

        # Calculate percentage loss weight for the current mill site
        persent_loss_weight[site_id] = calculatePersent(
            total_loss_time[site_id] if total_loss_time[site_id] else None,
            actual_working_time[site_id]
        )
        list_persent_loss_weight.append(persent_loss_weight[site_id])

    actual_working_time_all = Production.objects.filter(company__code = active, created__year = f'{now_date.year}' , created__month = f'{now_date.month}').annotate(working_time = ExpressionWrapper(F('actual_time') - F('total_loss_time'), output_field= models.DurationField())).aggregate(total_working_time=Sum('working_time'))['total_working_time']
    total_loss_time_all = Production.objects.filter(company__code = active, created__range = (start_date, end_date)).aggregate(s=Sum('total_loss_time'))["s"]
    persent_loss_weight_all = calculatePersent(total_loss_time_all if total_loss_time_all else None, actual_working_time_all)

    # Add list
    list_persent_loss_weight.append(persent_loss_weight_all)

    '''
    loade_en = time.time()# End measuring time
    loade_t = int((loade_en - loade_st) * 1000 - 100)
    request.session['loade_page'] = 0 if loade_t < 0 else loade_t# Convert to milliseconds    
    '''
    if not comp.biz or comp.biz.id == 1:#ธุรกิจเหมือง
        ####################################
        ######## data weight stock #########
        ####################################
        # เปลี่ยนเป็นเลือกระหว่างวันที่ 2024-04-10 -> data_sum_produc_all = Weight.objects.filter(bws__company__code__in = company_in, site__in = s_comp_id, date = previous_day, bws__weight_type = 2).aggregate(s=Sum("weight_total"))["s"]
        data_sum_produc_all = Weight.objects.filter(bws__company__code__in = company_in, site__in = s_comp_id, date__range=(start_date, end_date), bws__weight_type = 2).aggregate(s=Sum("weight_total"))["s"]

        data_sum_produc = []
        data_sum_produc.append(('Total', data_sum_produc_all))

        for site in s_comp:
            aggregated_value = Weight.objects.filter(
                bws__company__code__in=company_in,
                site=site['base_site_id'],
                date__range=(start_date, end_date),
                bws__weight_type=2
            ).aggregate(s=Sum("weight_total"))["s"]

            if aggregated_value:
                # Append a tuple (site_id, aggregated_value) to the list
                data_sum_produc.append((site['base_site_name'], aggregated_value))

        ####################################
        ########### chart stone ############
        ####################################
        #ดึงชนิดหินตามบริษัท ถ้าไม่มีดึงตามนี้ 'หิน 3/4', 'หิน 40/80', 'หินฝุ่น', 'หินคลุก A', 'หินคลุก B', 'อื่นๆ',
        try:
            set_stone = SetCompStone.objects.filter(comp__code = active)
            result_list = list(set_stone.values_list('stone', flat=True))

            text_value = result_list[0]
            stone_list = text_value.replace("'", "").split(',')
        except:
            stone_list = ['01ST', '07ST', '09ST', '10ST', '16ST']

        #หาชื่อหินจาก stone_list
        stone_name_list = list(BaseStoneType.objects.filter(base_stone_type_id__in = stone_list).values_list('base_stone_type_name', flat=True).order_by('base_stone_type_id'))
        stone_name_list.append('อื่นๆ')
        
        #หาจำนวนตันจาก stone_list
        sell_mill_list = getNumListStoneWeightChart(request, 1, stone_list, 1, company_in)
        sell_stock_list = getNumListStoneWeightChart(request, 1, stone_list, 4, company_in)
        sell_purchase_list = getNumListStoneWeightChart(request, 1, stone_list, 5, company_in)
        total_sell_list = getNumListStoneWeightChart(request, 1, stone_list, 6, company_in)

        stock_list = getNumListStoneWeightChart(request, 2, stone_list, 2, company_in)
        produce_list = getNumListStoneWeightChart(request, 2, stone_list, 3, company_in)

        sell_total = sum(sum(lst) for lst in [sell_mill_list, sell_stock_list, sell_purchase_list]) #รวม group sell
        produce_total = sum(produce_list)#รวม group produce

        context = { 
                    'previous_day':previous_day,
                    'start_day':start_day,
                    'end_day':end_day,
                    'actual_working_time_all':actual_working_time_all,
                    'sell_mill_list':sell_mill_list,
                    'sell_stock_list': sell_stock_list,
                    'sell_purchase_list': sell_purchase_list,
                    'total_sell_list': total_sell_list,
                    'stock_list':stock_list,
                    'produce_list':produce_list,
                    'data_sum_produc_all':data_sum_produc_all,
                    'data_sum_produc':data_sum_produc,
                    'sell_total':sell_total,
                    'produce_total':produce_total,
                    'list_date': list_date,
                    'list_goal_mill' : list_goal_mill,
                    'list_persent_loss_weight':list_persent_loss_weight,
                    'stone_name_list':stone_name_list,
                    'dashboard_page':'active',
                    active :"active",
        }
        return render(request, "index.html",context)
    
    elif comp.biz.id == 2:#ธุรกิจท่าเรือ
        store = BaseSiteStore.objects.filter(id__in = [1,2,3]).values('id', 'name').order_by('id')
        store_id = BaseSiteStore.objects.filter(id__in = [1,2,3]).values_list('id').order_by('id')

        ####################################
        ######## data weight stock #########
        ####################################
        # เปลี่ยนเป็นเลือกระหว่างวันที่ 2024-04-10 -> data_sum_produc_all = Weight.objects.filter(bws__company__code__in = company_in, site__in = s_comp_id, date = previous_day, bws__weight_type = 2).aggregate(s=Sum("weight_total"))["s"]
        sum_line_long = Weight.objects.filter(bws__company__code__in = company_in, mill__isnull = True, line_type = "สายยาว", date__range=(start_date, end_date), bws__weight_type = 1).aggregate(s=Sum("weight_total"))["s"]  or Decimal(0) # รับเข้า เป็นสายยาวทั้งหมด
        sum_cus_ot = Weight.objects.filter(Q(site__store = 2) | Q(site__isnull = True), mill__isnull = False, line_type = "สายยาว", bws__company__code__in = company_in, date__range=(start_date, end_date), bws__weight_type = 1).aggregate(s=Sum("weight_total"))["s"] or Decimal(0)#ขายภายนอก
        sum_ship = Weight.objects.filter(bws__company__code__in = company_in, site__store = 3, date__range=(start_date, end_date), bws__weight_type = 1).aggregate(s=Sum("weight_total"))["s"] or Decimal(0) #ขายลงเรือ

        data_sum_produc_all = sum_line_long + sum_cus_ot + sum_ship

        data_sum_produc = []
        data_sum_produc.append(('ยอดสะสมประจำเดือน', data_sum_produc_all))

        for i, st in enumerate(store):
            if i == 0: #16-07-2025 กราฟ รับเข้า เปลี่ยนการดึงข้อมูล เป็นสายยาวทั้งหมด
                aggregated_value = Weight.objects.filter(
                    mill__isnull = True,
                    bws__company__code__in=company_in,
                    line_type = "สายยาว",
                    date__range=(start_date, end_date),
                    bws__weight_type=1
                ).aggregate(s=Sum("weight_total"))["s"]
            elif i == 1: #31-07-2025 ขายภายนอก
                aggregated_value = Weight.objects.filter(
                    Q(site__store = st['id']) | Q(site__isnull = True),
                    mill__isnull = False,
                    bws__company__code__in=company_in,
                    line_type = "สายยาว",
                    date__range=(start_date, end_date),
                    bws__weight_type=1
                ).aggregate(s=Sum("weight_total"))["s"]
            elif i == 2:# ขายลงเรือ
                aggregated_value = Weight.objects.filter(
                    bws__company__code__in=company_in,
                    site__store = st['id'],
                    date__range=(start_date, end_date),
                    bws__weight_type=1
                ).aggregate(s=Sum("weight_total"))["s"]

            # Append a tuple (site_id, aggregated_value) to the list
            data_sum_produc.append((st['name'], aggregated_value))

        ####################################
        ########### chart stone ############
        ####################################
        #ดึงชนิดหินตามบริษัท ถ้าไม่มีดึงตามนี้ 'หิน 3/4', 'หิน 40/80', 'หินฝุ่น', 'หินคลุก A', 'หินคลุก B', 'อื่นๆ',
        try:
            set_stone = SetCompStone.objects.filter(comp__code = active)
            result_list = list(set_stone.values_list('stone', flat=True))

            text_value = result_list[0]
            stone_list = text_value.replace("'", "").split(',')
        except:
            stone_list = ['01ST', '07ST', '09ST', '10ST', '16ST']
        
        #หาชื่อหินจาก stone_list
        stone_name_list = list(BaseStoneType.objects.filter(base_stone_type_id__in = stone_list).values_list('base_stone_type_name', flat=True).order_by('base_stone_type_id'))
        stone_name_list.append('อื่นๆ')

        list_store_sites = [[] for _ in range(len(store_id))]
        cumulative_totals = [0] * len(store_id)

        # เตรียม ship name ตามวันที่
        ship_name_dict = {}
        ship_name_summary_by_date = {}

        if (3,) in store_id:
            ship_data = Weight.objects.filter(
                date__range=(start_date, end_date),
                site__store=3
            ).values('date', 'site__base_site_name').order_by('date').distinct()

            for item in ship_data:
                date_str = str(item['date'])
                name = item['site__base_site_name']
                ship_name_dict[date_str] = name  # ใช้ชื่อเรือเดียว
                if date_str in ship_name_summary_by_date:
                    if name not in ship_name_summary_by_date[date_str]:
                        ship_name_summary_by_date[date_str].append(name)
                else:
                    ship_name_summary_by_date[date_str] = [name]

        # สร้าง weights ตามแต่ละ store
        weights = {}
        for i, st_id in enumerate(store_id):
            if i == 0:  # สำหรับสายยาว
                weights[st_id] = Weight.objects.filter(
                    mill__isnull = True,
                    date__range=(start_date, end_date),
                    line_type="สายยาว"
                ).values('date').annotate(
                    cumulative_total=Sum('weight_total')
                ).order_by('date')
            elif i == 1:  # สำหรับแต่ละ store เช่น ขายภายนอก
                weights[st_id] = Weight.objects.filter(
                    Q(site__store = st_id) | Q(site__isnull = True),
                    mill__isnull = False,
                    date__range=(start_date, end_date),
                    line_type="สายยาว"
                ).values('date').annotate(
                    cumulative_total=Sum('weight_total')
                ).order_by('date')
            elif i == 2:  # สำหรับแต่ละ store เช่น ขายลงเรือ
                weights[st_id] = Weight.objects.filter(
                    date__range=(start_date, end_date),
                    site__store=st_id
                ).values('date').annotate(
                    cumulative_total=Sum('weight_total')
                ).order_by('date')

        # เตรียมข้อมูลสำหรับกราฟ
        tooltip_labels = [[] for _ in range(len(store_id))]

        for date in list_date:
            date_str = str(date)
            for i, st_id in enumerate(store_id):
                found = False
                for w in weights[st_id]:
                    if date_str == str(w['date']):
                        weight = float(w['cumulative_total'])
                        list_store_sites[i].append(weight)
                        if st_id == (3,):
                            ship_list = ship_name_summary_by_date.get(date_str, [])
                            label = f" {weight:.2f} ตัน / {', '.join(ship_list)}"
                        else:
                            label = f"{weight:.2f} ตัน"
                        tooltip_labels[i].append(label)
                        found = True
                        break
                if not found:
                    list_store_sites[i].append(0)
                    tooltip_labels[i].append("0 ตัน")

        list_store_site = []
        for i, store in enumerate(store):
            list_store_site.append((store['name'], list_store_sites[i]))

        ################## stock ########################
        stock_list = getNumListStoneWeightChart(request, 1, stone_list, 10, company_in)

        ################## stock port ###################
        ps = PortStock.objects.filter(company__code__in = company_in, created__range=(start_date, end_date)).values_list('id', flat=True).order_by('-created').first()
        qs_port_stock = PortStockStoneItem.objects.filter(pss__ps = ps).values('cus__customer_name', 'pss__stone__base_stone_type_name', 'total').order_by('-pss__ps__created')

        data = defaultdict(lambda: defaultdict(Decimal))

        port_stone_types = set()

        for row in qs_port_stock:
            customer = row['cus__customer_name']
            stone = row['pss__stone__base_stone_type_name']
            total = row['total']
            data[customer][stone] += total
            port_stone_types.add(stone)

        port_stone_types = sorted(port_stone_types)  # Sort columns

        # Optionally calculate row totals
        for customer in data:
            for stone in port_stone_types:
                if stone not in data[customer]:
                    data[customer][stone] = Decimal('0.00')
            data[customer]['__total__'] = sum(data[customer][stype] for stype in port_stone_types)

        port_stock_list = recursive_defaultdict_to_dict(data)

        context = { 
                    'previous_day':previous_day,
                    'start_day':start_day,
                    'end_day':end_day,
                    'actual_working_time_all':actual_working_time_all,
                    'stock_list':stock_list,
                    'port_stone_types':port_stone_types,
                    'port_stock_list': port_stock_list,
                    'data_sum_produc_all':data_sum_produc_all,
                    'data_sum_produc':data_sum_produc,
                    'list_date': list_date,
                    'list_goal_mill' : list_goal_mill,
                    'list_store_site' : list_store_site,
                    'tooltip_labels': tooltip_labels,
                    'list_persent_loss_weight':list_persent_loss_weight,
                    'stone_name_list':stone_name_list,
                    'dashboard_page':'active',
                    active :"active",
        }
        return render(request, "ndIndex.html",context)
    
    elif comp.biz.id == 3:#ธุรกิจขนส่ง
        chart_data = getChartTransport(start_date, end_date, None)
        slc_chart_data = getChartTransport(start_date, end_date, 'SLC')
        slt_chart_data = getChartTransport(start_date, end_date, 'SLT')
        ctm_chart_data = getChartTransport(start_date, end_date, 'CTM')
        uni_chart_data = getChartTransport(start_date, end_date, 'UNI')
        kt_chart_data = getChartTransport(start_date, end_date, 'KT')
        stps_chart_data = getChartTransport(start_date, end_date, 'STPS')
        tym_chart_data = getChartTransport(start_date, end_date, 'TYM')

        comp_list = ['SLC', 'SLT', 'CTM', 'UNI', 'KT', 'STPS', 'TYM']
        ####################################
        ######## data weight stock #########
        ####################################
        # เปลี่ยนเป็นเลือกระหว่างวันที่ 2024-04-10 -> data_sum_produc_all = Weight.objects.filter(bws__company__code__in = company_in, site__in = s_comp_id, date = previous_day, bws__weight_type = 2).aggregate(s=Sum("weight_total"))["s"]
        data_sum_produc_all = Weight.objects.filter(bws__company__code__in = comp_list, date__range=(start_date, end_date), bws__weight_type = 1, carry_type_name = "ส่งให้",).aggregate(sum_weight=Sum("weight_total"), count_num = Count("weight_id"))

        data_sum_produc = []
        data_sum_produc.append(('รวมยอดขนส่ง', data_sum_produc_all['sum_weight'], data_sum_produc_all['count_num']))

        for comp in comp_list:
            aggregated_value = Weight.objects.filter(
                bws__company__code = comp,
                date__range=(start_date, end_date),
                bws__weight_type=1,
                carry_type_name = "ส่งให้",
            ).aggregate(sum_weight=Sum("weight_total"), count_num = Count("weight_id"))

            b_com = BaseCompany.objects.get(code = comp)
            # Append a tuple (site_id, aggregated_value) to the list
            data_sum_produc.append((b_com.name, aggregated_value['sum_weight'], aggregated_value['count_num']))

        context = { 
                    'chart_data': json.dumps(chart_data, ensure_ascii=False),
                    'slc_chart_data': json.dumps(slc_chart_data, ensure_ascii=False),
                    'slt_chart_data': json.dumps(slt_chart_data, ensure_ascii=False),
                    'ctm_chart_data': json.dumps(ctm_chart_data, ensure_ascii=False),
                    'uni_chart_data': json.dumps(uni_chart_data, ensure_ascii=False),
                    'kt_chart_data': json.dumps(kt_chart_data, ensure_ascii=False),
                    'stps_chart_data': json.dumps(stps_chart_data, ensure_ascii=False),
                    'tym_chart_data': json.dumps(tym_chart_data, ensure_ascii=False),
                    'previous_day':previous_day,
                    'start_day':start_day,
                    'end_day':end_day,
                    'data_sum_produc_all':data_sum_produc_all,
                    'data_sum_produc':data_sum_produc,
                    'dashboard_page':'active',
                    active :"active",
        }
        return render(request, "thIndex.html",context)

def getChartTransport(start_date, end_date, company):
    # Step 1: Get top 10 car_team names ordered by total_weight
    base_filter = {
        "car_team__isnull": False,
        "date__range": (start_date, end_date),
        "carry_type_name": "ส่งให้"
    }
    if company:
        base_filter["bws__company__code"] = company

    raw_top_teams = Weight.objects.filter(**base_filter) \
        .values("car_team__car_team_name") \
        .annotate(total_weight=Sum("weight_total")) \
        .order_by("-total_weight")[:10]

    top_car_teams = [item["car_team__car_team_name"] for item in raw_top_teams]

    if not top_car_teams:
        return {"categories": [], "series": []}

    # Step 2: Get summarized data for those top 10 teams
    detail_filter = {
        "stone_type__isnull": False,
        "car_team__car_team_name__in": top_car_teams,
        "date__range": (start_date, end_date),
        "carry_type_name": "ส่งให้"
    }
    if company:
        detail_filter["bws__company__code"] = company

    queryset = Weight.objects.filter(**detail_filter) \
        .values("car_team__car_team_name", "stone_type__base_stone_type_name") \
        .annotate(
            sum_weight=Sum("weight_total"),
            num_count=Count("weight_id")
        )

    # Step 3: Prepare chart data
    car_teams = top_car_teams  # already ordered
    car_team_index = {team: idx for idx, team in enumerate(car_teams)}
    stone_types = sorted(set(item["stone_type__base_stone_type_name"] for item in queryset))
    stone_data = {stype: [0] * len(car_teams) for stype in stone_types}

    for item in queryset:
        team = item["car_team__car_team_name"]
        stype = item["stone_type__base_stone_type_name"]
        weight = float(item["sum_weight"])
        idx = car_team_index[team]
        stone_data[stype][idx] = weight

    chart_data = {
        "categories": car_teams,
        "series": [
            {
                "name": stype,
                "data": [
                    {
                        "x": car_teams[idx],
                        "y": stone_data[stype][idx],
                        "num_count": next((
                            item["num_count"]
                            for item in queryset
                            if item["car_team__car_team_name"] == car_teams[idx] and item["stone_type__base_stone_type_name"] == stype
                        ), 0)
                    }
                    for idx in range(len(car_teams))
                ]
            }
            for stype in stone_types
        ]
    }

    return chart_data

# Convert to plain dict
def recursive_defaultdict_to_dict(d):
    if isinstance(d, defaultdict):
        d = {k: recursive_defaultdict_to_dict(v) for k, v in d.items()}
    return d

def calculatePersent(num, num_all):
    persent = 0.0
    if num_all and num:
        persent = (num/num_all)*100
    return round(persent)

def is_scale(user):
    return user.groups.filter(name='scale').exists()

def is_account(user):
    return user.groups.filter(name='account').exists()

def is_edit_weight(user):
    return user.groups.filter(name='edit_weight').exists()

def is_edit_setting(user):
    return user.groups.filter(name='edit_setting').exists()

def is_view_weight(user):
    return user.groups.filter(name='view_weight').exists()

def is_edit_base_id(user):
    return user.groups.filter(name='edit_base_id').exists()

def is_edit_stock(user):
    return user.groups.filter(name='edit_stock').exists()

def is_approve_weight(user):
    return user.groups.filter(name='approve_weight').exists()

def loginPage(request):
    if request.method == 'POST':
        form = AuthenticationForm(data = request.POST)
        if form.is_valid():
            username = request.POST['username']
            password = request.POST['password']
            user = authenticate(username=username,password=password)
            #ถ้าล็อกอินสำเร็จไปหน้า home else ให้ไปสมัครใหม่
            if user is not None:
                login(request, user)
                #CPT*เลือกตามบริษัท 
                try:
                    user_profile = UserProfile.objects.get(user = request.user.id)
                    company = BaseCompany.objects.filter(userprofile = user_profile).first()
                except:
                    company = BaseCompany.objects.get(id = 1)
                request.session['company_code'] = company.code
                request.session['company'] = company.name

                #set session date in dashbord
                current_date_time = datetime.today()
                previous_date_time = current_date_time - timedelta(days=1)

                start_date = datetime.strptime(startDateInMonth(str(previous_date_time.strftime('%Y-%m-%d'))), "%Y-%m-%d")
                end_date = datetime.strptime(endDateInMonth(str(previous_date_time.strftime('%Y-%m-%d'))), "%Y-%m-%d")

                request.session['db_start_date'] = f'{start_date.strftime("%Y-%m-%d")}'
                request.session['db_end_date'] = f'{end_date.strftime("%Y-%m-%d")}'

                return redirect('home')
            else:
                return redirect('signUp')
    else:
        form = AuthenticationForm()
        #CPT*เลือกตามบริษัท 
        company = BaseCompany.objects.first()
        request.session['company_code'] = company.code
        request.session['company'] = company.name

    return render(request, 'account/login.html', {'form':form,})

def logoutUser(request):
    logout(request)
    return redirect('login')

@login_required(login_url='login')
def weightTable(request):
    ''' old บริษัทเดียว
    active = None
    if is_scale(request.user):
        us = UserScale.objects.filter(user = request.user).values_list('scale_id')
        data = Weight.objects.filter(scale_id__in = us).order_by('-date','weight_id')
    elif request.user.is_superuser or is_view_weight(request.user) or is_edit_weight(request.user) or is_account(request.user):
        data = Weight.objects.all().order_by('-date','weight_id')    
    '''

    #active : active คือแท็ปบริษัท active
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')


    #CPT*เลือกตามบริษัท
    '''
    if is_scale(request.user):
        us = UserScale.objects.filter(user = request.user).values_list('scale_id')
        data = Weight.objects.filter(scale_id__in = us).order_by('-date','weight_id')
    elif request.user.is_superuser or is_view_weight(request.user) or is_edit_weight(request.user) or is_account(request.user):
        data = Weight.objects.filter(bws__company__code__in = company_in).order_by('-date','weight_id')    
    '''
    data = Weight.objects.filter(bws__company__code__in = company_in).order_by('-date','weight_id')

    #กรองข้อมูล
    myFilter = WeightFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    weight = p.get_page(page)

    context = {'weight':weight,'filter':myFilter, 'weightTable_page':'active', 'is_view_weight' : is_view_weight(request.user), 'is_approve_weight' : is_approve_weight(request.user), 'is_scale' : is_scale(request.user), 'is_account' :is_account(request.user), active :"active",}
    return render(request, "weight/weightTable.html",context)

@login_required(login_url='login')
def approveWeight(request):
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')

    # Get distinct weight dates
    data = Weight.objects.filter(
        bws__company__code__in=company_in
    ).values_list('date', flat=True).order_by('-date').distinct()
    
    ap_data = ApproveWeight.objects.filter(company__code = active)

    myFilter = WeightFilter(request.GET, queryset=data)
    data = myFilter.qs

    # Pagination
    p = Paginator(data, 5)
    page = request.GET.get('page')
    weight = p.get_page(page)

    if request.method == 'POST':
        checkbox_data = request.POST.get('checkboxData')
        if checkbox_data:
            checkbox_data = json.loads(checkbox_data)

            # Prepare data
            date_list = [item['date'] for item in checkbox_data]
            approve_map = {item['date']: item['isChecked'] for item in checkbox_data}

            # Fetch existing approvals
            existing_apws = ApproveWeight.objects.filter(
                company__code=active, date__in=date_list
            )
            existing_apws_map = {str(apw.date): apw for apw in existing_apws}

            company = BaseCompany.objects.get(code=active)
            to_create = []
            to_update = []

            for date_str in date_list:
                is_checked = approve_map[date_str]

                if date_str in existing_apws_map:
                    apw = existing_apws_map[date_str]
                    apw.is_approve = is_checked
                    apw.update = datetime.now()
                    to_update.append(apw)
                else:
                    apw = ApproveWeight(
                        company=company,
                        date=date_str,
                        is_approve=is_checked,
                        update=datetime.now()
                    )
                    to_create.append(apw)

            # Create new entries
            if to_create:
                ApproveWeight.objects.bulk_create(to_create)

            # Update existing entries
            if to_update:
                ApproveWeight.objects.bulk_update(to_update, ['is_approve', 'update'])

            # Re-fetch all relevant ApproveWeight entries (now all are saved with pk)
            all_apws = ApproveWeight.objects.filter(
                company__code=active, date__in=date_list
            )
            full_apw_map = {str(apw.date): apw.is_approve for apw in all_apws}

            # Update related Weight.apw foreign keys
            weights = Weight.objects.filter(
                bws__company__code=active, date__in=date_list
            )

            for w in weights:
                is_apw = full_apw_map.get(str(w.date))
                w.is_apw = is_apw

            Weight.objects.bulk_update(weights, ['is_apw'])

        return redirect('weightTable')

    context = {'weight':weight, 'ap_data': ap_data, 'filter':myFilter, 'weightTable_page':'active', active :"active",}
    return render(request, "weight/approveWeight.html",context)

@login_required(login_url='login')
def editWeight(request, mode, weight_id):
    #loade_st = time.time()  # Start loade time

    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    weight_data = get_object_or_404(Weight, pk=weight_id)

    if company.biz.id == 1 and mode == 1: #ธุรกิจเหมือง
        template_name = "weight/editWeightSell.html"
        tmp_form_post = WeightForm(request.POST, request.FILES, instance=weight_data)
        tmp_form = WeightForm(instance=weight_data)
    elif company.biz.id == 1 and mode == 2: #ธุรกิจเหมือง
        template_name = "weight/editWeightStock.html"
        tmp_form_post = WeightStockForm(request.POST, request.FILES, instance=weight_data)
        tmp_form = WeightStockForm(instance=weight_data)
    elif company.biz.id == 2 and mode == 1: #ธุรกิจท่าเรือ
        template_name = "weight/editWeightPort.html"
        tmp_form_post = WeightPortForm(request.POST, request.FILES, instance=weight_data)
        tmp_form = WeightPortForm(instance=weight_data)

    #ถ้ารหัสกับชื่อ local และ center ไม่ตรงกันให้เลือกจากชื่อ 03/03/2025
    is_not_match_mill = False
    if weight_data.mill:
        mill = BaseMill.objects.get(mill_id = weight_data.mill.mill_id)
        center_mill = mill.mill_id + mill.mill_name #รหัสและชื่อบนหน้าเว็บ
        local_mill = weight_data.mill.mill_id + weight_data.mill_name #รหัสและชื่อจากตาชั่ง    
        if local_mill != center_mill:
            is_not_match_mill = True

    if request.method == 'POST':
        form = tmp_form_post
        if form.is_valid():
            #เก็บน้ำหนักและปลายทางก่อนแก้
            try:
                original_weight = Weight.objects.get(pk=form.instance.pk)
                original_weight_total = original_weight.weight_total
                if original_weight.site:
                    original_weight_site = original_weight.site.base_site_id
                if original_weight.customer:
                    original_weight_cus = original_weight.customer.customer_id
                if original_weight.stone_type:
                    original_weight_stone = original_weight.stone_type.base_stone_type_id
            except Weight.DoesNotExist:
                original_weight_total = None
                original_weight_site = None
                original_weight_cus = None
                original_weight_stone = None
        
            # log history เก็บข้อมูลก่อนแก้
            weight_form = form.save()

            weight_history = WeightHistory.objects.filter(weight_id = weight_form.pk).order_by('-update')[0]
            weight_history.user_update = request.user
            weight_history.save()

            #เครื่องขาย
            if mode == 1:
                if original_weight_total is not None and original_weight_total != weight_form.weight_total:
                    if weight_form.oil_content:
                        updateGasPrice(weight_form.bws.company.id, weight_form.date)
                        updateOilCostAndSell(weight_form.pk, weight_form.bws.company.id, weight_form.date)
            #เครื่องผลิต
            if mode == 2:#กรณีแก้ไขรายการชั่งผลิต update total StoneEstimateItem ด้วย และ capacity_per_hour
                if original_weight_total is not None and original_weight_total != weight_form.weight_total or original_weight_site is not None and original_weight_site != weight_form.site.base_site_id:
                    # update new site
                    updatePassScaleEstimate(weight_form.bws.company.id, weight_form.date, weight_form.site.base_site_id)
                    updateProductionCapacity(weight_form.bws.company.id, weight_form.date, weight_form.site.base_site_id)
                    # update old site
                    updatePassScaleEstimate(weight_form.bws.company.id, weight_form.date, original_weight_site)
                    updateProductionCapacity(weight_form.bws.company.id, weight_form.date, original_weight_site)#กรณีแก้ไขรายการชั่งผลิต update total StoneEstimateItem ด้วย และ capacity_per_hour
                    if  weight_form.stone_type:
                        updateProdStockStoneItem(weight_form.bws.company.id, weight_form.date)
            #ธุรกิจเหมือง
            if mode == 1  and company.biz.id == 1:#กรณีแก้ไขรายการชั่งขาย คำนวนราคาใหม่ด้วย
                if weight_form.stone_type:
                    updateSellStockStoneItem(weight_form.pk)
            #ธุรกิจท่าเรือ
            if mode == 1 and company.biz.id == 2:
                if original_weight_total is not None and original_weight_total != weight_form.weight_total or original_weight_cus is not None and original_weight_cus != weight_form.customer.customer_id or original_weight_stone is not None and original_weight_stone != weight_form.stone_type.base_stone_type_id:
                    updatePortStockStoneItem(weight_form.bws.company.id, weight_form.date, original_weight_cus, original_weight_stone)
                    updatePortStockStoneItem(weight_form.bws.company.id, weight_form.date, weight_form.customer.customer_id, weight_form.stone_type.base_stone_type_id)

            return redirect('weightTable')
    else:
        form = tmp_form

    '''
    loade_en = time.time()# End measuring time
    loade_t = int((loade_en - loade_st) * 1000 - 100)
    request.session['loade_page'] = 0 if loade_t < 0 else loade_t# Convert to milliseconds
    '''

    context = {'weightTable_page': 'active', 'form': form, 'weight': weight_data, 'is_edit_weight': is_edit_weight(request.user) , 'is_not_match_mill': is_not_match_mill, active :"active", 'disabledTab' : 'disabled'}
    return render(request, template_name, context)

def updateSellStockStoneItem(weight_id):
    tmp_ssn_id = None
    tmp_ss_item_id = None
    weight = Weight.objects.get(weight_id = weight_id)
    company = weight.bws.company
    date = weight.date
    stone = weight.stone_type.base_stone_type_id

    #รวมจำนวนขายทั้งหมด ไม่รวมใช้ภายในและอนุเคาราะห์
    try:
        ss_sell = StockStoneItem.objects.filter(source__id = 3, ssn__stk__company = company, ssn__stk__created = date, ssn__stone = stone).last()
        if ss_sell:
            #ขาย
            sell = Weight.objects.filter(~Q(site = '200PL') & ~Q(site = '300PL'), bws__company = company, bws__weight_type = 1, stone_type = stone, date = date).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
            tmp_ssn_id = ss_sell.ssn.id
            tmp_ss_item_id = ss_sell.id
            ss_sell.quantity = sell #ดึงจำนวนรวมขายมา
            ss_sell.save()
    except StockStoneItem.DoesNotExist:
        pass

    #อนุเคราะห์ (ปลายทาง 300PL)
    if weight.site and weight.site.base_site_id == '300PL':
        try:
            ss_aid = StockStoneItem.objects.filter(source__id = 10, ssn__stk__company = company, ssn__stk__created = date, ssn__stone = stone).last()
            if ss_aid:
                #อนุเคราะห์ (ปลายทาง 300PL)
                aid = Weight.objects.filter(bws__company = company, bws__weight_type = 1, stone_type = stone, date = date, site = '300PL').aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
                tmp_ssn_id = ss_aid.ssn.id
                tmp_ss_item_id = ss_aid.id
                ss_aid.quantity = aid #ดึงจำนวนรวมอนุเคราะห์
                ss_aid.save()
        except StockStoneItem.DoesNotExist:
            pass
    
    if tmp_ssn_id:
        #update total stock ของชนิดหินนี้
        ssn = StockStone.objects.get(id = tmp_ssn_id)
        ssn.total = calculateTotalStock(tmp_ssn_id)
        ssn.save()

    if tmp_ss_item_id:
        updateTotalStockInMonth(tmp_ss_item_id)#คำนวนยอดยกมา จากวันก่อนและคำนวน total stock ใหม่

def updatePortStockStoneItem(company, date, cus, stone):
    tmp_pss_id = None
    try:
        psi = PortStockStoneItem.objects.filter(pss__ps__company = company, pss__ps__created = date, cus = cus, pss__stone = stone).last()
        if psi:
            receive = Weight.objects.filter(bws__company = company, bws__weight_type = 1, stone_type = stone, date = date, customer = cus).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
            psi.receive = receive
            psi.total = psi.quoted + receive - (psi.pay + psi.loss + psi.sell_cus + psi.other)
            tmp_pss_id = psi.pss.id
            psi.save()
    except PortStockStoneItem.DoesNotExist:
        pass

    if tmp_pss_id:
        #update total stock ของชนิดหินนี้
        pss = PortStockStone.objects.get(id = tmp_pss_id)
        pss.total = PortStockStoneItem.objects.filter(pss = tmp_pss_id).aggregate(s=Sum("total"))["s"] or Decimal('0.0')
        pss.save()

    if psi:
        updateTotalPortStockInMonth(psi.id)#คำนวนยอดยกมา จากวันก่อนและคำนวน total stock ใหม่

def updateTotalPortStockInMonth(ps_id):
    psi = PortStockStoneItem.objects.get(id = ps_id)

    created = psi.pss.ps.created
    last_date = created.replace(day=1) + relativedelta(months=1, days=-1)
    stone = psi.pss.stone.base_stone_type_id
    company = psi.pss.ps.company.id
    cus = psi.cus.customer_id

    all_stone = PortStockStoneItem.objects.filter(pss__ps__created__range=(created, last_date), pss__stone = stone, pss__ps__company = company, cus = cus).order_by('pss__ps__created')
    old_quot = None

    for i in all_stone:
        if old_quot is not None:#2
            i.quoted = old_quot
            i.total = old_quot + i.receive - (i.pay + i.loss + i.sell_cus + i.other)
            old_total = i.total
            i.save()

            #update total stock ของชนิดหินนี้
            pss = PortStockStone.objects.get(id = i.pss.id)
            pss.total = PortStockStoneItem.objects.filter(pss = i.pss.id).aggregate(s=Sum("total"))["s"] or Decimal('0.0')
            pss.save()

        if old_quot is not None:#3
            old_quot = old_total #ดึงรายการ total ของวันหน้ามาเป็นยกมาของอีกวันนึง
        elif i.total is not None:#1
            old_quot = i.total

#คำนวนยอดยกมา จากวันก่อนและคำนวน total stock ใหม่
def updateTotalPortStockInMonthByDate(previous_day, company):

    stock = PortStock.objects.filter(created = previous_day, company = company).values_list('id', flat=True).first()
    ss_items = PortStockStoneItem.objects.filter(pss__ps = stock)

    for ss in ss_items:
        created = ss.pss.ps.created
        last_date = created.replace(day=1) + relativedelta(months=1, days=-1)
        stone = ss.pss.stone.base_stone_type_id
        company = ss.pss.ps.company.id
        cus = ss.cus.customer_id

        all_stone = PortStockStoneItem.objects.filter(pss__ps__created__range=(created, last_date), pss__stone = stone, pss__ps__company = company, cus = cus).order_by('pss__ps__created')
        old_quot = None

        for i in all_stone:
            if old_quot is not None:#2
                i.quoted = old_quot
                i.total = old_quot + i.receive - (i.pay + i.loss + i.sell_cus + i.other)
                old_total = i.total
                i.save()

                #update total stock ของชนิดหินนี้
                pss = PortStockStone.objects.get(id = i.pss.id)
                pss.total = PortStockStoneItem.objects.filter(pss = i.pss.id).aggregate(s=Sum("total"))["s"] or Decimal('0.0')
                pss.save()

            if old_quot is not None:#3
                old_quot = old_total #ดึงรายการ total ของวันหน้ามาเป็นยกมาของอีกวันนึง
            elif i.total is not None:#1
                old_quot = i.total


def updateProdStockStoneItem(company, date):
    #ดึงชนิดหินทั้งหมดใน stock ของวันนั้นที่เป็น source ผลิต
    try:
        ss_prod = StockStoneItem.objects.filter(source__id = 2, ssn__stk__company = company, ssn__stk__created = date)
        if ss_prod:
            for i in ss_prod:
                prod = StoneEstimateItem.objects.filter(se__created = date, stone_type = i.ssn.stone, se__company = company).aggregate(s=Sum("total"))["s"] or Decimal('0.0')
                i.quantity = prod #ดึงจำนวนรวมผลิตมา (estimate)
                i.save()

                if i.ssn.id:
                    #update total stock ของชนิดหินนี้
                    ssn = StockStone.objects.get(id = i.ssn.id)
                    ssn.total = calculateTotalStock(i.ssn.id)
                    ssn.save()
                    
                    updateTotalStockInMonth(i.id)#คำนวนยอดยกมา จากวันก่อนและคำนวน total stock ใหม่
    except StockStoneItem.DoesNotExist:
        pass

#คำนวนยอดยกมา จากวันก่อนและคำนวน total stock ใหม่
def updateTotalStockInMonth(ss_id):
    ss = StockStoneItem.objects.get(id = ss_id)

    created = ss.ssn.stk.created
    last_date = created.replace(day=1) + relativedelta(months=1, days=-1)

    all_stone = StockStoneItem.objects.filter(source__id = 1, ssn__stk__created__range=(created, last_date), ssn__stone = ss.ssn.stone, ssn__stk__company = ss.ssn.stk.company).order_by('ssn__stk__created')
    old_quot = None

    for i in all_stone:
        if old_quot is not None:
            i.quantity = old_quot
            i.save()
            #update total stock ของชนิดหินนี้
            ssn = StockStone.objects.get(id = i.ssn.id)
            ssn.total = calculateTotalStock(i.ssn.id)
            old_total = ssn.total
            ssn.save()

        if old_quot is not None:
            old_quot = old_total #ดึงรายการ total ของวันหน้ามาเป็นยกมาของอีกวันนึง
        elif i.ssn.total is not None:
            old_quot = i.ssn.total

#คำนวนยอดยกมา จากวันก่อนและคำนวน total stock ใหม่
def updateTotalStockInMonthByDate(previous_day, company):

    stock = Stock.objects.filter(created = previous_day, company = company).values_list('id', flat=True).first()
    ss_items = StockStoneItem.objects.filter(ssn__stk = stock)

    for ss in ss_items:
        created = ss.ssn.stk.created
        last_date = created.replace(day=1) + relativedelta(months=1, days=-1)

        all_stone = StockStoneItem.objects.filter(source__id = 1, ssn__stk__created__range=(created, last_date), ssn__stone = ss.ssn.stone, ssn__stk__company = ss.ssn.stk.company).order_by('ssn__stk__created')
        old_quot = None

        for i in all_stone:
            if old_quot is not None:
                i.quantity = old_quot
                i.save()
                #update total stock ของชนิดหินนี้
                ssn = StockStone.objects.get(id = i.ssn.id)
                ssn.total = calculateTotalStock(i.ssn.id)
                old_total = ssn.total
                ssn.save()

            if old_quot is not None:
                old_quot = old_total #ดึงรายการ total ของวันหน้ามาเป็นยกมาของอีกวันนึง
            elif i.ssn.total is not None:
                old_quot = i.ssn.total

def updateGasPrice(company_id, created):
    try:
        gp = GasPrice.objects.filter(created = created, company = company_id)
        sum_oil = Weight.objects.filter(date = created, bws__weight_type = 1, bws__company = company_id, oil_content__gt = 0,).aggregate(s=Sum("oil_content"))["s"] or Decimal('0.0')
        for i in gp:
            #i.total_cost = i.cost * sum_oil เอาราคาทุนน้ำมัน ออกก่อน 13/02/2025
            i.total_sell = i.sell * sum_oil
            i.save()
    except GasPrice.DoesNotExist or Weight.DoesNotExist:
        pass

def updateOilCostAndSell(weight_id, company_id, created):
    try:
        weight = Weight.objects.get(weight_id = weight_id)
        gp = GasPrice.objects.filter(created = created, company = company_id).first()
        if gp:
            #weight.oil_cost = gp.cost * weight.oil_content เอาราคาทุนน้ำมัน ออกก่อน 13/02/2025
            weight.oil_sell = gp.sell * weight.oil_content
            weight.save()
    except GasPrice.DoesNotExist or Weight.DoesNotExist:
        pass

def updateSumEstimateItem(company_id, created, site_id):
    se_item = StoneEstimateItem.objects.filter(se__company = company_id, se__created = created, se__site__base_site_id = site_id)
    for i in se_item:
        i.total = calculateSumEstimateByCompany(created, company_id, site_id, i.stone_type.base_stone_type_id)
        i.save()

def updateProductionCapacity(company_id, date, site_id):
    pd_item = Production.objects.filter(company = company_id, created = date, site = site_id)
    for i in pd_item:
        i.capacity_per_hour = calculatProductionCapacity(company_id, date, i.site, i.line_type)
        i.save()

def searchDataCustomer(request):
    if 'customer_id' in request.GET and 'weight_id' in request.GET:
        customer_id = request.GET.get('customer_id')

        site = BaseCustomerSite.objects.filter(customer = customer_id).values('site__base_site_id','site__base_site_name')
    data = {
        'site_list': list(site),
    }
    return JsonResponse(data)

def searchDataBaesCustomer(request):
    if 'customer_id' in request.GET :
        customer_id = request.GET.get('customer_id')
        try:
            customer = BaseCustomer.objects.get(customer_id = customer_id)
            val = customer.customer_id
        except BaseCustomer.DoesNotExist:
            val = None
    data = {
        'val': val,
    }
    return JsonResponse(data)

def setDataCustomer(request):
    if 'customer_id' in request.GET:
        customer_id = request.GET.get('customer_id')
        qs = BaseCustomer.objects.get(customer_id = customer_id)
        val = qs.customer_id + ":" + qs.customer_name
    data = {
        'val': val,
    }
    return JsonResponse(data)

def setDataSite(request):
    if 'site_id' in request.GET:
        site_id = request.GET.get('site_id')
        qs = BaseSite.objects.get(base_site_id = site_id)
        val = qs.base_site_id + ":" + qs.base_site_name
    data = {
        'val': val,
    }
    return JsonResponse(data)

def setDataCarryType(request):
    if 'transport_id' in request.GET:
        transport_id = request.GET.get('transport_id')
        qs = BaseTransport.objects.get(base_transport_id = transport_id)
        val = qs.base_carry_type.base_carry_type_name  
    data = {
        'val': val,
    }
    return JsonResponse(data)

def createCustomerId(request):
    if 'job_type_id' in request.GET:
        job_type_id = request.GET.get('job_type_id')
        weight_type_id = request.GET.get('weight_type_id')

        if weight_type_id == '1' and job_type_id:
            missing_customer_id  = generateCodeId('BaseCustomer', 3, 1, job_type_id)
        elif weight_type_id == '2':
            missing_customer_id  = generateCodeId('BaseCustomer', 1, 2, None)
        else:
            missing_customer_id  = None

        val = missing_customer_id
    data = {
        'val': val,
    }
    return JsonResponse(data)

def createCarId(request):
    if 'car_team_id' in request.GET:
        car_team_id = request.GET.get('car_team_id')
        missing_customer_id  = generateCodeId('BaseCar', 3, None, car_team_id)
        val = missing_customer_id
    data = {
        'val': val,
    }
    return JsonResponse(data)

def searchNumCalQ(request):
    if 'stone_type_id' in request.GET:
        stone_type_id = request.GET.get('stone_type_id')
        qs = BaseStoneType.objects.get(base_stone_type_id = stone_type_id)
        val = qs.cal_q
    data = {
        'val': val,
    }
    return JsonResponse(data)

def searchTeamFromCar(request):
    if 'car_registration_name' in request.GET :
        car_registration_name = request.GET.get('car_registration_name')

        team = BaseCar.objects.filter(car_name = car_registration_name).values('base_car_team__car_team_id','base_car_team__car_team_name')
    data = {
        'team_list': list(team),
    }
    return JsonResponse(data)

def autocompalteCustomer(request):
    if 'term' in request.GET:
        term = request.GET.get('term')
        qs = BaseCustomer.objects.filter(Q(customer_id__icontains = term) | Q(customer_name__icontains = term))[:15]
        titles = list()
        for obj in qs:
            titles.append(obj.customer_id +":"+ obj.customer_name)
    return JsonResponse(titles, safe=False)

def autocompalteSite(request):
    if 'term' in request.GET:
        term = request.GET.get('term')
        qs = BaseSite.objects.filter(Q(base_site_id__icontains = term) | Q(base_site_name__icontains = term))[:15]
        titles = list()
        for obj in qs:
            titles.append(obj.base_site_id +":"+ obj.base_site_name)
    return JsonResponse(titles, safe=False)

def excelProductionByStone(request, my_q, list_date):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    comp = BaseCompany.objects.get(code = active)

    if comp.biz.id == 1:
        # Query ข้อมูลขาย
        #ดึงข้อมูลต้นทางกองสต็อค ที่มีในรายการชั่งของบริษัทนี้
        stock_name = Weight.objects.filter(my_q, Q(mill_name__contains='สต็อค') | Q(mill_name__contains='สต๊อก'), ~Q(site = '200PL') & ~Q(site = '300PL'), bws__weight_type = 1, bws__company__code = active).values_list('mill_id').order_by('mill_id').distinct()

        #ดึงข้อมูลต้นทางกองสต็อคและโรงโม่ของบริษัท
        m_comp_id = BaseMill.objects.filter(Q(m_comp__code = active) | Q(mill_id__in = stock_name)).values_list('mill_id').order_by('mill_id')
        data = Weight.objects.filter(my_q, ~Q(site = '200PL') & ~Q(site = '300PL'), mill__in = m_comp_id, bws__weight_type = 1).order_by('date','mill','stone_type').values_list('date','mill_name', 'stone_type_name').annotate(sum_weight_total = Sum('weight_total'))

        # Query ข้อมูลผลิตรวม
        s_comp_id = BaseSite.objects.filter(s_comp__code = active).values_list('base_site_id').order_by('base_site_id')
        data_sum_produc = Weight.objects.filter(my_q, site__in = s_comp_id, bws__weight_type = 2).order_by('date','site').values_list('date','site_name').annotate(sum_weight_total = Sum('weight_total'))
    
    elif comp.biz.id == 2:
        #ดึงข้อมูลต้นทางกองสต็อคและโรงโม่ของบริษัท
        data = Weight.objects.filter(my_q, site__store__in = [2,3], bws__weight_type = 1).order_by('date','site','stone_type').values_list('date','site_name', 'stone_type_name').annotate(sum_weight_total = Sum('weight_total'))

        # Query ข้อมูลรับเข้า
        s_comp_id = Weight.objects.filter(my_q, line_type = "สายยาว").values_list('customer__customer_id').order_by('customer__customer_id').distinct()
        data_sum_produc = Weight.objects.filter(my_q, customer__in = s_comp_id, line_type = "สายยาว", bws__weight_type = 1).order_by('date','customer').values_list('date', 'customer__customer_name').annotate(sum_weight_total=Sum("weight_total"))

    # Create a new workbook and get the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    if data or data_sum_produc:
        date_style = NamedStyle(name='custom_datetime', number_format='DD/MM/YYYY')
        
        # Create a set of all unique mill and stone values
        mills = set()
        stones = set()
        for item in data:
            mills.add(item[1])
            stones.add(item[2]) 

        mill_col_list = []

        
        # Create a list of colors for each line_type
        mill_colors = [generate_pastel_color() for i  in range(len(mills) + 1)]

        column_index = 2 + len(s_comp_id)
        for mill in mills:
            worksheet.cell(row=1, column=column_index, value=f'ยอดขาย{mill}')
            worksheet.merge_cells(start_row=1, start_column = column_index, end_row=1, end_column=(column_index + len(stones)) -1 )
            
            cell = worksheet.cell(row=1, column=column_index)
            cell.alignment = Alignment(horizontal='center')

            info = {}
            info['mill'] = mill
            info['strat_col'] = column_index
            info['end_col'] = column_index + len(stones)
            mill_col_list.append(info)

            #อัพเดทจำนวน col ตามชนิดหิน
            column_index += len(stones)

        #set color in header in row 1-2
        for row in worksheet.iter_rows(min_row=1, max_row=2):
            # Set the background color for each cell in the column
            for cell in row:
                #cell.border = Border(top=side, bottom=side, left=side, right=side)
                cell.alignment = Alignment(horizontal='center')
                line_index = (cell.column - ( len(s_comp_id) + 2 )) // (len(stones))
                fill_color = mill_colors[line_index % len(mill_colors)]
                fill = PatternFill(start_color=fill_color, fill_type="solid")
                cell.fill = fill

        # Write headers row 2 to the worksheet
        column_index = 2 + len(s_comp_id)
        for mill in mills:
            for stone in stones:
                worksheet.cell(row=2, column=column_index, value=stone).alignment = Alignment(horizontal='center')
                column_index += 1

        # Create a dictionary to store data by date, mill, and stone
        date_data = {}

        # Loop through the data and populate the dictionary  
        for item in data:
            date = item[0]
            mill = item[1]
            stone = item[2]
            value = item[3]

            if date not in date_data:
                date_data[date] = {}

            if mill not in date_data[date]:
                date_data[date][mill] = {}

            date_data[date][mill][stone] = value

        row_index = 3
        for idl, ldate in enumerate(list_date):
            #เขียนวันที่ใน worksheet column 1
            worksheet.cell(row=idl+3, column=1, value=ldate).style = date_style
            worksheet.cell(row=idl+3, column=1).alignment = Alignment(horizontal='center')

            for date, mill_data in date_data.items():
                #เขียน weight total ของแต่ละหินใน worksheet
                if worksheet.cell(row=idl+3, column = 1).value == date:
                    column_index = 2 + len(s_comp_id)
                    for mill in mills:
                        stone_data = mill_data.get(mill, {})
                        for stone in stones:
                            value = stone_data.get(stone, '')
                            worksheet.cell(row=idl+3, column=column_index, value=value).number_format = '#,##0.00'
                            column_index += 1
                    #row_index += 1
            row_index += 1    

        #นำข้อมูลการผลิตมาเรียง
        sorted_queryset = sorted(data_sum_produc, key=lambda x: x[0])

        # Create a dictionary to store the summed values by date and mill_name
        summed_values = {}
        for date, mill_name, value in sorted_queryset:
            key = (date, mill_name)
            summed_values[key] = summed_values.get(key, 0) + float(value)

        mill_produc_list = []

        # Create headers
        headers = ['Date'] + list(set(row[1] for row in sorted_queryset))
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value =  f'ยอดผลิต{header}' if comp.biz.id == 1 else f'ยอดรับเข้า{header}'
            cell.alignment = Alignment(horizontal='center')
            worksheet.merge_cells(start_row=1, start_column = col_num, end_row=2, end_column=col_num)

            info = {}
            if header != 'Date':
                info['mill'] = header
                info['col'] = col_num
                mill_produc_list.append(info)

        # Fill in the data ยืด วันที่ จาก วันที่ขายทั้งหมด set(row[0] for row in data หากยึด วันที่ตามวันที่ผลิต set(row[0] for row in sorted_queryset
        for row_num, date in enumerate(sorted(set(row for row in list_date)), 2):
            #worksheet.cell(row=row_num, column=4, value=date)
            row_num += 1 
            for col_num, mill_name in enumerate(headers[1:], 2):
                key = (date, mill_name)
                value = summed_values.get(key, '')
                worksheet.cell(row=row_num, column=col_num, value=value).number_format = '#,##0.00'
        

        # Write headers row 1 to the worksheet
        worksheet.cell(row=1, column=1, value='Date')

        worksheet.cell(row=row_index, column=1, value='รวมทั้งสิ้น')
        sum_by_col = Decimal('0.00')
        for col in range(2, column_index):
            for row in range(3, row_index):
                sum_by_col = sum_by_col + Decimal( worksheet.cell(row=row, column=col).value or '0.00' )
            worksheet.cell(row=row_index, column=col, value=sum_by_col).number_format = '#,##0.00'
            worksheet.cell(row=row_index, column=col).font = Font(bold=True)
            sum_by_col = Decimal('0.00')

        '''
        #คิดเป็นเปอร์เซ็น
        worksheet.cell(row=row_index+1, column=1, value="เปอร์เซ็นต์เฉลี่ย")
        for col, produc in zip(mill_col_list, mill_produc_list):
            if col['mill'] == produc['mill']:
                for i in range(col['strat_col'], col['end_col']):
                    sum_produc_val = Decimal(worksheet.cell(row=row_index, column = produc['col']).value or '1.00' )
                    val = Decimal(worksheet.cell(row=row_index, column = i).value or '1.00' )
                    percent = int(val/sum_produc_val * 100)

                    worksheet.cell(row=row_index+1, column=i, value = " " if val == Decimal('1.00') else f'{percent}%').alignment = Alignment(horizontal='right')
                    worksheet.cell(row=row_index+1, column=i).font = Font(color="FF0000")        
        '''

        # Set the column widths
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[2].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width
            worksheet.column_dimensions[column].height = 20

        side = Side(border_style='thin', color='000000')
        set_border(worksheet, side)
        worksheet.freeze_panes = "B3" #freeze
    else:
        worksheet.cell(row = 1, column = 1, value = f'ไม่มีข้อมูลยอดขายตามประเภทหินของเดือนนี้')

    # Save workbook into memory
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    size = output.getbuffer().nbytes

    # Generator to stream file in chunks
    def file_iterator(buffer, chunk_size=8192):
        while True:
            data = buffer.read(chunk_size)
            if not data:
                break
            yield data

    response = StreamingHttpResponse(
        file_iterator(output),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="sales_daily_({active}).xlsx"'
    response["Content-Length"] = str(size)
    return response

def exportExcelProductionByStone(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    doc_id = request.GET.get('doc_id') or None
    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None
    customer_name = request.GET.get('customer_name') or None
    stone_type = request.GET.get('stone_type') or None

    my_q = Q()
    if doc_id is not None:
        my_q &= Q(doc_id__icontains = doc_id)
    if start_created is not None:
        my_q &= Q(date__gte = start_created)
    if end_created is not None:
        my_q &=Q(date__lte = end_created)
    if customer_name is not None :
        my_q &=Q(customer_name__icontains = customer_name)
    if stone_type is not None :
        my_q &=Q(stone_type_name__icontains = stone_type)

    my_q &= Q(bws__company__code__in = company_in)
    my_q &= ~Q(customer_name ='ยกเลิก')
   
    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    startDate = datetime.strptime(start_created or startDateInMonth(previous_date_time.strftime('%Y-%m-%d')), "%Y-%m-%d").date()
    endDate = datetime.strptime(end_created or previous_date_time.strftime('%Y-%m-%d'), "%Y-%m-%d").date()

    #สร้าง list ระหว่าง start_date และ end_date
    list_date = [startDate+timedelta(days=x) for x in range((endDate-startDate).days + 1)]

    response = excelProductionByStone(request, my_q, list_date)
    return response

def exportExcelProductionByStoneInDashboard(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    ''' แบบเก่าดึง รายงานผลิตหิน ตามเดือนนั้นๆ 09/05/2024
    #ดึงรายงานของเดือนนั้นๆ
    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    end_created = previous_date_time.strftime('%Y-%m-%d')
    start_created = startDateInMonth(end_created)
    '''
    end_created = request.session['db_end_date']
    start_created = request.session['db_start_date']

    my_q = Q()
    if start_created is not None:
        my_q &= Q(date__gte = start_created)
    if end_created is not None:
        my_q &=Q(date__lte = end_created)

    my_q &= Q(bws__company__code__in = company_in)
    my_q &= ~Q(customer_name ='ยกเลิก')

    #เปลี่ยนออกเป็น ดึงรายงานของเดือนนั้นๆเท่านั้น
    startDate = datetime.strptime(start_created, "%Y-%m-%d").date()
    endDate = datetime.strptime(end_created, "%Y-%m-%d").date()

    #สร้าง list ระหว่าง start_date และ end_date
    list_date = [startDate+timedelta(days=x) for x in range((endDate-startDate).days + 1)]

    response = excelProductionByStone(request, my_q, list_date)
    return response

def excelProductionByStoneAndMonth(request, my_q, list_date):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    comp = BaseCompany.objects.get(code = active)

    if comp.biz.id == 1:
        # Query ข้อมูลขาย
        #ดึงข้อมูลต้นทางกองสต็อค ที่มีในรายการชั่งของบริษัทนี้
        stock_name = Weight.objects.filter(my_q, Q(mill_name__contains='สต็อค') | Q(mill_name__contains='สต๊อก'), ~Q(site = '200PL') & ~Q(site = '300PL'), bws__weight_type = 1, bws__company__code = active).values_list('mill_id').order_by('mill_id').distinct()
        
        #ดึงข้อมูลต้นทางกองสต็อคและโรงโม่ของบริษัท
        m_comp_id = BaseMill.objects.filter(Q(m_comp__code = active) | Q(mill_id__in = stock_name)).values_list('mill_id').order_by('mill_id')
        data = Weight.objects.filter(my_q, ~Q(site = '200PL') & ~Q(site = '300PL'), mill__in = m_comp_id, bws__weight_type = 1).annotate(
            month=ExtractMonth('date'),
            year=ExtractYear('date')
        ).values_list('year', 'month', 'mill_name', 'stone_type_name').annotate(
            sum_weight_total=Sum('weight_total')
        ).order_by('year', 'month', 'mill_name', 'stone_type_name')

        # Query ข้อมูลผลิตรวม
        s_comp_id = BaseSite.objects.filter(s_comp__code = active).values_list('base_site_id').order_by('base_site_id')
        data_sum_produc = Weight.objects.filter(my_q, site__in = s_comp_id, bws__weight_type = 2).annotate(
            month=ExtractMonth('date'),
            year=ExtractYear('date')
        ).values_list('year', 'month', 'site_name').annotate(
            sum_weight_total=Sum('weight_total')
        ).order_by('year', 'month', 'site_name')

    elif comp.biz.id == 2:
        # Query ข้อมูลขาย
        #ดึงข้อมูลต้นทางกองสต็อค ที่มีในรายการชั่งของบริษัทนี้
        data = Weight.objects.filter(my_q, site__store__in = [2,3], bws__weight_type = 1).annotate(
            month=ExtractMonth('date'),
            year=ExtractYear('date')
        ).values_list('year', 'month', 'site_name', 'stone_type_name').annotate(
            sum_weight_total=Sum('weight_total')
        ).order_by('year', 'month', 'site_name', 'stone_type_name')

        s_comp_id = Weight.objects.filter(my_q, line_type = "สายยาว").values_list('customer__customer_id').order_by('customer__customer_id').distinct()
        data_sum_produc = Weight.objects.filter(my_q, customer__in = s_comp_id, line_type = "สายยาว", bws__weight_type = 1).annotate(
            month=ExtractMonth('date'),
            year=ExtractYear('date')
        ).values_list('year', 'month', 'customer__customer_name').annotate(
            sum_weight_total=Sum('weight_total')
        ).order_by('year', 'month', 'customer__customer_name')

    # Create a new workbook and get the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    thai_months = ['', 'ม.ค.', 'ก.พ.', 'มี.ค.', 'เม.ย.', 'พ.ค.', 'มิ.ย.','ก.ค.', 'ส.ค.', 'ก.ย.', 'ต.ค.', 'พ.ย.', 'ธ.ค.']

    if data or data_sum_produc:
        date_style = NamedStyle(name='custom_datetime', number_format='DD/MM/YYYY')
        
        # Create a set of all unique mill and stone values
        mills = set()
        stones = set()
        for item in data:
            mills.add(item[2])
            stones.add(item[3]) 

        mill_col_list = []

        # Create a list of colors for each line_type
        mill_colors = [generate_pastel_color() for i  in range(len(mills) + 1)]

        column_index = 2 + len(s_comp_id)
        for mill in mills:
            worksheet.cell(row=1, column=column_index, value=f'ยอดขาย{mill}')
            worksheet.merge_cells(start_row=1, start_column = column_index, end_row=1, end_column=(column_index + len(stones)) -1 )
            
            cell = worksheet.cell(row=1, column=column_index)
            cell.alignment = Alignment(horizontal='center')

            info = {}
            info['mill'] = mill
            info['strat_col'] = column_index
            info['end_col'] = column_index + len(stones)
            mill_col_list.append(info)

            #อัพเดทจำนวน col ตามชนิดหิน
            column_index += len(stones)

        #set color in header in row 1-2
        for row in worksheet.iter_rows(min_row=1, max_row=2):
            # Set the background color for each cell in the column
            for cell in row:
                #cell.border = Border(top=side, bottom=side, left=side, right=side)
                cell.alignment = Alignment(horizontal='center')
                line_index = (cell.column - (len(s_comp_id) + 2)) // (len(stones))
                fill_color = mill_colors[line_index % len(mill_colors)]
                fill = PatternFill(start_color=fill_color, fill_type="solid")
                cell.fill = fill

        # Write headers row 2 to the worksheet
        column_index = 2 + len(s_comp_id)
        for mill in mills:
            for stone in stones:
                worksheet.cell(row=2, column=column_index, value=stone).alignment = Alignment(horizontal='center')
                column_index += 1

        # Create a dictionary to store data by date, mill, and stone
        date_data = {}

        # Loop through the data and populate the dictionary  
        for item in data:
            date = (item[0], item[1])
            mill = item[2]
            stone = item[3]
            value = item[4]

            if date not in date_data:
                date_data[date] = {}

            if mill not in date_data[date]:
                date_data[date][mill] = {}

            date_data[date][mill][stone] = value

        row_index = 3
        for idl, ldate in enumerate(list_date):
                #เขียนวันที่ใน worksheet column 1
                formatted_date = f"{thai_months[ldate[1]]} {ldate[0]}"
                worksheet.cell(row=idl+3, column=1, value = str(ldate))
                worksheet.cell(row=idl+3, column=1).alignment = Alignment(horizontal='center')

                for date, mill_data in date_data.items():
                    #เขียน weight total ของแต่ละหินใน worksheet
                    if str(worksheet.cell(row=idl+3, column = 1).value) == str(date):
                        column_index = 2 + len(s_comp_id)
                        for mill in mills:
                            stone_data = mill_data.get(mill, {})
                            for stone in stones:
                                value = stone_data.get(stone, '')
                                worksheet.cell(row=idl+3, column=column_index, value=value).number_format = '#,##0.00'
                                column_index += 1
                        #row_index += 1
                worksheet.cell(row=idl+3, column=1, value = formatted_date) #เปลี่ยน วันที่เป็นเดือนไทย
                row_index += 1
 
        #นำข้อมูลการผลิตมาเรียง
        sorted_queryset = sorted(data_sum_produc, key=lambda x: (x[0], x[1]))

        # Create a dictionary to store the summed values by date and mill_name
        summed_values = {}
        for year, month, mill_name, value in sorted_queryset:
            key = ((year, month), mill_name)
            summed_values[key] = summed_values.get(key, 0) + float(value)

        mill_produc_list = []

        # Create headers
        headers = ['Date'] + list(set(row[2] for row in sorted_queryset))
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = f'ยอดผลิต{header}'
            cell.alignment = Alignment(horizontal='center')
            worksheet.merge_cells(start_row=1, start_column = col_num, end_row=2, end_column=col_num)

            info = {}
            if header != 'Date':
                info['mill'] = header
                info['col'] = col_num
                mill_produc_list.append(info)


        # Fill in the data ยืด วันที่ จาก วันที่ขายทั้งหมด set(row[0] for row in data หากยึด วันที่ตามวันที่ผลิต set(row[0] for row in sorted_queryset
        for row_num, date in enumerate(sorted(set(row for row in list_date)), 2):
            #worksheet.cell(row=row_num, column=4, value=date)
            row_num += 1 
            for col_num, mill_name in enumerate(headers[1:], 2):
                key = (date, mill_name)
                value = summed_values.get(key, '')
                worksheet.cell(row=row_num, column=col_num, value=value).number_format = '#,##0.00'
        

        # Write headers row 1 to the worksheet
        worksheet.cell(row=1, column=1, value='Date')

        worksheet.cell(row=row_index, column=1, value='รวมทั้งสิ้น')
        sum_by_col = Decimal('0.00')
        for col in range(2, column_index):
            for row in range(3, row_index):
                sum_by_col = sum_by_col + Decimal( worksheet.cell(row=row, column=col).value or '0.00' )
            worksheet.cell(row=row_index, column=col, value=sum_by_col).number_format = '#,##0.00'
            worksheet.cell(row=row_index, column=col).font = Font(bold=True)
            sum_by_col = Decimal('0.00')

        '''
        #คิดเป็นเปอร์เซ็น
        worksheet.cell(row=row_index+1, column=1, value="เปอร์เซ็นต์เฉลี่ย")
        for col, produc in zip(mill_col_list, mill_produc_list):
            if col['mill'] == produc['mill']:
                for i in range(col['strat_col'], col['end_col']):
                    sum_produc_val = Decimal(worksheet.cell(row=row_index, column = produc['col']).value or '1.00' )
                    val = Decimal(worksheet.cell(row=row_index, column = i).value or '1.00' )
                    percent = int(val/sum_produc_val * 100)

                    worksheet.cell(row=row_index+1, column=i, value = " " if val == Decimal('1.00') else f'{percent}%').alignment = Alignment(horizontal='right')
                    worksheet.cell(row=row_index+1, column=i).font = Font(color="FF0000")       
        '''


        # Set the column widths
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[2].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width
            worksheet.column_dimensions[column].height = 20

        side = Side(border_style='thin', color='000000')
        set_border(worksheet, side)
        worksheet.freeze_panes = "B3" #freeze
    else:
        worksheet.cell(row = 1, column = 1, value = f'ไม่มีข้อมูลยอดขายตามประเภทหินของเดือนนี้')

    # Save workbook into memory
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    size = output.getbuffer().nbytes

    # Generator to stream file in chunks
    def file_iterator(buffer, chunk_size=8192):
        while True:
            data = buffer.read(chunk_size)
            if not data:
                break
            yield data

    response = StreamingHttpResponse(
        file_iterator(output),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="sales_monthly_({active}).xlsx"'
    response["Content-Length"] = str(size)
    return response

def exportExcelProductionByStoneAndMonthInDashboard(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    #ดึงรายงานของเดือนนั้นๆ
    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    end_created = previous_date_time.strftime('%Y-%m-%d')
    start_created = startDateInMonth(end_created)

    my_q = Q()
    my_q &= Q(bws__company__code__in = company_in)
    my_q &= ~Q(customer_name ='ยกเลิก')

    #ดึงรายงานของปีนั้นๆทั้งหมด
    startDate = datetime.strptime(start_created, "%Y-%m-%d").date()
    endDate = datetime.strptime(end_created, "%Y-%m-%d").date()

    # สร้าง list ระหว่าง start_date และ end_date ในรูปแบบ (year, month)
    list_year_month = [(year, month) for year in range(startDate.year, endDate.year+1) for month in range(1, 13)]

    response = excelProductionByStoneAndMonth(request, my_q, list_year_month)
    return response

def exportExcelProductionByStoneAndMonth(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    doc_id = request.GET.get('doc_id') or None
    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None
    customer_name = request.GET.get('customer_name') or None
    stone_type = request.GET.get('stone_type') or None

    my_q = Q()
    if doc_id is not None:
        my_q &= Q(doc_id__icontains = doc_id)
    if start_created is not None:
        my_q &= Q(date__gte = start_created)
    if end_created is not None:
        my_q &=Q(date__lte = end_created)
    if customer_name is not None :
        my_q &=Q(customer_name__icontains = customer_name)
    if stone_type is not None :
        my_q &=Q(stone_type_name__icontains = stone_type)

    my_q &= Q(bws__company__code__in = company_in)
    my_q &= ~Q(customer_name ='ยกเลิก')
   
    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    startDate = datetime.strptime(start_created or startDateInMonth(previous_date_time.strftime('%Y-%m-%d')), "%Y-%m-%d").date()
    endDate = datetime.strptime(end_created or previous_date_time.strftime('%Y-%m-%d'), "%Y-%m-%d").date()

    # สร้าง list ระหว่าง start_date และ end_date ในรูปแบบ (year, month)
    list_year_month = [(year, month) for year in range(startDate.year, endDate.year+1) for month in range(1, 13)]

    response = excelProductionByStoneAndMonth(request, my_q, list_year_month)
    return response

@login_required(login_url='login')
def viewProduction(request):
    #active : active คือแท็ปบริษัท active
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')

    data = Production.objects.filter(company__code__in = company_in).order_by('-created', 'site')

    #กรองข้อมูล
    myFilter = ProductionFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    product = p.get_page(page)

    context = {'production_page':'active', 'product': product,'filter':myFilter, active :"active",}
    return render(request, "production/viewProduction.html",context)

@login_required(login_url='login')
def summaryProduction(request):
    #active : active คือแท็ปบริษัท active
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    
    ''' แบบเก่าดึง Performance การผลิต ตามเดือนนั้นๆ 09/05/2024
    #ดึงข้อมูลย้อนหลัง 1 วัน
    previous_date_time = date_object - timedelta(days=1)

    end_created = previous_date_time.strftime('%Y-%m-%d')
    start_created = startDateInMonth(end_created)
    '''
    #ดึงข้อมูลวันนี้
    date_object = datetime.today()

    end_created = request.session['db_end_date']
    start_created = request.session['db_start_date']

    start_day = datetime.strptime(start_created, "%Y-%m-%d")
    end_day = datetime.strptime(end_created, "%Y-%m-%d")

    #หาวันสุดท้ายที่คีย์ข้อมูลผลิต
    date_last_pd = Production.objects.filter(company__code__in = company_in, created__range=(start_created, end_created)).values_list('created', flat=True).order_by('-created').first()

    b_site = Production.objects.filter(company__code__in = company_in).values('site').distinct()

    real_pd = Weight.objects.filter(bws__company__code__in = company_in, site__in = b_site, date__range=(start_created, date_last_pd), bws__weight_type = 2).values('site__base_site_id', 'site__base_site_name').order_by('site__base_site_id').annotate(sum_weight = Sum("weight_total"))

    pd = Production.objects.filter(company__code__in = company_in, created__range=(start_created, date_last_pd)).values('site__base_site_id', 'site__base_site_name', 'pd_goal__accumulated_goal').order_by('site__base_site_id').annotate(count=Count('site__base_site_id') 
        , sum_goal = Sum('goal'), sum_loss = Sum('total_loss_time'), sum_actual = Sum('actual_time'), sum_run = Sum('run_time'), percent_p = ExpressionWrapper(F('sum_run') / F('sum_actual'), output_field= models.DecimalField())
        , sum_uncontrol=Sum(Case(When(uncontrol_time__isnull=True, then=Value(timedelta(0))), default='uncontrol_time', output_field = models.DurationField()))
        , sum_loss_n_un = ExpressionWrapper(F('sum_loss') - F('sum_uncontrol'), output_field= models.DurationField())
        , working_time = ExpressionWrapper(F('sum_actual') - F('sum_uncontrol') , output_field= models.DurationField()), working_time_de = ExpressionWrapper(F('sum_actual') - F('sum_uncontrol') , output_field= models.IntegerField())
        , stone_time = ExpressionWrapper(F('working_time') - F('sum_loss_n_un') , output_field= models.DurationField()), stone_time_de = ExpressionWrapper(F('working_time') - F('sum_loss_n_un') , output_field= models.IntegerField())
        , percent_a = ExpressionWrapper(F('stone_time') / F('working_time') * 100, output_field= models.DecimalField())
        , percent_goal = ExpressionWrapper(F('sum_goal') / F('pd_goal__accumulated_goal') * 100, output_field= models.IntegerField()), loss_weight = ExpressionWrapper(F('pd_goal__accumulated_goal') - F('sum_goal'), output_field= models.FloatField())
        , capacity = ExpressionWrapper(F('sum_goal') / (F('working_time_de')/1000000/3600), output_field= models.DecimalField())
        , percent_loss = ExpressionWrapper(F('sum_loss_n_un') / F('working_time') * 100, output_field= models.DecimalField()))

    # M = เครื่องจักรหลัก, S = เครื่องจักรรอง ไว้แสดงข้อมูลเท่านั้น
    pd_loss_mc = ProductionLossItem.objects.filter(production__company__code__in = company_in, production__created__range=(start_created, date_last_pd), mc_type__kind = 'M').order_by('production__site__base_site_id').values('production__site__base_site_id', 'mc_type').annotate(sum_time = Sum('loss_time'))

    mc_type_time  = ProductionMachineItem.objects.filter(production__company__code__in = company_in, production__created__range=(start_created, date_last_pd)).values('mc_type__id','mc_type__name').distinct()
    pd_mc_time = ProductionMachineItem.objects.filter(production__company__code__in = company_in, production__created__range=(start_created, date_last_pd)).order_by('production__site__base_site_id').values('production__site__base_site_id', 'mc_type__name', 'mc_type').annotate(sum_diff_time = Sum('diff_time'))
    
    mc_loos_type = ProductionLossItem.objects.filter(production__company__code__in = company_in, production__created__range=(start_created, date_last_pd), mc_type__kind = 'S').order_by('mc_type__id').values('mc_type__name', 'loss_type__name').distinct()
    pd_loss_pro = ProductionLossItem.objects.filter(production__company__code__in = company_in, production__created__range=(start_created, date_last_pd), mc_type__kind = 'S').order_by('production__site__base_site_id', 'mc_type__id').values('production__site__base_site_id', 'mc_type__id', 'mc_type__name', 'loss_type__name').annotate(sum_time = Sum('loss_time'))
    mc_type  = BaseMachineType.objects.filter(kind = 'M')

    s_comp_id = BaseSite.objects.filter(s_comp__code = active).values_list('base_site_id').order_by('base_site_id')

    s_target = BaseSite.objects.filter(s_comp__code = active).values('base_site_id', 'target').order_by('base_site_id')
    
    list_ls_name = [[] for _ in range(len(s_comp_id))]
    list_ls_val = [[] for _ in range(len(s_comp_id))]
    list_ls = []

    for i, mill_id in enumerate(s_comp_id):
        list_ls_name[i] = getLossNameByMill(company_in, mill_id, start_created, date_last_pd, 1)
        list_ls_val[i] = getLossNameByMill(company_in, mill_id, start_created, date_last_pd, 2)
        list_ls.append((list_ls_name[i], list_ls_val[i]))

    pd_loss_all = ProductionLossItem.objects.filter(production__company__code__in = company_in, production__created__range=(start_created, date_last_pd)).order_by('production__site__base_site_id').values('production__site__base_site_id', 'mc_type__name').annotate(sum_time = Sum('loss_time'))

    context = {'dashboard_page':'active','pd':pd,
               'pd_loss_mc':pd_loss_mc,
               'mc_type_time': mc_type_time,
               'pd_mc_time': pd_mc_time,
               'pd_loss_pro':pd_loss_pro,
               'date_object':date_object, 'mc_type':mc_type,
               'list_ls': list_ls,
               'pd_loss_all':pd_loss_all,
               'mc_loos_type':mc_loos_type,
               'real_pd':real_pd,
               's_target':s_target,
               'start_day':start_day,
               'end_day': end_day,
               'last_day': date_last_pd,
               active :"active",
    }
    return render(request, "production/summaryProduction.html",context)

def extract_month_year(date):
    return date.strftime("%Y-%m")

@login_required(login_url='login')
def monthlyProduction(request):
    #active : active คือแท็ปบริษัท active
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    end_created = request.session['db_end_date']
    start_created = request.session['db_start_date']

    current_date_time = datetime.strptime(end_created, "%Y-%m-%d")
    last_day = calendar.monthrange(current_date_time.year, current_date_time.month)[1]

    first_day_in_year = f"{current_date_time.year}-01-01"
    last_day_in_month = f"{current_date_time.year}-{current_date_time.month:02d}-{last_day:02d}"

    this_year = current_date_time.year
    current_year = current_date_time.year - 1
    numeric_month = current_date_time.month

    s_comp = BaseSite.objects.filter(s_comp__code = active).values_list('base_site_id', flat=True).order_by('base_site_id')
    s_comp_name = BaseSite.objects.filter(s_comp__code = active).values_list('base_site_name', flat=True).order_by('base_site_name')
    #ดึงข้อมูล 2025 ขึ้นไป
    stone_id = StoneEstimateItem.objects.filter(se__site__in = s_comp, se__created__range=(first_day_in_year, last_day_in_month), percent__gt = 0).values_list('stone_type__base_stone_type_id', flat=True).order_by('stone_type__base_stone_type_id').distinct() #ดึงเฉพาะ stone_id ที่มีการคีย์ percent > 0

    date_data = StoneEstimateItem.objects.filter(se__site__in = s_comp, se__created__range=(first_day_in_year, last_day_in_month), stone_type__in = stone_id
    ).annotate(
        year=ExtractYear('se__created'),
        month=ExtractMonth('se__created'),
    ).values_list('year', 'month', 'se__site__base_site_name', 'stone_type__base_stone_type_name'
    ).annotate(
        sum=Coalesce(Sum('total'), Value(0), output_field=models.DecimalField()),
    ).order_by('se__site', 'se__created', 'stone_type')
    
    aggregated_results = {}
    produc_run_results = {}
    produc_work_results = {}
    produc_capacity_results = {}
    produc_hour_per_day_results = {}

    
    all_month_years = [f"{current_date_time.year}-{str(month).zfill(2)}" for month in range(1, numeric_month + 1)]
    default_month_years = [
        (int(my.split('-')[0]), int(my.split('-')[1])) for my in all_month_years
    ]
    
    all_thai_months = ['ม.ค.', 'ก.พ.', 'มี.ค.', 'เม.ย.', 'พ.ค.', 'มิ.ย.','ก.ค.', 'ส.ค.', 'ก.ย.', 'ต.ค.', 'พ.ย.', 'ธ.ค.']
    thai_months = all_thai_months[:numeric_month]

    for dt in date_data:
        year, month, site_name, stone_type_name, sum_total = dt

        month_year = f"{year}-{str(month).zfill(2)}"

        if site_name not in aggregated_results:
            aggregated_results[site_name] = {}
        if stone_type_name not in aggregated_results[site_name]:
            aggregated_results[site_name][stone_type_name] = {}

        for my in all_month_years:
            if my not in aggregated_results[site_name][stone_type_name]:
                aggregated_results[site_name][stone_type_name][my] = Decimal(0)

        aggregated_results[site_name][stone_type_name][month_year] += sum_total

    ###################### start สรุปข้อมูลผลิต #####################
    product_data = list(
        Production.objects.filter(
            site__in=s_comp,
            created__range=(first_day_in_year, last_day_in_month)
        ).annotate(
            year=ExtractYear('created'),
            month=ExtractMonth('created'),
            working_time=ExpressionWrapper(F('actual_time') - F('total_loss_time'), output_field=models.DurationField()),
            hour_per_day=ExpressionWrapper(F('actual_time') / (F('actual_time') - F('total_loss_time')), output_field=models.DecimalField()),
        ).values_list(
            'year', 'month', 'site__base_site_name'
        ).annotate(
            sum_run=Sum('run_time'),
            sum_total_working_time=Sum('working_time'),
            sum_hour_per_day=Sum('hour_per_day'),
            sum_capacity_per_hour=Sum('capacity_per_hour'),
        )
    )

    #เพิ่ม ข้อมูล default product_data หากไม่มีข้อมูลผลิตของ โรงโม่ ในบริษัทนั้นๆ
    existing_sites = {pd[2] for pd in product_data}
    missing_sites = set(s_comp_name) - existing_sites

    for site in missing_sites:
        for year, month in default_month_years:
            product_data.append((
                year, month, site,
                Decimal(0),      # sum_run
                timedelta(0),    # sum_total_working_time
                Decimal(0),      # sum_hour_per_day
                Decimal(0),      # sum_capacity_per_hour
            ))

    for pd in product_data:
        year, month, site_name, sum_run, sum_total_working_time, sum_hour_per_day, sum_capacity_per_hour  = pd
        month_year = f"{year}-{str(month).zfill(2)}"
        
        update_results(all_month_years, 1, produc_run_results, site_name, month_year, sum_run)
        update_results(all_month_years, 1, produc_work_results, site_name, month_year, sum_total_working_time)
        update_results(all_month_years, 2, produc_capacity_results, site_name, month_year, sum_capacity_per_hour)
        update_results(all_month_years, 2, produc_hour_per_day_results, site_name, month_year, sum_hour_per_day)
        ###################### end สรุปข้อมูลผลิต ####################

    ################ start รวมทุกๆโรงโม่ ############################
    totals = {}  # Initialize a dictionary to hold totals for each stone type
    for site_name, stone_data in aggregated_results.items():
        for stone_type_name, date_data in stone_data.items():
            for created_date, value in date_data.items():
                if stone_type_name not in totals:
                    totals[stone_type_name] = {}
                if created_date not in totals[stone_type_name]:
                    totals[stone_type_name][created_date] = Decimal(0)
                totals[stone_type_name][created_date] += value

    # Add totals to results under a special key like "Total"
    total_values = {}
    for stone_type_name, date_data in totals.items():
        total_values[stone_type_name] = {}
        for created_date, total_value in date_data.items():
            total_values[stone_type_name][created_date] = total_value

    aggregated_results["Total"] = total_values
    ################ end รวมทุกๆโรงโม่ ###############################

    ################ start รวมทุกชนิดหินในเดือนนั้นๆ ####################
    sum_aggregated = {}
    for site_name, site_data in aggregated_results.items():
        for stone_type, stone_type_data in site_data.items():
            for month_year, result in stone_type_data.items():
                if site_name not in sum_aggregated:
                    sum_aggregated[site_name] = {}
                if month_year not in sum_aggregated[site_name]:
                    sum_aggregated[site_name][month_year] = 0
                    
                sum_aggregated[site_name][month_year] += result
    ################ end รวมทุกชนิดหินในเดือนนั้นๆ ######################
    data_stone_old_year = None
    data_run_old_year = None
    data_work_old_year = None
    data_cap_old_year = None
    data_hpd_old_year = None

    if this_year == 2025:
        data_stone_old_year = strToArrList(active, 'weight')
        data_run_old_year = strToArrList(active, 'prod_run')
        data_work_old_year = strToArrList(active, 'prod_work')
        data_cap_old_year = strToArrList(active, 'prod_cap')
        data_hpd_old_year = strToArrList(active, 'prod_hpd') 
    elif this_year > 2025:
        data_stone_old_year = strToArrListOldYear(active, 'weight', s_comp, current_year)
        data_run_old_year = strToArrListOldYear(active, 'prod_run', s_comp, current_year)
        data_work_old_year = strToArrListOldYear(active, 'prod_work', s_comp, current_year)
        data_cap_old_year = strToArrListOldYear(active, 'prod_cap', s_comp, current_year)
        data_hpd_old_year = strToArrListOldYear(active, 'prod_hpd', s_comp, current_year)

    context = {
               'aggregated_results':aggregated_results,
               'produc_run_results': produc_run_results,
               'produc_work_results': produc_work_results,
               'produc_capacity_results': produc_capacity_results,
               'produc_hour_per_day_results': produc_hour_per_day_results,
               'sum_aggregated': sum_aggregated,
               'data_stone_old_year': data_stone_old_year,
               'data_run_old_year': data_run_old_year,
               'data_work_old_year': data_work_old_year,
               'data_cap_old_year': data_cap_old_year,
               'data_hpd_old_year': data_hpd_old_year,
               'now_year': current_date_time.year,
               'current_year': current_year,
               'thai_months': thai_months,
                active :"active",
              }
    return render(request, "production/monthlyProduction.html",context)

def strToArrList(active, field):
    try:
        queryset_string = SetWeightOY.objects.filter(comp__code = active).values_list(field)
        data = list(queryset_string[0])
        data_old_year = ast.literal_eval(data[0])
    except IndexError:
        data_old_year = None

    return data_old_year

def strToArrListOldYear(active, field, s_comp, current_year):
    try:
        queryset_string = None
        if field == "weight":
            queryset = StoneEstimateItem.objects.filter(se__company__code = active, se__created__year = current_year
                ).values_list('se__site__base_site_name', 'stone_type__base_stone_type_name'
                ).annotate(
                    A = Coalesce(Sum('total'), Value(0), output_field=models.DecimalField()),
                    B = Coalesce(Sum('id'), Value(0), output_field=models.DecimalField()),
                ).order_by('se__site', 'stone_type')
            queryset_string = transform_queryset(queryset, active)

        else:
            pd_data = Production.objects.filter(company__code=active, site__in=s_comp, created__year = current_year
            ).annotate(
                working_time=ExpressionWrapper(F('actual_time') - F('total_loss_time'), output_field=models.DurationField()),
                hour_per_day=ExpressionWrapper(F('actual_time') / (F('actual_time') - F('total_loss_time')), output_field=models.DecimalField()),
            ).values_list(
                'site__base_site_name'
            ).annotate(
                sum_run=Sum('run_time'),
                sum_total_working_time=Sum('working_time'),
                sum_hour_per_day=Sum('hour_per_day'),
                sum_capacity_per_hour=Sum('capacity_per_hour'),
            )
            if field == "prod_run":
                queryset_string = {
                    name: {
                        'A': f"{td_to_hours(td):,.2f}".replace('.', ':')
                    }
                    for name, td, _, _, _ in pd_data
                }
            elif field == "prod_work":
                queryset_string = {
                    name: {
                        'A': f"{td_to_hours(td):,.2f}".replace('.', ':')
                    }
                    for name, _, td, _, _ in pd_data
                }
            elif field == "prod_cap":
                queryset_string = {
                    name: {
                        'A': f"{td:,.2f}"
                    }
                    for name, _, _, td, _ in pd_data
                }
            elif field == "prod_hpd":
                queryset_string = {
                    name: {
                        'A': f"{td:,.2f}"
                    }
                    for name, _, _, td, _ in pd_data
                }
        data_old_year = queryset_string
    except IndexError:
        data_old_year = None

    return data_old_year

def clean_text(text):
    """ลบช่องว่างซ้อน"""
    return re.sub(r'\s+', ' ', text).strip()

def fmt(n):
    """format ตัวเลขเป็น 1,234"""
    return f"{int(n):,}" if n else '0'

def td_to_hours(td):
    return td.total_seconds() / 3600

def transform_queryset(queryset, active):
    result = defaultdict(dict)
    plant_order = {}# เก็บ id โรงโม่
    total_A = defaultdict(Decimal)
    total_B = defaultdict(Decimal)

    for plant, stone, a, plant_id in queryset:
        num_mount = StoneEstimateItem.objects.filter(
                se__company__code = active, se__created__year = '2025', se__site__base_site_name = plant, stone_type__base_stone_type_name = stone
                ).annotate(month=ExtractMonth('se__created')
                ).values('month'
                ).annotate(total=Count('id')
                ).order_by('month')
        
        plant = clean_text(plant)
        stone = clean_text(stone)

        a = a or Decimal('0')
        plant_id = int(plant_id)

        # เก็บ id โรงโม่ (ใช้เรียง)
        plant_order.setdefault(plant, plant_id)
        # ===== คำนวณค่า B (ปรับสูตรได้) =====
        b = a / Decimal(len(num_mount))

        result[plant][stone] = {
            'A': fmt(a),
            'B': fmt(b),
        }
        total_A[stone] += a
        total_B[stone] += b

    # ===== เรียงโรงโม่ตาม id =====
    sorted_result = OrderedDict(
        sorted(result.items(), key=lambda x: plant_order[x[0]])
    )
    # ===== เพิ่ม Total (อยู่ล่างสุด) =====
    sorted_result['Total'] = {}
    for stone in total_A:
        sorted_result['Total'][stone] = {
            'A': fmt(total_A[stone]),
            'B': fmt(total_B[stone]),
        }
    return sorted_result

def update_results(all_month_years, format,  dictionary, key1, key2, value):
    if value is None:
        value = Decimal(0)

    if key1 not in dictionary:
        dictionary[key1] = {}
    for my in all_month_years:
        if my not in dictionary[key1]:
            if format == 1:
                dictionary[key1][my] = timedelta(hours=0, minutes=0)
            elif format == 2:
                dictionary[key1][my] = Decimal(0)
    dictionary[key1][key2] = value

def getLossNameByMill(company_in, site, start_created, end_created, mode):
    list_loss = []
    pd_loss = ProductionLossItem.objects.filter(production__company__code__in = company_in, production__site = site, production__created__range=(start_created, end_created)).order_by('mc_type').values('production__site__base_site_id', 'mc_type__name').annotate(sum_time = Sum('loss_time'))
    pd_loss_all = ProductionLossItem.objects.filter(production__company__code__in = company_in, production__site = site, production__created__range=(start_created, end_created)).aggregate(s=Sum('loss_time'))["s"]
    for i in pd_loss:
        if mode == 1:
            list_loss.append(i['mc_type__name'])
        elif mode == 2:
            list_loss.append(calculatePersent(i['sum_time'], pd_loss_all))
    return list_loss

def calculatorDiffTime(request, start_time, end_time):
    difference = None
    if start_time is None:
        start_time = timedelta(hours=0, minutes=0) 
    if end_time is None:
        end_time = timedelta(hours=0, minutes=0)
    difference = end_time - start_time
    return difference

def calculatorDiffWorkRealTime(plan, uncontrol, loss):
    difference = None
    if plan is None:
        plan = timedelta(hours=0, minutes=0) 
    if uncontrol is None:
        uncontrol = timedelta(hours=0, minutes=0)
    if loss is None:
        loss = timedelta(hours=0, minutes=0)
    difference = plan - uncontrol - loss
    return difference

def calculatorDiff(request, val1, val2):
    difference = None
    if val1 is None:
        val1 = Decimal('0.0')
    if val2 is None:
        val2 = Decimal('0.0')
    difference = val2 - val1
    return difference

def setDurationTime(request, duration):
    result = None
    if duration is not None:
        if str(duration).startswith('0:'):
            _ , hours, minutes  = map(int, str(duration).split(':'))
        else:
            hours, minutes, _  = map(int, str(duration).split(':'))
        result = timedelta(hours=hours, minutes=minutes)

    return result

def searchProductionGoal(request):
    if 'site_id' in request.GET and 'line_type_id' in request.GET and 'created' in request.GET and 'pd_id' in request.GET and 'company' in request.GET:
        site_id = request.GET.get('site_id')
        line_type_id = request.GET.get('line_type_id')
        created =  request.GET.get('created')
        pd_id =  request.GET.get('pd_id')
        company =  request.GET.get('company')


        date_object = datetime.strptime(created, "%Y-%m-%d")

        #เอาออก line_type__id = line_type_id เพราะโรงโม่เดียวกันใช้เป้าผลิตเท่ากัน
        pd_goal = ProductionGoal.objects.filter(company__code = company, date__year = f'{date_object.year}' , date__month = f'{date_object.month}' , site = site_id).order_by('-id')[:1].values('site', 'line_type', 'date' , 'accumulated_goal', 'id')
        #if pd_id == '' create mode , else edit mode
        if pd_id == '':
            have_production = Production.objects.filter(company__code = company, created = created, site = site_id, line_type__id = line_type_id ).exists()
        else:
            have_production = Production.objects.filter(~Q(id = pd_id), company__code = company, created = created, site = site_id, line_type__id = line_type_id ).exists()
        #ดึงข้อมูล line 1 มาเพื่อไป set default ใน line อื่นๆ
        pd_line1 = Production.objects.filter(company__code = company, created = created, site = site_id, line_type__id = 1).values('plan_start_time', 'plan_end_time')
        
        
    data = {
        'pd_goal_list': list(pd_goal),
        'have_production' :have_production,
        'pd_line1': list(pd_line1),
    }
    
    return JsonResponse(data)

def createProduction(request):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    base_loss_type = BaseLossType.objects.all()

    ProductionLossItemFormSet = modelformset_factory(ProductionLossItem, fields=('mc_type', 'loss_type', 'loss_time'), extra=20, widgets={'loss_time': forms.TimeInput(format='%H:%M', attrs={'class':'form-control', 'type': 'time'}),})
    ProductionMachineItemFormSet = modelformset_factory(ProductionMachineItem, fields=('mc_type', 'mile_start', 'mile_end'), extra=10, widgets={})
    if request.method == 'POST':
        pd_goal_form = ProductionGoalForm(request.POST)
        production_form = ProductionForm(request, request.POST or None, initial={'company': company})
        formset = ProductionLossItemFormSet(request.POST, prefix='loss')
        mc_formset = ProductionMachineItemFormSet(request.POST, prefix='machine')
        if production_form.is_valid() and formset.is_valid() and pd_goal_form.is_valid() and mc_formset.is_valid():
            production = production_form.save()

            if pd_goal_form.cleaned_data['pk_goal']:
                pd_goal = ProductionGoal.objects.get(id = pd_goal_form.cleaned_data['pk_goal'])
                pd_goal.accumulated_goal = pd_goal_form.cleaned_data['accumulated_goal']
                pd_goal.company = production.company
                pd_goal.save()
            else:
                pd_goal = ProductionGoal.objects.create(accumulated_goal = pd_goal_form.cleaned_data['accumulated_goal'])
                pd_goal.site = production.site
                pd_goal.line_type = production.line_type
                pd_goal.date = production.created
                pd_goal.company = production.company
                pd_goal.save()

            production.pd_goal = pd_goal

            formset_instances = formset.save(commit=False)
            for form, instance in zip(formset.forms, formset_instances):
                mc_type = form.cleaned_data.get('mc_type')
                loss_type = form.cleaned_data.get('loss_type')
                loss_time = form.cleaned_data.get('loss_time')
                if mc_type and loss_type and loss_time:  # only save if mc_type is filled
                    instance.production = production
                    instance.save()

            mc_formset_instances = mc_formset.save(commit=False)
            for mc_form, mc_instance in zip(mc_formset.forms, mc_formset_instances):
                mc_type = mc_form.cleaned_data.get('mc_type')
                mile_start = mc_form.cleaned_data.get('mile_start')
                mile_end = mc_form.cleaned_data.get('mile_end')
                if mc_type and mile_start and mile_end:  # only save if mc_type is filled
                    mc_instance.production = production
                    mc_instance.save()

            #คำนวนเวลารวมในการสูญเสีย
            total_loss_time = ProductionLossItem.objects.filter(production = production).aggregate(s=Sum("loss_time"))["s"]
            production.total_loss_time = total_loss_time if total_loss_time else timedelta(hours=0, minutes=0)
            #คำนวนเวลารวมในการสูญเสีย uncontrol
            total_uncontrol_time = ProductionLossItem.objects.filter(production = production, mc_type = 7).aggregate(s=Sum("loss_time"))["s"]
            production.uncontrol_time = total_uncontrol_time if total_uncontrol_time else timedelta(hours=0, minutes=0)
            production.save()

            #คำนวน capacity per hour
            production.capacity_per_hour = calculatProductionCapacity(production.company, production.created, production.site, production.line_type)
            production.save()

            return redirect('viewProduction')
    else:
        production_form = ProductionForm(request, initial={'company': company})
        pd_goal_form = ProductionGoalForm(initial={'company': company})
        formset = ProductionLossItemFormSet(queryset=ProductionLossItem.objects.none(), prefix='loss')
        mc_formset = ProductionMachineItemFormSet(queryset=ProductionMachineItem.objects.none() , prefix='machine')

    context = {'production_page':'active', 'pd_goal_form': pd_goal_form, 'form': production_form, 'formset': formset, 'mc_formset': mc_formset, active :"active", 'disabledTab' : 'disabled'}
    return render(request, "production/createProduction.html",context)

def editProduction(request, pd_id):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    pd_data = Production.objects.get(id = pd_id)

    #หาบันทึกปฎิบัติการของวันนี้ เพื่อเช็คไม่ให้ save mill และ line ซ้ำกัน
    production_on_day = Production.objects.filter(~Q(id = pd_data.id), created = datetime.today()).values('site', 'line_type', 'created')

    if request.method == "POST":
        formset = ProductionLossItemInlineFormset(request.POST, request.FILES, instance=pd_data)
        mc_formset = ProductionMachineItemInlineFormset(request.POST, request.FILES, instance=pd_data)
        form = ProductionForm(request, request.POST, request.FILES, instance=pd_data)
        pd_goal_form = ProductionGoalForm(request.POST, request.FILES, instance=pd_data.pd_goal)

        if form.is_valid() and formset.is_valid() and pd_goal_form.is_valid() and mc_formset.is_valid():
            # save production
            production = form.save()

            #หา id production Goal ใหม่
            find_pd_goal = ProductionGoal.objects.filter(company__code = company, date__year = f'{production.created.year}', date__month = f'{production.created.month}', site = production.site).last()
            production.pd_goal = find_pd_goal

            # save ProductionLossItem
            instances = formset.save(commit=False)
            for instance in instances:
                instance.save()
            for obj in formset.deleted_objects:
                obj.delete()
            formset.save_m2m()

            # save ProductionMachineItem
            mc_instances = mc_formset.save(commit=False)
            for mc_instance in mc_instances:
                mc_instance.save()
            for mc_obj in mc_formset.deleted_objects:
                mc_obj.delete()
            mc_formset.save_m2m()

            #คำนวนเวลารวมในการสูญเสีย
            total_loss_time = ProductionLossItem.objects.filter(production = production).aggregate(s=Sum("loss_time"))["s"]
            production.total_loss_time = total_loss_time if total_loss_time else timedelta(hours=0, minutes=0)
            #คำนวนเวลารวมในการสูญเสีย uncontrol
            total_uncontrol_time = ProductionLossItem.objects.filter(production = production, mc_type = 7).aggregate(s=Sum("loss_time"))["s"]
            production.uncontrol_time = total_uncontrol_time if total_uncontrol_time else timedelta(hours=0, minutes=0)
            production.save()

            #update เป้าผลิตสะสม production Goal ใหม่
            pd_goal = ProductionGoal.objects.get(id = find_pd_goal.id)
            pd_goal.accumulated_goal = pd_goal_form.cleaned_data['accumulated_goal']
            pd_goal.save()

            updateProductionCapacity(production.company, production.created, production.site)#คำนวน capacity per hour

            return redirect('viewProduction')
    else:
        formset = ProductionLossItemInlineFormset(instance=pd_data)
        mc_formset = ProductionMachineItemInlineFormset(instance=pd_data)
        form = ProductionForm(request, instance=pd_data)
        pd_goal_form = ProductionGoalForm(instance=pd_data.pd_goal)

    context = {'production_page':'active', 'pd_goal_form': pd_goal_form, 'form': form, 'formset': formset, 'mc_formset': mc_formset, 'pd': pd_data, 'production_on_day': production_on_day, active :"active", 'disabledTab' : 'disabled'}
    return render(request, "production/editProduction.html",context)

def calculatProductionCapacity(company_id, date, site_id, line_type_id):
    result = Decimal('0.0')
    data_sum_produc = Weight.objects.filter(bws__company = company_id, site=site_id, date = date, bws__weight_type = 2).aggregate(s=Sum("weight_total"))["s"]
    wk_time = Production.objects.filter(company = company_id, site=site_id, line_type = line_type_id, created = date).annotate(working_time_de = ExpressionWrapper(F('actual_time') - F('total_loss_time') , output_field= models.DecimalField())).aggregate(total_working_time=Sum('working_time_de'))['total_working_time']
    
    if data_sum_produc and wk_time:
        result = data_sum_produc/(wk_time/1000000/3600)
    return result

def removeProduction(request, pd_id):
    pd = Production.objects.get(id = pd_id)

    #ลบ ProductionLossItem ใน Production ด้วย
    items = ProductionLossItem.objects.filter(production = pd)
    items.delete()

    #ลบ ProductionMachineItem ใน Production ด้วย
    mc_items = ProductionMachineItem.objects.filter(production = pd)
    mc_items.delete()
    
    #ลบ Production ทีหลัง
    pd.delete()
    return redirect('viewProduction')

#หาวันแรกของเดือนนี้
def startDateInMonth(day):
    dt = datetime.strptime(f"{day}", '%Y-%m-%d')
    result = dt.replace(day=1).date()
    return f"{result}"

#หาวันสุดท้ายของเดือนนี้
def endDateInMonth(day):
    dt = datetime.strptime(f"{day}", '%Y-%m-%d')
    day = calendar.monthrange(dt.year, dt.month)[1]
    result = dt.replace(day=day).date()
    return f"{result}"

def calculatCapacityPerHour(request, data_sum_produc, wk_time):
    result = Decimal('0.0')
    if data_sum_produc and wk_time:
        result = data_sum_produc/(wk_time/1000000/3600)
    return result

def formatHourMinute(time):
    result = None
    if time:
       #result = (datetime.min + time).strftime("%H:%M") or None
       result = f'{time}'[:-3]
    return result

def excelProductionAndLoss(request, my_q, sc_q):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    pd_sites = Production.objects.filter(my_q).values_list('site', flat=True).distinct()
    sites = BaseSite.objects.filter(base_site_id__in = pd_sites)

    sunday_fill = PatternFill(start_color="f59393", fill_type="solid") #ถ้าเป็นวันอาทิตย์ แถวเป็นสีแดง

    workbook = openpyxl.Workbook()
    if sites:
        for site in sites:
            #ดึงเวลาสูญเสีย
            count_loss = ProductionLossItem.objects.filter(sc_q, production__site = site).order_by('mc_type__id', 'loss_type__id').values('production__site__base_site_id', 'mc_type__name', 'loss_type__name').annotate(sum_time = Sum('loss_time'))
            #ดึงเวลาเครื่องจักร
            count_mc = ProductionMachineItem.objects.filter(sc_q, production__site = site).order_by('mc_type__id').values('mc_type__name').distinct()

            sheet = workbook.create_sheet(title=site.base_site_name)

            # Fetch distinct line types for the current mill
            line_types = Production.objects.filter(my_q, site=site).values_list('line_type', flat=True).distinct()

            line_type =  BaseLineType.objects.filter(id__in=line_types)

            # Create a list of colors for each line_type
            line_type_colors = [generate_pastel_color() for i  in range(len(line_type) + 1)]

            column_index = 2
            for line in line_type:
                sheet.cell(row=1, column = column_index, value = line.name)
                sheet.merge_cells(start_row=1, start_column = column_index, end_row=1, end_column= (column_index + len(count_loss) + (len(count_mc) * 3) + 17) -1 )
                sheet.cell(row=1, column=column_index).alignment = Alignment(horizontal='center')
                column_index += len(count_loss) + (len(count_mc) * 3) + 17

            headers2 = ['Date']
            for i in  range(len(line_type)):
                headers2.extend(['เป้าต่อวัน','เป้าสะสม(ตัน)', 'ชั่วโมงตามแผน', 'ชั่วโมงตามแผน', 'ชั่วโมงทำงาน','ชั่วโมงกำหนดจริง', 'ชั่วโมงกำหนดจริง', 'ชั่วโมงกำหนดจริง', 'ชั่วโมงเดินเครื่อง', 'ชั่วโมงเดินเครื่อง', 'ชั่วโมงเดินเครื่อง'])
                headers2.extend([cl['mc_type__name'] for cl in count_loss])
                headers2.extend(['รวมเวลาสูญเสีย','ชั่วโมงการทำงานจริง'])
                headers2.extend([
                    item
                    for cl in count_mc
                    for item in [
                        cl['mc_type__name'],
                        cl['mc_type__name'],
                        cl['mc_type__name']
                    ]
                ])            
                headers2.extend(['ยอดผลิต (ตัน)','ยอดผลิตสะสม','กำลังการผลิต (ตัน/ชั่วโมง)','หมายเหตุ',])

            sheet.append(headers2)

            merge_cells_num = 0
            headers3 = ['Date']
            for i in  range(len(line_type)):
                headers3.extend(['เป้าต่อวัน','เป้าสะสม(ตัน)', '(เริ่ม)', '(สิ้นสุด)', 'ชั่วโมงทำงาน', '(เริ่ม)', '(สิ้นสุด)', 'ชั่วโมงกำหนดจริง', '(เริ่ม)', '(สิ้นสุด)', 'ชั่วโมงเดินเครื่อง'])
                headers3.extend([cl['loss_type__name'] for cl in count_loss])
                headers3.extend(['รวมเวลาสูญเสีย','ชั่วโมงการทำงานจริง'])
                headers3.extend([
                    item
                    for cl in count_mc
                    for item in [
                        'เลขไมล์เริ่มต้น ' + cl['mc_type__name'],
                        'เลขไมล์สิ้นสุด ' + cl['mc_type__name'],
                        'เวลาเดินเครื่อง ' + cl['mc_type__name']
                    ]
                ])
                headers3.extend(['ยอดผลิต (ตัน)','ยอดผลิตสะสม','กำลังการผลิต (ตัน/ชั่วโมง)','หมายเหตุ',])

                # merge_cells headers เป้าต่อวัน, เป้าสะสม(ตัน),ชั่วโมงทำงาน,ชั่วโมงเดินเครื่อง
                sheet.merge_cells(start_row=2, start_column = 2 + merge_cells_num , end_row=3, end_column = 2 + merge_cells_num)
                sheet.merge_cells(start_row=2, start_column = 3 + merge_cells_num , end_row=3, end_column = 3 + merge_cells_num)
                sheet.merge_cells(start_row=2, start_column = 6 + merge_cells_num , end_row=3, end_column = 6 + merge_cells_num)
                sheet.merge_cells(start_row=2, start_column = 9 + merge_cells_num , end_row=3, end_column = 9 + merge_cells_num)
                sheet.merge_cells(start_row=2, start_column = 12 + merge_cells_num , end_row=3, end_column = 12 + merge_cells_num)
                sheet.merge_cells(start_row = 2, start_column = 4 + merge_cells_num , end_row = 2, end_column = 5 + merge_cells_num)
                sheet.merge_cells(start_row = 2, start_column = 7 + merge_cells_num , end_row = 2, end_column = 8 + merge_cells_num)
                sheet.merge_cells(start_row = 2, start_column = 10 + merge_cells_num , end_row = 2, end_column = 11 + merge_cells_num)

                #ช่องหลังจาก loos_item
                sheet.merge_cells(start_row=2, start_column = 13 + merge_cells_num + len(count_loss) , end_row=3, end_column = 13 + merge_cells_num + len(count_loss))
                sheet.merge_cells(start_row=2, start_column = 14 + merge_cells_num + len(count_loss) , end_row=3, end_column = 14 + merge_cells_num + len(count_loss))

                #ช่องหลังจาก loos_item และ mc_item
                sheet.merge_cells(start_row=2, start_column = 15 + merge_cells_num + len(count_loss) + (len(count_mc) * 3) , end_row=3, end_column = 15 + merge_cells_num + len(count_loss)+ (len(count_mc) * 3))            
                sheet.merge_cells(start_row=2, start_column = 16 + merge_cells_num + len(count_loss) + (len(count_mc) * 3), end_row=3, end_column = 16 + merge_cells_num + len(count_loss)+ (len(count_mc) * 3))
                sheet.merge_cells(start_row=2, start_column = 17 + merge_cells_num + len(count_loss) + (len(count_mc) * 3), end_row=3, end_column = 17 + merge_cells_num + len(count_loss)+ (len(count_mc) * 3))
                sheet.merge_cells(start_row=2, start_column = 18 + merge_cells_num + len(count_loss) + (len(count_mc) * 3), end_row=3, end_column = 18 + merge_cells_num + len(count_loss)+ (len(count_mc) * 3))

                merge_cells_num += len(count_loss) + (len(count_mc) * 3) + 17

            sheet.cell(row=1, column = 1, value = 'วัน/เดือน/ปี')
            sheet.merge_cells(start_row=1, start_column = 1, end_row=3, end_column=1)
            sheet.append(headers3)

            # Fetch distinct 'created' dates for the current mill
            created_dates = Production.objects.filter(my_q, site=site).values_list('created', flat=True).order_by('created').distinct()
            first_date = created_dates.first()
            pd_year = datetime.strptime(str(first_date), '%Y-%m-%d').year

            for created_date in created_dates:
                row = [created_date]
                row_sum = ['']
                row_persent_loss = ['']
                row_persent_accumulated_produc = ['']
                sum_capacity_per_hour = Decimal('0.0')
                
                date_from_accumulated = startDateInMonth(created_date)

                for line_type in BaseLineType.objects.filter(id__in=line_types):
                    production = Production.objects.filter(site = site, line_type = line_type, created = created_date).annotate( count=Count('site__base_site_id') 
                        , sum_loss = Sum('total_loss_time'), sum_actual = Sum('actual_time'), sum_uncontrol=Sum(Case(When(uncontrol_time__isnull=True, then=Value(timedelta(0))), default='uncontrol_time', output_field = models.DurationField()))
                        , sum_loss_n_un = ExpressionWrapper(F('sum_loss') - F('sum_uncontrol'), output_field= models.DurationField()) 
                        , sum_work_real =  ExpressionWrapper(F('sum_actual') - F('sum_loss'), output_field= models.DurationField()) ).first()

                    accumulated_goal = Production.objects.filter(site = site, line_type = line_type, created__range=(date_from_accumulated, created_date)).aggregate(s=Sum("goal"))["s"]

                    data_sum_produc = Weight.objects.filter(site=site, date = created_date, bws__weight_type = 2).aggregate(s=Sum("weight_total"))["s"]
                    wk_time = Production.objects.filter(site=site, line_type = line_type, created = created_date).annotate(working_time_de = ExpressionWrapper(F('actual_time') - F('total_loss_time') , output_field= models.DecimalField())).aggregate(total_working_time=Sum('working_time_de'))['total_working_time']

                    accumulated_produc = Weight.objects.filter(site=site ,date__range=(date_from_accumulated, created_date) , bws__weight_type = 2).aggregate(s=Sum("weight_total"))["s"]

                    #3) sum_by_mill = Production.objects.filter(my_q, site=site, line_type = line_type).distinct().aggregate(Sum('plan_time'),Sum('run_time'),Sum('total_loss_time'))
                    #4) cal_by_mill = Production.objects.filter(my_q, site=site, line_type = line_type).distinct().annotate(working_time = ExpressionWrapper(F('run_time') - F('total_loss_time'), output_field= models.DurationField())).aggregate(total_working_time=Sum('working_time'))['total_working_time']

                    capacity_per_hour = (
                        production.capacity_per_hour 
                        if pd_year > 2024 and production is not None 
                        else calculatCapacityPerHour(request, data_sum_produc, wk_time)
                    ) #if ปี capacity_per_hour > 2024 ดึงแบบใหม่
                        
                    if production:
                        row.extend([production.goal, accumulated_goal , formatHourMinute(production.plan_start_time), formatHourMinute(production.plan_end_time), formatHourMinute(production.plan_time), formatHourMinute(production.actual_start_time), formatHourMinute(production.actual_end_time) , formatHourMinute(production.actual_time), formatHourMinute(production.run_start_time) if production.run_start_time else production.mile_run_start_time  , formatHourMinute(production.run_end_time) if production.run_end_time else production.mile_run_end_time, formatHourMinute(production.run_time)])
                    else:
                        row.extend(['' for i in range(17)])

                    if  count_loss:
                        for i in range(len(count_loss)):
                            tmp_mc = sheet.cell(row=2, column = i+13).value
                            tmp_loss = sheet.cell(row=3, column = i+13).value
                            lss = ProductionLossItem.objects.filter(production = production, mc_type__name = tmp_mc, loss_type__name = tmp_loss).aggregate(s=Sum('loss_time'))['s']
                            if lss:
                                row.extend([formatHourMinute(lss)])
                            else:
                                row.extend(['-'])
                    else:
                        row.extend(['-' for i in range(len(count_loss))])

                    #รวม, ชั่วโมงการทำงานจริง
                    if  production:
                        row.extend([formatHourMinute(production.sum_loss_n_un), formatHourMinute(production.sum_work_real)])

                    if  count_mc:
                        for i in range(len(count_mc)):
                            tmp_mc = sheet.cell(row=2, column = 15 + len(count_loss) + (i * 3)).value
                            mcc = (
                                ProductionMachineItem.objects
                                .filter(production=production, mc_type__name=tmp_mc)
                                .values('mile_start', 'mile_end', 'diff_time')
                                .first()
                            )

                            if mcc:
                                row.extend([mcc['mile_start'],mcc['mile_end'],formatHourMinute(mcc['diff_time'])])
                            else:
                                row.extend(['-', '-', '-'])
                    else:
                        row.extend(['-' for i in range(len(count_mc) * 3)])
                

                    # ยอดผลิต (ตัน), ยอดผลิตสะสม, กำลังการผลิต (ตัน/ชั่วโมง), หมายเหตุ
                    if  production:
                        row.extend([data_sum_produc, accumulated_produc, capacity_per_hour, production.note,])
                        sum_capacity_per_hour += capacity_per_hour

                    ''' 1) ล่างสุดตัวหนังสือสีแดง sum ทั้งหมด
                    row_sum.extend([len(created_dates), '' , '', 'ชั่วโมงทำงานรวม', formatHourMinute(sum_by_mill['plan_time__sum']), '', '', formatHourMinute(sum_by_mill['run_time__sum'])])
                    row_sum.extend(['eiei'+formatHourMinute(pd_loos_item['sum_loss_time']) for pd_loos_item in ProductionLossItem.objects.filter(production__site=site, production__line_type = line_type).order_by('mc_type__id', 'loss_type__id').values('loss_type__id').annotate(sum_loss_time=Coalesce(Sum('loss_time'), None))])

                    row_sum.extend([formatHourMinute(sum_by_mill['total_loss_time__sum']), formatHourMinute(cal_by_mill), 'diff จากเป้า' , calculatorDiff(request, accumulated_goal , accumulated_produc) , sum_capacity_per_hour/len(created_dates),''])

                    loss_items = ProductionLossItem.objects.filter(
                        production__site=site,
                        production__line_type=line_type
                    ).order_by('loss_type__id').values('loss_type__id').annotate(
                        sum_loss_time=Coalesce(Sum('loss_time'), None)
                    )

                    row_persent_accumulated_produc.extend(['', '' , '', '', '', '', '', ''])
                    row_persent_accumulated_produc.extend(['C' for i in range(len(count_loss))])
                    row_persent_accumulated_produc.extend(['', '', '' , str(round(calculatorDiff(request, accumulated_goal , accumulated_produc) / accumulated_goal, 2)) + "%" if accumulated_goal and accumulated_produc else None , '',''])

                    row_persent_loss.extend(['', '' , '', '', '', '', '% ชม.สูญเสีย ต่อ ชม.ทำงานจริง', ''])
                    row_persent_loss.extend([str(round(pd_loos_item['sum_loss_time'] / sum_by_mill['total_loss_time__sum'] * 100, 2)) + "%" if pd_loos_item['sum_loss_time'] else None for pd_loos_item in loss_items])
                    row_persent_loss.extend(['100%', '', '' , '' , '',''])                
                    '''


                sheet.append(row)

                #ถ้าเป็นวันอาทิตย์ แถวเป็นสีแดง
                if created_date.weekday() == 6:  #วันอาทิตย์
                    for cell in sheet[sheet.max_row]:
                        cell.fill = sunday_fill

            if len(created_dates) > 0:
                sheet.append(row_sum)
                sheet.append(row_persent_accumulated_produc)
                sheet.append(row_persent_loss)
                # 2) ล่างสุดตัวหนังสือสีแดง sum ทั้งหมด sheet.cell(row = len(created_dates) + 4, column = 1, value = f'จำนวนวันทำงาน' )        

            # Set column width and border for all columns
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)  # Get the letter of the current column
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass

                adjusted_width = (max_length + 2) * 1.2  # Adjust the width based on content length
                sheet.column_dimensions[column_letter].width = adjusted_width
                
                # Set border for each cell in the column
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for cell in column:
                    cell.alignment = Alignment(horizontal='center')

                    # ชั่วโมงการทำงานจริง line 1
                    if len(line_types) > 0 and cell.column == len(count_loss)  + 14:
                        cell.font = Font(color="0000FF")  # Blue text
                    # ชั่วโมงการทำงานจริง line 2
                    if  len(line_types) > 1 and cell.column == (len(count_loss) * 2) + (len(count_mc) * 3) + 31:
                        cell.font = Font(color="0000FF")  # Blue text
                    # ชั่วโมงการทำงานจริง line 3
                    if  len(line_types) > 2 and cell.column == (len(count_loss) * 3)+  ((len(count_mc) * 3) * 2) + 48:
                        cell.font = Font(color="0000FF")  # Blue text

                    # 2 row สุดท้าย ไม่ใส่ border และ set ตัวหนังสือสีแดง
                    if cell.row > sheet.max_row - 3:
                        cell.font = Font(color="FF0000")
                    else:
                        cell.border = border
            
            column_index = 2
            for line_index, line in enumerate(line_types):
                # Set the background color for the current line_type
                fill = PatternFill(start_color=line_type_colors[line_index % len(line_type_colors)], fill_type="solid")
                sheet.cell(row=1, column=column_index).fill = fill
                column_index += len(count_loss) + (len(count_mc) * 3) + 17

            for row in sheet.iter_rows(min_row=1, max_row=3):
                # Set the background color for each cell in the column
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    line_index = (cell.column - 2) // (len(count_loss) + (len(count_mc) * 3) + 17)
                    fill_color = line_type_colors[line_index % len(line_type_colors)]
                    fill = PatternFill(start_color=fill_color, fill_type="solid")
                    cell.fill = fill

            #1) merge_cells loos header mc_type
            num_loss = 0
            for col_index in range(12, sheet.max_column):
                if sheet.cell(row=2, column=col_index).value == sheet.cell(row=2, column=col_index + 1).value:
                    num_loss += 1
                else:
                    sheet.merge_cells(start_row=2, start_column=col_index - num_loss, end_row=2, end_column=col_index)
                    num_loss = 0
            
            sheet.freeze_panes = "B4" #freeze

        workbook.remove(workbook['Sheet'])
    else:
        worksheet = workbook.active
        worksheet.cell(row = 1, column = 1, value = f'ไม่มีข้อมูลบันทึกปฎิบัติการโรงโม่ดือนนี้')

    # Save workbook into memory
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    size = output.getbuffer().nbytes

    # Generator to stream file in chunks
    def file_iterator(buffer, chunk_size=8192):
        while True:
            data = buffer.read(chunk_size)
            if not data:
                break
            yield data

    response = StreamingHttpResponse(
        file_iterator(output),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="production_record_({active}).xlsx"'
    response["Content-Length"] = str(size)
    return response

def exportExcelProductionAndLoss(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None
    site = request.GET.get('site') or None

    date_object = datetime.today()
    previous_date_time = date_object - timedelta(days=1)

    if end_created is None:
        end_created = previous_date_time.strftime('%Y-%m-%d')
    if start_created is None:
        start_created = startDateInMonth(end_created)

    my_q = Q()
    if start_created is not None:
        my_q &= Q(created__gte = start_created)
    if end_created is not None:
        my_q &=Q(created__lte = end_created)
    if site is not None:
        my_q &=Q(site = site)
    my_q &=Q(company__code__in = company_in)

    sc_q = Q()
    if start_created is not None:
        sc_q &= Q(production__created__gte = start_created)
    if end_created is not None:
        sc_q &=Q(production__created__lte = end_created)
    sc_q &=Q(production__company__code__in = company_in)
    
    response = excelProductionAndLoss(request, my_q, sc_q)
    return response

def exportExcelProductionAndLossDashboard(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    '''แบบเก่าดึง บันทึกปฏิบัติการงานโรงโม่ ตามเดือนนั้นๆ 09/05/2024
    #ดึงรายงานของเดือนนั้นๆ
    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)


    end_created = previous_date_time.strftime('%Y-%m-%d')
    start_created = startDateInMonth(end_created)
    '''
    end_created = request.session['db_end_date']
    start_created = request.session['db_start_date']

    my_q = Q()
    if start_created is not None:
        my_q &= Q(created__gte = start_created)
    if end_created is not None:
        my_q &=Q(created__lte = end_created)
    my_q &=Q(company__code__in = company_in)

    sc_q = Q()
    if start_created is not None:
        sc_q &= Q(production__created__gte = start_created)
    if end_created is not None:
        sc_q &=Q(production__created__lte = end_created)
    sc_q &=Q(production__company__code__in = company_in)
    
    response = excelProductionAndLoss(request, my_q, sc_q)
    return response

@login_required(login_url='login')
def viewStoneEstimate(request):
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')

    data = StoneEstimate.objects.filter(company__code__in = company_in).order_by('-created', 'site')

    #กรองข้อมูล
    myFilter = StoneEstimateFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    stone_estimate = p.get_page(page)

    context = {'stone_estimate_page':'active', 'stone_estimate': stone_estimate,'filter':myFilter, active :"active",}
    return render(request, "stoneEstimate/viewStoneEstimate.html",context)

def calculateSumEstimateByCompany(created, company, site_id, stone_type_id):
    w = Decimal('0.0')
    se_item = StoneEstimateItem.objects.filter(se__company = company, se__created = created, se__site__base_site_id = site_id, stone_type = stone_type_id).values('se__created','percent')
    for i in se_item:
        crush = Weight.objects.filter(bws__company = company, site = site_id, bws__weight_type = 2 , date = i['se__created']).aggregate(s = Sum("weight_total"))["s"] or Decimal('0.0')
        w += calculateEstimate(i['percent'], crush)
    return w

def createStoneEstimate(request):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    site_qs = BaseSite.objects.filter(weight_type = 2, s_comp__code = active)
    SITE_CHOICES = [('', '---------')] + [(str(site.base_site_id), site.base_site_name) for site in site_qs]
    ND_SITE_CHOICES = [('', '---------')] + [(str(site.base_site_id), site.base_site_name) for site in site_qs]

    base_stone_type = BaseStoneType.objects.filter(is_stone_estimate = True)
    StoneEstimateItemFormSet = modelformset_factory(StoneEstimateItem, fields=('stone_type', 'percent', 'qty', 'site_id', 'qty_site', 'nd_site_id', 'nd_qty_site', 'total'), extra=len(base_stone_type), 
        widgets={
        'site_id': Select(choices=SITE_CHOICES),
        'nd_site_id': Select(choices=ND_SITE_CHOICES)
    })

    if request.method == 'POST':
        se_form = StoneEstimateForm(request, request.POST)
        formset = StoneEstimateItemFormSet(request.POST)
        if se_form.is_valid() and formset.is_valid():
            se = se_form.save()

            formset_instances = formset.save(commit=False)
            for instance in formset_instances:
                instance.se = se
                instance.save()

            updatePassOtherEstimate(se.company.id, se.created)
            updateProdStockStoneItem(se.company.id, se.created)#คำนวณ stock
            return redirect('viewStoneEstimate')
    else:
        initial_data = (
            [{'qty_site': 0.0} for _ in range(len(base_stone_type))] +
            [{'nd_qty_site': 0.0} for _ in range(len(base_stone_type))]
        )

        se_form = StoneEstimateForm(request, initial={'company': company})
        formset = StoneEstimateItemFormSet(queryset=StoneEstimateItem.objects.none(), initial=initial_data)

    context = {'stone_estimate_page':'active', 'se_form': se_form, 'formset' : formset, 'base_stone_type': base_stone_type, active :"active", 'disabledTab' : 'disabled'}
    return render(request, "stoneEstimate/createStoneEstimate.html",context)

def editStoneEstimate(request, se_id):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    se_data = StoneEstimate.objects.get(id = se_id)
    original_site = se_data.site.base_site_id

    estimate = get_object_or_404(StoneEstimate, pk=se_id)
    FormsetClass = StoneEstimateItemInlineFormset

    if request.method == "POST":
        formset = FormsetClass(request.POST, instance=estimate, form_kwargs={'company_code': request.session['company_code']})
        se_form = StoneEstimateForm(request, request.POST, request.FILES, instance=estimate)
        
        if se_form.is_valid() and formset.is_valid():
            se = se_form.save(commit=False)
            se.save()

            # Save related items
            instances = formset.save(commit=False)
            for instance in instances:
                if instance.stone_type:
                    instance.save()

            for obj in formset.deleted_objects:
                obj.delete()
            formset.save_m2m()

            updatePassOtherEstimate(se.company.id, se.created)
            updateProdStockStoneItem(se.company.id, se.created)

            return redirect('viewStoneEstimate')
    else:
        formset = FormsetClass(instance=estimate, form_kwargs={'company_code': request.session['company_code']})
        se_form = StoneEstimateForm(request, instance=se_data)

    context = {'stone_estimate_page':'active', 'se_form': se_form, 'formset' : formset,'se': se_data, active :"active", 'disabledTab' : 'disabled'}
    return render(request, "stoneEstimate/editStoneEstimate.html",context)

def removeStoneEstimate(request, se_id):
    se = StoneEstimate.objects.get(id = se_id)
    tmp_company = se.company.id
    tmp_created = se.created
    tmp_site = se.site.base_site_id
    #ลบ StoneEstimateItem ใน StoneEstimate ด้วย
    items = StoneEstimateItem.objects.filter(se = se)
    items.delete()
    #ลบ StoneEstimate ทีหลัง
    se.delete()

    updateSumEstimateItem(tmp_company, tmp_created, tmp_site)#กรณีแก้ไข update total StoneEstimateItem
    updateProdStockStoneItem(tmp_company, tmp_created)#คำนวณ stock
    
    return redirect('viewStoneEstimate')

#update Estimate จากตาชั่ง
def updatePassScaleEstimate(company, created, site):
    try:
        sum_weight = Weight.objects.filter(date = created, bws__weight_type = 2, bws__company = company, site__base_site_id = site).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
        es = StoneEstimate.objects.filter(site = site, created = created, company = company).first()
        if es:
            es.scale = sum_weight
            es.total = es.topup + es.other + es.scale
            es.save()
            updateEstimate(company, site, created)
    except StoneEstimate.DoesNotExist:
        pass

#update Estimate จากโรงโม่อื่น
def updatePassOtherEstimate(company, created):
    all_site =  StoneEstimateItem.objects.filter(se__company = company, se__created = created, site_id__isnull = False).values('site_id').distinct()
    for i in all_site:
        try:
            #ดึงข้อมูล qty ทั้งหมด
            sum_qty_site = StoneEstimateItem.objects.filter(site_id = i['site_id'], se__created = created, se__company = company).aggregate(s=Sum("qty_site"))["s"] or Decimal('0.0')
            sum_nd_qty_site = StoneEstimateItem.objects.filter(nd_site_id = i['site_id'], se__created = created, se__company = company).aggregate(s=Sum("nd_qty_site"))["s"] or Decimal('0.0')
            sum_all = sum_qty_site + sum_nd_qty_site

            #ดึงน้ำหนักจากตาชั่งผลิตทั้งหมด
            sum_weight = Weight.objects.filter(date = created, bws__weight_type = 2, bws__company = company, site__base_site_id = i['site_id']).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
            
            es = StoneEstimate.objects.filter(site = i['site_id'], created = created, company = company).first()
            if es:
                es.other = sum_all
                es.scale = sum_weight
                es.total = es.topup + es.other + es.scale
                es.save()
                updateEstimate(company, i['site_id'], created)
        except StoneEstimate.DoesNotExist:
            pass

    nd_all_site =  StoneEstimateItem.objects.filter(se__company = company, se__created = created, nd_site_id__isnull = False).values('nd_site_id').distinct()
    for i in nd_all_site:
        try:
            #ดึงข้อมูล qty ทั้งหมด
            sum_qty_site = StoneEstimateItem.objects.filter(site_id = i['nd_site_id'], se__created = created, se__company = company).aggregate(s=Sum("qty_site"))["s"] or Decimal('0.0')
            sum_nd_qty_site = StoneEstimateItem.objects.filter(nd_site_id = i['nd_site_id'], se__created = created, se__company = company).aggregate(s=Sum("nd_qty_site"))["s"] or Decimal('0.0')
            sum_all = sum_qty_site + sum_nd_qty_site

            #ดึงน้ำหนักจากตาชั่งผลิตทั้งหมด
            sum_weight = Weight.objects.filter(date = created, bws__weight_type = 2, bws__company = company, site__base_site_id = i['nd_site_id']).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')

            es = StoneEstimate.objects.filter(site = i['nd_site_id'], created = created, company = company).first()
            if es:
                es.other = sum_all
                es.scale = sum_weight
                es.total = es.topup + es.other + es.scale
                es.save()
                updateEstimate(company, i['nd_site_id'], created)
        except StoneEstimate.DoesNotExist:
            pass

#คำนวน total จาก percent -> qty , qty_site
def updateEstimate(company, site, created):
    all_total = StoneEstimate.objects.filter(
        company=company, created=created, site=site
    ).values_list('total', flat=True).first()
        
    se_item = StoneEstimateItem.objects.filter(
        se__company=company,
        se__created=created,
        se__site__base_site_id=site
    ).values('id', 'percent', 'qty_site')

    for i in se_item:
        qty = calculateEstimate(i['percent'], Decimal(all_total))
        if i['qty_site']:
            total = qty - i['qty_site']
        else:
            total = qty
        StoneEstimateItem.objects.filter(id=i['id']).update(qty=qty, total=total)

def searchStoneEstimate(request):
    if 'site_id' in request.GET and 'created' in request.GET and 'se_id' in request.GET and 'company' in request.GET:
        site_id = request.GET.get('site_id')
        created =  request.GET.get('created')
        se_id =  request.GET.get('se_id')
        company =  request.GET.get('company')

        #if se_id == '' create mode , else edit mode
        if se_id == '':
            have_estimate = StoneEstimate.objects.filter(company__code = company, created = created, site = site_id).exists()
        else:
            have_estimate = StoneEstimate.objects.filter(~Q(id = se_id),company__code = company, created = created, site = site_id).exists()
        #ดึงเปอร์เซ็นคำนวนหินเปอร์ที่คีย์ไปล่าสุด
        last_se = StoneEstimate.objects.filter(company__code = company, site = site_id).order_by('-created').first()
        last_se_item = StoneEstimateItem.objects.filter(se = last_se).values('stone_type', 'percent')
        
    data = {
        'have_estimate' :have_estimate,
        'last_se_item': list(last_se_item),
    }
    
    return JsonResponse(data)

def calculateEstimate(percent, sum_all):
    result = Decimal(0.0)
    if percent and sum_all:
        result = Decimal(sum_all) * Decimal(percent)/100
    return result

def calculateEstimateToString(percent, sum_all):
    #result = Decimal(0.0) ลดการปริ้น 0
    result = None
    if percent and sum_all:
        result = Decimal(sum_all) * Decimal(percent)/100
        result = f"{result:.2f}"
    return result

def calculateSumEstimateToString(stone_type, site, customer_name, list_date, time_in, time_out):
    result = Decimal(0)
    tmp = Decimal(0)

    for ld in list_date:
        try:
            percent = StoneEstimateItem.objects.get(se__created = ld, se__site = site, stone_type = stone_type).percent
        except:
            percent = None
        sum_all = Weight.objects.filter( Q(time_out__gte=time_in) & Q(time_out__lte=time_out) , bws__weight_type = 2 , date = ld, customer_name = customer_name, site = site).aggregate(s_weight = Sum("weight_total"))['s_weight']
        
        if percent and sum_all:
            tmp = Decimal(sum_all) * Decimal(percent)/100
            result = result + tmp
    return f"{result:.2f}" if result != 0.00 else None

def exportExcelStoneEstimateAndProduction(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None

    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    tmp_end_created = previous_date_time.strftime('%Y-%m-%d')
    tmp_start_created = startDateInMonth(tmp_end_created)

    if start_created is None:
        start_created = tmp_start_created
    if end_created is None:
        end_created = tmp_end_created

    my_q = Q()
    if start_created is not None:
        my_q &= Q(created__gte = start_created)
    if end_created is not None:
        my_q &=Q(created__lte = end_created)
    my_q &=Q(company__code__in = company_in)

    sc_q = Q()
    if start_created is not None:
        sc_q &= Q(date__gte = start_created)
    if end_created is not None:
        sc_q &=Q(date__lte = end_created)
    sc_q &=Q(bws__company__code__in = company_in)
    
    response = excelStoneEstimateAndProduction(request, my_q, sc_q)
    return response

def exportExcelStoneEstimateAndProductionDashboard(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    ''' แบบเก่าดึง รายงานผลิตแยกผู้รับเหมาและชนิดหิน ตามเดือนนั้นๆ 09/05/2024
    #ดึงรายงานของเดือนนั้นๆ
    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    end_created = previous_date_time.strftime('%Y-%m-%d')
    start_created = startDateInMonth(end_created)
    '''
    
    end_created = request.session['db_end_date']
    start_created = request.session['db_start_date']

    my_q = Q()
    if start_created is not None:
        my_q &= Q(created__gte = start_created)
    if end_created is not None:
        my_q &=Q(created__lte = end_created)
    my_q &=Q(company__code__in = company_in)

    sc_q = Q()
    if start_created is not None:
        sc_q &= Q(date__gte = start_created)
    if end_created is not None:
        sc_q &=Q(date__lte = end_created)
    sc_q &=Q(bws__company__code__in = company_in)
    
    response = excelStoneEstimateAndProduction(request, my_q, sc_q)
    return response


def excelStoneEstimateAndProduction(request, my_q, sc_q):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    date_style = NamedStyle(name='custom_datetime', number_format='DD/MM/YYYY')

    se_site = StoneEstimate.objects.filter(my_q).values_list('site',flat=True).distinct()
    sites = BaseSite.objects.filter(base_site_id__in = se_site)

    se_id = StoneEstimate.objects.filter(my_q).values_list('id',flat=True).distinct()
    #ดึงเฉพาะชนิดหิน estimate
    base_stone_type = StoneEstimateItem.objects.select_related('stone_type').filter(se__in = se_id).order_by('stone_type').values_list('stone_type__base_stone_type_name', flat=True).distinct()

    #list_customer_name = ['สมัย','วีระวุฒิ','NCK']
    #เปลี่ยนการตั้งค่า 13-02-2024 list_customer_name = BaseCustomer.objects.filter(is_stone_estimate = True).values_list('customer_name', flat=True)
    list_customer_name = BaseSEC.objects.filter(company__code__in = company_in).values_list('customer__customer_name', flat=True)

    workbook = openpyxl.Workbook()
    if sites:
        for site in sites:
            sheet = workbook.create_sheet(title=site.base_site_name)

            list_time = BaseTimeEstimate.objects.filter(site = site).values('time_from', 'time_to', 'time_name')
            #ดึงชนิดหินที่มีคำว่าเข้าโม่
            mill_type = Weight.objects.filter(sc_q, bws__weight_type = 2, site = site).order_by('mill_name').values_list('mill_name', flat=True).distinct()

            tmp_stock_name = "กองสต็อค" + site.base_site_name
            try:
                stock_type = BaseSite.objects.get(base_site_name = tmp_stock_name)
                stock_type_name = stock_type.base_site_name
            except:
                stock_type_name = "ไม่มีกองสต็อค"

            #weight_stone_type = BaseStoneType.objects.filter(base_stone_type_name__in=weight_stone_types)

            column_index = 2
            sheet.cell(row=1, column = column_index, value = "พนักงาน")
            sheet.merge_cells(start_row=1, start_column = column_index, end_row=2, end_column= (column_index + 2) -1 )
            sheet.cell(row=1, column=column_index).alignment = Alignment(horizontal='center')

            column_index += 2
            sheet.cell(row=1, column = column_index, value = "ชม.ทำงาน")
            sheet.merge_cells(start_row=1, start_column = column_index, end_row=2, end_column= column_index)
            sheet.cell(row=1, column=column_index).alignment = Alignment(horizontal='center')

            column_index += 1
            sheet.cell(row=1, column = column_index, value = "หินเขา")
            sheet.cell(row=1, column=column_index).alignment = Alignment(horizontal='center')

            column_index += 1
            for mt in mill_type:
                sheet.cell(row=1, column = column_index, value = mt)
                sheet.merge_cells(start_row=1, start_column = column_index, end_row=1, end_column= (column_index + 2) -1 )
                sheet.cell(row=1, column=column_index).alignment = Alignment(horizontal='center')
                column_index += 2

            sheet.cell(row=1, column = column_index, value = stock_type_name)
            sheet.merge_cells(start_row=1, start_column = column_index, end_row=1, end_column= (column_index + 2) -1 )
            sheet.cell(row=1, column=column_index).alignment = Alignment(horizontal='center')
            column_index += 2

            sheet.cell(row=1, column = column_index, value = "หินเข้าโม่ทั้งหมด")
            sheet.merge_cells(start_row=1, start_column = column_index, end_row=1, end_column= (column_index + 2) -1 )
            sheet.cell(row=1, column=column_index).alignment = Alignment(horizontal='center')

            column_index += 2
            sheet.cell(row=1, column = column_index, value = "สต็อกคงเหลือ")

            column_index += 1
            sheet.cell(row=1, column = column_index, value = 'ประเภทหินตัน')
            sheet.merge_cells(start_row=1, start_column = column_index, end_row=1, end_column= (column_index + len(base_stone_type)) -1 )
            sheet.cell(row=1, column=column_index).alignment = Alignment(horizontal='center')
            column_index += len(base_stone_type)

            sheet.cell(row=1, column = column_index, value = 'หินเข้าโม่รวม(ตัน)')
            sheet.merge_cells(start_row=1, start_column = column_index, end_row=2, end_column= column_index)
            sheet.cell(row=1, column=column_index).alignment = Alignment(horizontal='center')


            headers2 = ['Date','พนักงาน', 'กะ', 'ชม.ทำงาน', 'ที่ผลิตได้',]
            for i in range(len(mill_type) + 1):
                headers2.extend(['เที่ยว','ตัน',])

            headers2.extend(['เที่ยว','ตัน',])

            headers2.extend(['AAA'])
            headers2.extend([i for i in base_stone_type])
            headers2.extend(['หินเข้าโม่รวม(ตัน)', 'ผลิตตัน/ชม.', 'หมายเหตุ'])

            sheet.cell(row=1, column = 1, value = 'วัน/เดือน/ปี')
            #merge_cells วัน/เดือน/ปี
            sheet.merge_cells(start_row=1, start_column = 1, end_row=2, end_column=1)
            sheet.append(headers2)

            # Fetch distinct 'created' dates for the current site
            created_dates = StoneEstimate.objects.filter(my_q, site = site).values_list('created', flat=True).order_by('created').distinct()
            
            first_date = created_dates.first()
            es_year = datetime.strptime(str(first_date), '%Y-%m-%d').year

            row_index = 3
            for created_date in created_dates:
                len_row_index = 0
                total_working_time = None
                production_note = None
                production_cph = None
                for i in range(len(list_customer_name)):
                    for j, time in enumerate(list_time):
                        len_row_index +=1

                        #ชั่วโมงทำงาน
                        total_working_time = Production.objects.filter(created = created_date, site = site).distinct().annotate(working_time = ExpressionWrapper(F('run_time') - F('total_loss_time'), output_field= models.DurationField())).aggregate(total_working_time=Sum('working_time'))['total_working_time']
                        #หมายเหตุ
                        production_note = Production.objects.filter(site = site, created = created_date).values_list('note', flat=True).first()
                        #capacity_per_hour
                        production_cph = Production.objects.filter(site = site, created = created_date).values_list('capacity_per_hour', flat=True).first()
                        #หินเขา
                        mountain1  = Weight.objects.filter(Q(time_out__gte=time['time_from']) & Q(time_out__lte=time['time_to']), Q(mill = '001MA') | Q(mill = '002MA'), Q(site = site) | Q(site__base_site_name = stock_type_name), bws__weight_type = 2, date = created_date, customer_name = list_customer_name[i]).aggregate(s_weight = Sum("weight_total"))

                        #หินเข้าโม่ทั้งหมด
                        crush1 = Weight.objects.filter(Q(time_out__gte=time['time_from']) & Q(time_out__lte=time['time_to']), bws__weight_type = 2 , date = created_date, customer_name = list_customer_name[i], site = site).aggregate(s_weight = Sum("weight_total"), c_weight=Count('weight_total'))

                        #กองสต็อกตามโรงโม่
                        stock1 = Weight.objects.filter(Q(time_out__gte=time['time_from']) & Q(time_out__lte=time['time_to']), bws__weight_type = 2 , date = created_date, customer_name = list_customer_name[i], site__base_site_name = stock_type_name).aggregate(s_weight = Sum("weight_total"), c_weight=Count('weight_total'))

                        #สร้างแถว 1
                        row1 = [created_date, list_customer_name[i], str(time['time_name']), formatHourMinute(total_working_time), mountain1['s_weight']]

                        for mill in mill_type:

                            weight_time1 = Weight.objects.filter(Q(time_out__gte=time['time_from']) & Q(time_out__lte=time['time_to']), bws__weight_type = 2, mill_name = mill, site = site, date = created_date, customer_name = list_customer_name[i]).aggregate(s_weight = Sum("weight_total"), c_weight=Count('weight_total'))
                            if weight_time1:
                                row1.extend([weight_time1['c_weight'], weight_time1['s_weight']])
                            else:
                                row1.extend(['' for i in range(3)])

                        row1.extend([stock1['c_weight'], stock1['s_weight']])
                        row1.extend([crush1['c_weight'], crush1['s_weight']])
                        row1.extend(['1'])
                        row1.extend([calculateEstimateToString(se_item, crush1['s_weight']) for se_item in StoneEstimateItem.objects.filter(se__created = created_date, se__site = site).order_by('stone_type').values_list('percent', flat=True)])
                        sheet.append(row1)

                    #merge_cells พนักงาน
                    sheet.merge_cells(start_row = row_index + len_row_index -2 , start_column = 2, end_row = row_index + len_row_index -1, end_column=2)

                sheet.cell(row = row_index + len_row_index, column=1, value='รวมทั้งหมด')
                sum_by_col = Decimal('0.00')
                for col in range(5, column_index):
                    for row in range(row_index, len_row_index + row_index):
                        sum_by_col = sum_by_col + Decimal( sheet.cell(row=row, column=col).value or '0.00' )
                    sheet.cell(row=row_index + len_row_index, column=col, value=sum_by_col).number_format = '#,##0.00'
                    sheet.cell(row=row_index + len_row_index, column=col).font = Font(bold=True)
                    sum_by_col = Decimal('0.00')
                row_index +=  len_row_index + 1

                sum_in_row = Decimal('0.00')
                for row in range(3, row_index):
                    for col in range(column_index - len(base_stone_type), column_index):
                        sum_in_row = sum_in_row + Decimal( sheet.cell(row=row, column=col).value or '0.00' )
                    sheet.cell(row=row, column=column_index, value=sum_in_row).number_format = '#,##0.00'
                    sheet.cell(row=row, column=column_index).font = Font(color="FF0000", bold=True)
                    sum_in_row = Decimal('0.00')

                #หินเข้าโม่รวม(ตัน) ของวันนั้นๆ
                sum_crush = sheet.cell(row=row_index-1, column=column_index).value
                capacity_per_hour = 0
                if es_year > 2024 and production_cph:#if ปี capacity_per_hour > 2024 ดึงแบบใหม่
                    capacity_per_hour = production_cph
                elif es_year <= 2024 and total_working_time:
                    (h, m) = str(format_duration(total_working_time)).split(':')
                    decimal_time = int(h) + (int(m) / 100)
                    decimal_time = Decimal(decimal_time)
                    #ผลิตตัน/ชม
                    capacity_per_hour = sum_crush/decimal_time

                sheet.cell(row = (row_index - 1 ) - len_row_index, column=column_index+1, value = f"{capacity_per_hour:.2f}")
                sheet.merge_cells(start_row = (row_index - 1 ) - len_row_index, start_column = column_index+1, end_row = (row_index - 1 ), end_column=column_index+1)

                #หมายเหตุ
                sheet.cell(row = (row_index - 1 ) - len_row_index, column=column_index+2, value = production_note)
                sheet.merge_cells(start_row = (row_index - 1 ) - len_row_index, start_column = column_index+2, end_row = (row_index - 1 ), end_column=column_index+2)

                #merge_cells วันที่, ชม.ทำงาน
                sheet.merge_cells(start_row = (row_index - 1 ) - len_row_index, start_column = 1, end_row = (row_index - 1 ), end_column=1)
                sheet.merge_cells(start_row = (row_index - 1 ) - len_row_index, start_column = 4, end_row = (row_index - 1 ), end_column=4)  

            # Total last
            len_row_index_total = 0
            for i in range(len(list_customer_name)):

                for j, time in enumerate(list_time):
                    len_row_index_total += 1

                    #ชั่วโมงทำงาน
                    total_working_time_tt = Production.objects.filter(my_q, site = site).distinct().annotate(working_time = ExpressionWrapper(F('run_time') - F('total_loss_time'), output_field= models.DurationField())).aggregate(total_working_time=Sum('working_time'))['total_working_time']

                    #หินเขา
                    mountain_tt  = Weight.objects.filter(sc_q, Q(time_out__gte=time['time_from']) & Q(time_out__lte=time['time_to']), Q(mill = '001MA') | Q(mill = '002MA'), Q(site = site) | Q(site__base_site_name = stock_type_name), bws__weight_type = 2, customer_name = list_customer_name[i]).aggregate(s_weight = Sum("weight_total"))

                    #หินเข้าโม่ทั้งหมด
                    crush_tt = Weight.objects.filter(sc_q, Q(time_out__gte=time['time_from']) & Q(time_out__lte=time['time_to']), bws__weight_type = 2, customer_name = list_customer_name[i], site = site).aggregate(s_weight = Sum("weight_total"), c_weight=Count('weight_total'))

                    #กองสต็อกตามโรงโม่
                    stock_tt = Weight.objects.filter(sc_q, Q(time_out__gte=time['time_from']) & Q(time_out__lte=time['time_to']), bws__weight_type = 2, customer_name = list_customer_name[i], site__base_site_name = stock_type_name).aggregate(s_weight = Sum("weight_total"), c_weight=Count('weight_total'))
                    
                    #สร้างแถว
                    row_tt = ["รวม", list_customer_name[i], str(time['time_name']), formatHourMinute(total_working_time_tt), mountain_tt['s_weight']]

                    for mill in mill_type:

                        weight_time_tt = Weight.objects.filter(sc_q, Q(time_out__gte=time['time_from']) & Q(time_out__lte=time['time_to']), bws__weight_type = 2, mill_name = mill, site = site, customer_name = list_customer_name[i]).aggregate(s_weight = Sum("weight_total"), c_weight=Count('weight_total'))
                        if weight_time_tt:
                            row_tt.extend([weight_time_tt['c_weight'], weight_time_tt['s_weight']])
                        else:
                            row_tt.extend(['' for i in range(3)])

                    row_tt.extend([stock_tt['c_weight'], stock_tt['s_weight']])
                    row_tt.extend([crush_tt['c_weight'], crush_tt['s_weight']])
                    row_tt.extend(['1'])
                    row_tt.extend([calculateSumEstimateToString(stone_type , site, list_customer_name[i], created_dates, time['time_from'], time['time_to']) for stone_type in StoneEstimateItem.objects.filter(se__created = created_date, se__site = site).order_by('stone_type').values_list('stone_type', flat=True)])
                    row_tt.extend([crush_tt['s_weight']])
                    sheet.append(row_tt)

                #merge_cells พนักงาน
                sheet.merge_cells(start_row = row_index + len_row_index_total -2 , start_column = 2, end_row = row_index + len_row_index_total -1, end_column=2)

            #merge_cells วันที่, ชม.ทำงาน
            sheet.merge_cells(start_row = row_index, start_column = 1, end_row = (row_index + len_row_index_total -1), end_column=1)
            sheet.merge_cells(start_row = row_index, start_column = 4, end_row = (row_index + len_row_index_total -1), end_column=4)
            sheet.merge_cells(start_row = row_index, start_column = column_index+1, end_row = (row_index + len_row_index_total -1), end_column=column_index+1)
            sheet.merge_cells(start_row = row_index, start_column = column_index+2, end_row = (row_index + len_row_index_total -1), end_column=column_index+2)


            # Set background color for the merged cells
            fill = PatternFill(start_color='F5CBA7', end_color='F5CBA7', fill_type='solid')  # Replace 'FF0000' with your desired color code
            sheet.cell(row = row_index + len_row_index_total, column=1, value='รวมทั้งหมด').fill = fill
            sheet.merge_cells(start_row = row_index + len_row_index_total, start_column = 1, end_row = row_index + len_row_index_total, end_column=4)

            sum_by_col_tt = Decimal('0.00')
            for col in range(5, column_index):
                for row in range(row_index, len_row_index_total + row_index):
                    sum_by_col_tt = sum_by_col_tt + Decimal( sheet.cell(row=row, column=col).value or '0.00' )
                sheet.cell(row=row_index + len_row_index_total, column=col, value=sum_by_col_tt).number_format = '#,##0.00'
                sheet.cell(row=row_index + len_row_index_total, column=col).font = Font(bold=True)
                sum_by_col_tt = Decimal('0.00')
            row_index +=  len_row_index_total + 1

            # Set column width and border for all columns
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)  # Get the letter of the current column

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value) + 2
                    except:
                        pass

                    if cell.column == 5:
                        cell.fill = PatternFill(start_color='93eef5', end_color='93eef5', fill_type='solid')

                    column_crush = len(mill_type) * 2 + 8
                    if cell.column == column_crush or cell.column == column_crush + 1:
                        cell.fill = PatternFill(start_color='FFD548', end_color='FFD548', fill_type='solid')

                    if cell.column == column_crush + 2:
                        cell.fill = PatternFill(start_color='F7BF94', end_color='F7BF94', fill_type='solid')

                adjusted_width = (max_length + 2) * 1.2  # Adjust the width based on content length
                sheet.column_dimensions[column_letter].width = adjusted_width

                # Set border for each cell in the column
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                for cell in column:
                    if cell.row < 3:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.border = border
            
        workbook.remove(workbook['Sheet'])
    else:
        worksheet = workbook.active
        worksheet.cell(row = 1, column = 1, value = f'ไม่มีข้อมูลรายงานผลิตแยกผู้รับเหมาและชนิดหินดือนนี้')

    # Save workbook into memory
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    size = output.getbuffer().nbytes

    # Generator to stream file in chunks
    def file_iterator(buffer, chunk_size=8192):
        while True:
            data = buffer.read(chunk_size)
            if not data:
                break
            yield data

    response = StreamingHttpResponse(
        file_iterator(output),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="Prod_daily_({active}).xlsx"'
    response["Content-Length"] = str(size)

    #workbook.save(response)
    return response

def exportExcelEstimate(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None
    site = request.GET.get('site') or None

    my_q = Q()
    if start_created is not None:
        my_q &= Q(se__created__gte = start_created)
    if end_created is not None:
        my_q &=Q(se__created__lte = end_created)
    if site is not None:
        my_q &=Q(se__site = site)

    my_q &= Q(se__company__code__in = company_in)
   
    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    startDate = datetime.strptime(start_created or startDateInMonth(previous_date_time.strftime('%Y-%m-%d')), "%Y-%m-%d").date()
    endDate = datetime.strptime(end_created or previous_date_time.strftime('%Y-%m-%d'), "%Y-%m-%d").date()

    #สร้าง list ระหว่าง start_date และ end_date
    list_date = [startDate+timedelta(days=x) for x in range((endDate-startDate).days + 1)]

    response = excelEstimate(request, my_q, list_date)
    return response

def exportExcelEstimateDashboard(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    
    end_created = request.session['db_end_date']
    start_created = request.session['db_start_date']

    my_q = Q()
    if start_created is not None:
        my_q &= Q(se__created__gte = start_created)
    if end_created is not None:
        my_q &=Q(se__created__lte = end_created)

    my_q &= Q(se__company__code__in = company_in)
    
    #เปลี่ยนออกเป็น ดึงรายงานของเดือนนั้นๆเท่านั้น
    startDate = datetime.strptime(start_created, "%Y-%m-%d").date()
    endDate = datetime.strptime(end_created, "%Y-%m-%d").date()

    #สร้าง list ระหว่าง start_date และ end_date
    list_date = [startDate+timedelta(days=x) for x in range((endDate-startDate).days + 1)]

    response = excelEstimate(request, my_q, list_date)
    return response

def excelEstimate(request, my_q, list_date):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    stone_id = StoneEstimateItem.objects.filter(Q(percent__gt = 0) & my_q).values_list('stone_type__base_stone_type_id', flat=True).order_by('stone_type__base_stone_type_id').distinct() #ดึงเฉพาะ stone_id ที่มีการคีย์ percent > 0
    data = StoneEstimateItem.objects.filter(Q(stone_type__in = stone_id) & my_q).order_by('se__created', 'se__site', 'stone_type').values_list('se__created', 'se__site__base_site_name', 'stone_type__base_stone_type_name', 'total')

    # Create a new workbook and get the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    if data:
        worksheet.cell(row=1, column=1, value='Date')
        worksheet.merge_cells(start_row=1, start_column = 1, end_row=2, end_column=1)

        date_style = NamedStyle(name='custom_datetime', number_format='DD/MM/YYYY')
        
        # Create a set of all unique mill and stone values
        sites = set()
        stones = set()
        for item in data:
            sites.add(item[1])
            stones.add(item[2]) 

        site_col_list = []
        
        # Create a list of colors for each line_type
        site_colors = [generate_pastel_color() for i  in range(len(sites) + 1)]

        column_index = 2
        for st in sites:
            worksheet.cell(row=1, column=column_index, value=f'การผลิตหิน {st}')
            worksheet.merge_cells(start_row=1, start_column = column_index, end_row=1, end_column=(column_index + len(stones) + 4) -1 )
            
            cell = worksheet.cell(row=1, column=column_index)
            cell.alignment = Alignment(horizontal='center')

            info = {}
            info['st'] = st
            info['strat_col'] = column_index
            info['end_col'] = column_index + len(stones) + 4
            site_col_list.append(info)

            #อัพเดทจำนวน col ตามที่มา
            column_index += len(stones) + 4

        #set color in header in row 1-2
        for row in worksheet.iter_rows(min_row=1, max_row=2):
            # Set the background color for each cell in the column
            for cell in row:
                #cell.border = Border(top=side, bottom=side, left=side, right=side)
                cell.alignment = Alignment(horizontal='center')
                line_index = (cell.column - 2) // (len(stones) + 4)
                fill_color = site_colors[line_index % len(site_colors)]
                fill = PatternFill(start_color=fill_color, fill_type="solid")
                cell.fill = fill

        # Write headers row 2 to the worksheet
        column_index = 2
        for st in sites:
            # top up ไม่ผ่านตาชั่ง, จากโรงโม่อื่น, จากตาชั่ง, รวมเข้าโม่
            worksheet.cell(row=2, column=column_index, value="Top up ไม่ผ่านตาชั่ง").alignment = Alignment(horizontal='center')
            worksheet.cell(row=2, column=column_index).font = Font(bold=True, color="0000FF")
            column_index += 1
            worksheet.cell(row=2, column=column_index, value="จากโรงโม่อื่น").alignment = Alignment(horizontal='center')
            worksheet.cell(row=2, column=column_index).font = Font(bold=True, color="0000FF")
            column_index += 1
            worksheet.cell(row=2, column=column_index, value="จากตาชั่ง").alignment = Alignment(horizontal='center')
            worksheet.cell(row=2, column=column_index).font = Font(bold=True, color="0000FF")
            column_index += 1
            worksheet.cell(row=2, column=column_index, value=f'รวมเข้า{st}').alignment = Alignment(horizontal='center')
            worksheet.cell(row=2, column=column_index).font = Font(bold=True, color="0000FF")
            column_index += 1
            for sou in stones:
                worksheet.cell(row=2, column=column_index, value=sou).alignment = Alignment(horizontal='center')
                column_index += 1
                
        # Create a dictionary to store data by date, mill, and stone
        date_data = {}

        # Loop through the data and populate the dictionary  
        for item in data:
            date = item[0]
            site = item[1]
            stone = item[2]
            value = item[3]

            if date not in date_data:
                date_data[date] = {}

            if site not in date_data[date]:
                date_data[date][site] = {}

            date_data[date][site][stone] = value

        row_index = 3
        for idl, ldate in enumerate(list_date):
            #เขียนวันที่ใน worksheet column 1
            worksheet.cell(row=idl+3, column=1, value=ldate).style = date_style
            worksheet.cell(row=idl+3, column=1).alignment = Alignment(horizontal='center')

            for date, site_data in date_data.items():
                #เขียน weight total ของแต่ละหินใน worksheet
                if worksheet.cell(row=idl+3, column = 1).value == date:
                    column_index = 2
                    for site in sites:
                        # top up ไม่ผ่านตาชั่ง, จากโรงโม่อื่น, จากตาชั่ง, รวมเข้าโม่
                        et = StoneEstimate.objects.filter(created = date, site__base_site_name = site, company__code__in = company_in).values_list('topup', 'other', 'scale', 'total').first() or (0, 0, 0, 0)
                        worksheet.cell(row=idl+3, column=column_index, value=et[0]).number_format = '#,##0.00'
                        column_index += 1
                        worksheet.cell(row=idl+3, column=column_index, value=et[1]).number_format = '#,##0.00'
                        column_index += 1
                        worksheet.cell(row=idl+3, column=column_index, value=et[2]).number_format = '#,##0.00'
                        column_index += 1
                        worksheet.cell(row=idl+3, column=column_index, value=et[3]).number_format = '#,##0.00'
                        column_index += 1
                        stone_data = site_data.get(site, {})
                        for stone in stones:
                            value = stone_data.get(stone, '')
                            worksheet.cell(row=idl+3, column=column_index, value=value).number_format = '#,##0.00'
                            column_index += 1
                    #row_index += 1
            row_index += 1

        worksheet.cell(row=row_index, column=1, value='รวมทั้งสิ้น')
        sum_by_col = Decimal('0.00')
        for col in range(2, column_index):
            for row in range(3, row_index):
                sum_by_col = sum_by_col + Decimal( worksheet.cell(row=row, column=col).value or '0.00' )
            worksheet.cell(row=row_index, column=col, value=sum_by_col).number_format = '#,##0.00'
            worksheet.cell(row=row_index, column=col).font = Font(bold=True, color="FF0000")
            sum_by_col = Decimal('0.00')

        # Set the column widths
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[2].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width
            worksheet.column_dimensions[column].height = 20

        side = Side(border_style='thin', color='000000')
        set_border(worksheet, side)
        worksheet.freeze_panes = "B3" #freeze
    else:
        worksheet.cell(row = 1, column = 1, value = f'ไม่มีข้อมูลผลิตหินประจำวันของเดือนนี้')

    # Save workbook into memory
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    size = output.getbuffer().nbytes

    # Generator to stream file in chunks
    def file_iterator(buffer, chunk_size=8192):
        while True:
            data = buffer.read(chunk_size)
            if not data:
                break
            yield data

    response = StreamingHttpResponse(
        file_iterator(output),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="estimate_stone_({active}).xlsx"'
    response["Content-Length"] = str(size)
    return response

################### BaesMill ####################
@login_required(login_url='login')
def settingBaseMill(request):
    #active : active คือแท็ปบริษัท active
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')

    data = BaseMill.objects.all().order_by('-mill_id')

    #กรองข้อมูล
    myFilter = BaseMillFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 15)
    page = request.GET.get('page')
    base_mill = p.get_page(page)

    context = {'setting_page':'active', 'setting_base_mill_page': 'active', 'base_mill': base_mill,'filter':myFilter, 'is_edit_setting': is_edit_setting(request.user), active :"active",}
    return render(request, "manage/BaseMill/baseMill.html",context)


def createBaseMill(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    form = BaseMillForm(request.POST or None, initial={'mill_id': generateCodeId('BaseMill', 1, None, None), 'user_created': request.user}) 
    if form.is_valid(): 
        new_contact = form.save(commit = False)
        duplicate = BaseMill.objects.filter(mill_id = new_contact.pk).exists()
        if duplicate:
            form.add_error(None, 'มีรหัสนี้อยู่แล้ว กรุณาเปลี่ยนรหัสใหม่.')
        else:
            try:
                new_contact.save()
            except IntegrityError:
                form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
            else:
                return redirect('settingBaseMill')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_mill_page': 'active',
        'table_name' : 'ต้นทาง',
        'text_mode' : 'เพิ่ม',
        'id_name' : '#id_mill_id',
        'mode' : 0,
        active :"active",
    }

    return render(request, "manage/BaseMill/formBaseMill.html", context)

def editBaseMill(request, id):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    obj = get_object_or_404(BaseMill, mill_id = id)
 
    form = BaseMillForm(request.POST or None, instance = obj)
    if form.is_valid():
        try:
            mill_form = form.save()

            # update weight ด้วย
            weights = Weight.objects.filter(mill_id = mill_form.pk)
            weights.update(mill_name = mill_form.mill_name)#iiiiiiiiii
        except IntegrityError:
            form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
        else:
            return redirect('settingBaseMill')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_mill_page': 'active',
        'table_name' : 'ต้นทาง',
        'text_mode' : 'เปลี่ยน',
        'id_name' : '#id_mill_id',
        'mode' : 1,
        active :"active",
    }

    return render(request, "manage/BaseMill/formBaseMill.html", context)

################### BaseJobType ####################
@login_required(login_url='login')
def settingBaseJobType(request):
    #active : active คือแท็ปบริษัท active
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')

    data = BaseJobType.objects.all().order_by('base_job_type_id')

    #กรองข้อมูล
    myFilter = BaseJobTypeFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 15)
    page = request.GET.get('page')
    base_job_type = p.get_page(page)

    context = {'setting_page':'active', 'setting_base_job_type_page': 'active', 'base_job_type': base_job_type,'filter':myFilter, 'is_edit_setting': is_edit_setting(request.user), active :"active",}
    return render(request, "manage/baseJobType.html",context)

def createBaseJobType(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    form = BaseJobTypeForm(request.POST or None, initial={'user_created': request.user}) 
    if form.is_valid(): 
        new_contact = form.save(commit = False)
        duplicate = BaseJobType.objects.filter(base_job_type_id = new_contact.pk).exists()
        if duplicate:
            form.add_error(None, 'มีรหัสนี้อยู่แล้ว กรุณาเปลี่ยนรหัสใหม่.')
        else:
            try:
                new_contact.save()
            except IntegrityError:
                form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
            else:
                return redirect('settingBaseJobType')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_job_type_page': 'active',
        'table_name' : 'ประเภทงานของลูกค้า',
        'text_mode' : 'เพิ่ม',
        'id_name' : '#id_base_job_type_id',
        'mode' : 0,
        active :"active",
    }
    return render(request, "manage/formBase.html", context)

def editBaseJobType(request, id):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    obj = get_object_or_404(BaseJobType, base_job_type_id = id)
 
    form = BaseJobTypeForm(request.POST or None, instance = obj)
    if form.is_valid():
        try:
            form.save()
        except IntegrityError:
            form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
        else:
            return redirect('settingBaseJobType')
 
    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_job_type_page': 'active',
        'table_name' : 'ประเภทงานของลูกค้า',
        'text_mode' : 'เปลี่ยน',
        'id_name' : '#id_base_job_type_id',
        'mode' : 1,
        active :"active",
    }
 
    return render(request, "manage/formBase.html", context)

################### BaesStoneType ####################
@login_required(login_url='login')
def settingBaseStoneType(request):
    #active : active คือแท็ปบริษัท active
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')

    data = BaseStoneType.objects.all().order_by('-base_stone_type_id')

    #กรองข้อมูล
    myFilter = BaseStoneTypeFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 15)
    page = request.GET.get('page')
    base_stone_type = p.get_page(page)

    context = {'setting_page':'active', 'setting_base_stone_type_page': 'active', 'base_stone_type': base_stone_type,'filter':myFilter, 'is_edit_setting': is_edit_setting(request.user), active :"active",}
    return render(request, "manage/baseStoneType.html",context)

def createBaseStoneType(request):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    form = BaseStoneTypeForm(request.POST or None, initial={'base_stone_type_id': generateCodeId('BaseStoneType', 1, None, None), 'user_created': request.user}) 
    if form.is_valid(): 
        new_contact = form.save(commit = False)
        duplicate = BaseStoneType.objects.filter(base_stone_type_id = new_contact.pk).exists()
        if duplicate:
            form.add_error(None, 'มีรหัสนี้อยู่แล้ว กรุณาเปลี่ยนรหัสใหม่.')
        else:
            try:
                new_contact.save()
            except IntegrityError:
                form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
            else:
                return redirect('settingBaseStoneType')
            
    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_stone_type_page': 'active',
        'table_name' : 'ชนิดหิน',
        'text_mode' : 'เพิ่ม',
        'id_name' : '#id_base_stone_type_id',
        'mode' : 0,
        active :"active",
    }

    return render(request, "manage/formBase.html", context)

def editBaseStoneType(request, id):
    active = request.session['company_code']
    company_in = findCompanyIn(request)            

    obj = get_object_or_404(BaseStoneType, base_stone_type_id = id)
 
    form = BaseStoneTypeForm(request.POST or None, instance = obj)
    if form.is_valid():
        try:
            stone_type_form = form.save()

            # update weight ด้วย
            weights = Weight.objects.filter(stone_type_id = stone_type_form.pk)
            weights.update(stone_type_name = stone_type_form.base_stone_type_name)#iiiiiiiiii
        except IntegrityError:
            form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
        else:
            return redirect('settingBaseStoneType')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_stone_type_page': 'active',
        'table_name' : 'ชนิดหิน',
        'text_mode' : 'เปลี่ยน',
        'id_name' : '#id_base_stone_type_id',
        'mode' : 1,
        active :"active",
    }

    return render(request, "manage/formBase.html", context)

################### BaesScoop ####################
@login_required(login_url='login')
def settingBaseScoop(request):
    #active : active คือแท็ปบริษัท active
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')

    data = BaseScoop.objects.all().order_by('-scoop_id')

    #กรองข้อมูล
    myFilter = BaseScoopFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 15)
    page = request.GET.get('page')
    base_scoop = p.get_page(page)

    context = {'setting_page':'active', 'setting_base_scoop_page': 'active', 'base_scoop': base_scoop,'filter':myFilter, 'is_edit_setting': is_edit_setting(request.user), active :"active",}
    return render(request, "manage/baseScoop.html",context)

def createBaseScoop(request):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)
        
    form = BaseScoopForm(request.POST or None, initial={'scoop_id': generateCodeId('BaseScoop', 1, None, None), 'user_created': request.user}) 
    if form.is_valid(): 
        new_contact = form.save(commit = False)
        duplicate = BaseScoop.objects.filter(scoop_id = new_contact.pk).exists()
        if duplicate:
            form.add_error(None, 'มีรหัสนี้อยู่แล้ว กรุณาเปลี่ยนรหัสใหม่.')
        else:
            try:
                new_contact.save()
            except IntegrityError:
                form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
            else:
                return redirect('settingBaseScoop')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_scoop_page': 'active',
        'table_name' : 'ผู้ตัก',
        'text_mode' : 'เพิ่ม',
        'id_name' : '#id_scoop_id',
        'mode' : 0,
        active :"active",
    }

    return render(request, "manage/formBase.html", context)

def editBaseScoop(request, id):
    active = request.session['company_code']
    company_in = findCompanyIn(request) 
            
    obj = get_object_or_404(BaseScoop, scoop_id = id)
 
    form = BaseScoopForm(request.POST or None, instance = obj)
    if form.is_valid():
        try:
            scoop_form = form.save()

            # update weight ด้วย
            weights = Weight.objects.filter(scoop_id = scoop_form.pk)
            weights.update(scoop_name = scoop_form.scoop_name)#iiiiiiiiii
        except IntegrityError:
            form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
        else:
            return redirect('settingBaseScoop')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_scoop_page': 'active',
        'table_name' : 'ผู้ตัก',
        'text_mode' : 'เปลี่ยน',
        'id_name' : '#id_scoop_id',
        'mode' : 1,
        active :"active",
    }

    return render(request, "manage/formBase.html", context)

################### BaseCarTeam ####################
@login_required(login_url='login')
def settingBaseCarTeam(request):
    #active : active คือแท็ปบริษัท active
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')

    data = BaseCarTeam.objects.all().order_by('-car_team_id')

    #กรองข้อมูล
    myFilter = BaseCarTeamFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 15)
    page = request.GET.get('page')
    base_car_team = p.get_page(page)

    context = {'setting_page':'active', 'setting_base_car_team_page': 'active', 'base_car_team': base_car_team,'filter':myFilter, 'is_edit_setting': is_edit_setting(request.user), active :"active",}
    return render(request, "manage/baseCarTeam.html",context)

def createBaseCarTeam(request):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    form = BaseCarTeamForm(request.POST or None, initial={'car_team_id': generateCodeId('BaseCarTeam', 2, None, None), 'user_created': request.user}) 
    if form.is_valid(): 
        new_contact = form.save(commit = False)
        #new_contact.oil_customer_id = generateOilCustomerId(new_contact.pk) #สร้างรหัสลูกค้าน้ำมัน auto
        duplicate = BaseCarTeam.objects.filter(car_team_id = new_contact.pk).exists()
        if duplicate:
            form.add_error(None, 'มีรหัสนี้อยู่แล้ว กรุณาเปลี่ยนรหัสใหม่.')
        else:
            try:
                new_contact.save()
            except IntegrityError:
                form.add_error(None, 'มีรหัสลูกค้าน้ำมัน หรือมีชื่อทีมนี้อยู่แล้ว กรุณาเปลี่ยนรหัสลูกค้าน้ำมันหรือชื่อทีมใหม่.')
            else:
                return redirect('settingBaseCarTeam')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_car_team_page': 'active',
        'table_name' : 'ทีม',
        'text_mode' : 'เพิ่ม',
        'id_name' : '#id_car_team_id',
        'mode' : 0,
        active :"active",
    }

    return render(request, "manage/formBase.html", context)

def editBaseCarTeam(request, id):
    active = request.session['company_code']
    company_in = findCompanyIn(request) 

    obj = get_object_or_404(BaseCarTeam, car_team_id = id)
 
    form = BaseCarTeamForm(request.POST or None, instance = obj)
    if form.is_valid():
        try:
            car_team_form = form.save()

            # update weight ด้วย
            weights = Weight.objects.filter(car_team_id = car_team_form.pk)
            weights.update(car_team_name = car_team_form.car_team_name)#iiiiiiiiii
        except IntegrityError:
            form.add_error(None, 'มีรหัสลูกค้าน้ำมัน หรือมีชื่อทีมนี้อยู่แล้ว กรุณาเปลี่ยนรหัสลูกค้าน้ำมันหรือชื่อทีมใหม่.')
        else:
            return redirect('settingBaseCarTeam')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_car_team_page': 'active',
        'table_name' : 'ทีม',
        'text_mode' : 'เปลี่ยน',
        'id_name' : '#id_car_team_id',
        'mode' : 1,
        active :"active",
    }

    return render(request, "manage/formBase.html", context)

################### BaseCar ####################
@login_required(login_url='login')
def settingBaseCar(request):
    #active : active คือแท็ปบริษัท active
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')

    data = BaseCar.objects.all().order_by('-car_id')

    #กรองข้อมูล
    myFilter = BaseCarFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 15)
    page = request.GET.get('page')
    base_car = p.get_page(page)

    context = {'setting_page':'active', 'setting_base_car_page': 'active', 'base_car': base_car,'filter':myFilter, 'is_edit_setting': is_edit_setting(request.user), active :"active",}
    return render(request, "manage/BaseCar/baseCar.html",context)

def createBaseCar(request):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    form = BaseCarForm(request.POST or None, initial={'user_created': request.user}) 
    if form.is_valid(): 
        new_contact = form.save(commit = False)
        duplicate = BaseCar.objects.filter(car_id = new_contact.pk).exists()
        if duplicate:
            form.add_error(None, 'มีรหัสนี้อยู่แล้ว กรุณาเปลี่ยนรหัสใหม่.')
        else:
            try:
                new_contact.save()
            except IntegrityError:
                form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
            else:
                return redirect('settingBaseCar')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_car_page': 'active',
        'table_name' : 'รถร่วม',
        'text_mode' : 'เพิ่ม',
        'id_name' : '#id_car_id',
        'mode' : 0,
        active :"active",
    }

    return render(request, "manage/BaseCar/formBaseCar.html", context)

def editBaseCar(request, id):
    active = request.session['company_code']
    company_in = findCompanyIn(request)         

    obj = get_object_or_404(BaseCar, car_id = id)
 
    form = BaseCarForm(request.POST or None, instance = obj)
    if form.is_valid():
        try:
            car_form = form.save()

            '''
            # update weight ด้วย
            weights = Weight.objects.filter(scoop_id = scoop_form.pk)
            weights.update(scoop_name = scoop_form.scoop_name)         
            '''
        except IntegrityError:
            form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
        else:
            return redirect('settingBaseCar')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_car_page': 'active',
        'table_name' : 'รถร่วม',
        'text_mode' : 'เปลี่ยน',
        'id_name' : '#id_car_id',
        'mode' : 1,
        active :"active",
    }

    return render(request, "manage/BaseCar/formBaseCar.html", context)

################### BaesSite ####################
@login_required(login_url='login')
def settingBaseSite(request):
    #active : active คือแท็ปบริษัท active
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')

    data = BaseSite.objects.all().order_by('-base_site_id')

    #กรองข้อมูล
    myFilter = BaseSiteFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 15)
    page = request.GET.get('page')
    base_site = p.get_page(page)

    context = {'setting_page':'active', 'setting_base_site_page': 'active', 'base_site': base_site,'filter':myFilter, 'is_edit_setting': is_edit_setting(request.user), active :"active",}
    return render(request, "manage/BaseSite/baseSite.html",context)

def createBaseSite(request):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    form = BaseSiteForm(request.POST or None, initial={'base_site_id': generateCodeId('BaseSite', 1, None, None), 'user_created': request.user})
    if form.is_valid(): 
        new_contact = form.save(commit = False)
        duplicate = BaseSite.objects.filter(base_site_id = new_contact.pk).exists()
        if duplicate:
            form.add_error(None, 'มีรหัสนี้อยู่แล้ว กรุณาเปลี่ยนรหัสใหม่.')
        else:
            try:
                new_contact.save()
            except IntegrityError:
                form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
            else:
                return redirect('settingBaseSite')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_site_page': 'active',
        'table_name' : 'ปลายทาง',
        'text_mode' : 'เพิ่ม',
        'id_name' : '#id_base_site_id',
        'mode' : 0,
        active :"active",
    }

    return render(request, "manage/BaseSite/formBaseSite.html", context)

def editBaseSite(request, id):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    obj = get_object_or_404(BaseSite, base_site_id = id)
 
    form = BaseSiteForm(request.POST or None, instance = obj)
    if form.is_valid():
        try:
            site_form = form.save()

            # update weight ด้วย
            weights = Weight.objects.filter(site_id = site_form.pk)
            weights.update(site_name = site_form.base_site_name)#iiiiiiiiii
        except IntegrityError:
            form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
        else:
            return redirect('settingBaseSite')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_site_page': 'active',
        'table_name' : 'ปลายทาง',
        'text_mode' : 'เปลี่ยน',
        'id_name' : '#id_base_site_id',
        'mode' : 1,
        active :"active",
    }

    return render(request, "manage/BaseSite/formBaseSite.html", context)

################### BaesCustomer ####################
@login_required(login_url='login')
def settingBaseCustomer(request):
    #active : active คือแท็ปบริษัท active
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')

    data = BaseCustomer.objects.filter(is_disable = False).order_by('-weight_type_id','-customer_id')

    #กรองข้อมูล
    myFilter = BaseCustomerFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 15)
    page = request.GET.get('page')
    base_customer = p.get_page(page)

    context = {'setting_page':'active', 'setting_base_customer_page': 'active', 'base_customer': base_customer,'filter':myFilter, 'is_edit_setting': is_edit_setting(request.user), active :"active",}
    return render(request, "manage/BaseCustomer/baseCustomer.html",context)

def createBaseCustomer(request):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    form = BaseCustomerForm(request.POST or None, initial={'user_created': request.user}) 
    if form.is_valid(): 
        new_contact = form.save(commit = False)
        duplicate = BaseCustomer.objects.filter(customer_id = new_contact.pk).exists()
        if duplicate:
            form.add_error(None, 'มีรหัสนี้อยู่แล้ว กรุณาเปลี่ยนรหัสใหม่.')
        else:
            try:
                new_contact.save()
            except IntegrityError:
                form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
            else:
                return redirect('settingBaseCustomer')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_customer_page': 'active',
        'table_name' : 'ลูกค้า',
        'text_mode' : 'เพิ่ม',
        'id_name' : '#id_customer_id',
        'is_edit_base_id': is_edit_base_id(request.user),
        'mode' : 0,
        active :"active",
    }

    return render(request, "manage/BaseCustomer/formBaseCustomer.html", context)

def editBaseCustomer(request, id):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    obj = get_object_or_404(BaseCustomer, customer_id = id)
 
    form = BaseCustomerForm(request.POST or None, instance = obj)
    if form.is_valid():
        try:
            customer_form = form.save()

            # update weight ด้วย
            weights = Weight.objects.filter(customer_id = customer_form.pk)
            weights.update(customer_name = customer_form.customer_name)#iiiiiiiiii
        except IntegrityError:
            form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
        else:
            return redirect('settingBaseCustomer')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_customer_page': 'active',
        'table_name' : 'ลูกค้า',
        'text_mode' : 'เปลี่ยน',
        'id_name' : '#id_customer_id',
        'is_edit_base_id': is_edit_base_id(request.user),
        'mode' : 1,
        active :"active",
    }

    return render(request, "manage/BaseCustomer/formBaseCustomer.html", context)

################### BaseDriver ####################
@login_required(login_url='login')
def settingBaseDriver(request):
    #active : active คือแท็ปบริษัท active
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')

    data = BaseDriver.objects.all().order_by('-driver_id')

    #กรองข้อมูล
    myFilter = BaseDriverFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 15)
    page = request.GET.get('page')
    base_driver = p.get_page(page)

    context = {'setting_page':'active', 'setting_base_driver_page': 'active', 'base_driver': base_driver,'filter':myFilter, 'is_edit_setting': is_edit_setting(request.user), active :"active",}
    return render(request, "manage/baseDriver.html",context)

def createBaseDriver(request):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    form = BaseDriverForm(request.POST or None, initial={'driver_id': generateCodeId('BaseDriver', 1, None, None), 'user_created': request.user}) 
    if form.is_valid(): 
        new_contact = form.save(commit = False)
        duplicate = BaseDriver.objects.filter(driver_id = new_contact.pk).exists()
        if duplicate:
            form.add_error(None, 'มีรหัสนี้อยู่แล้ว กรุณาเปลี่ยนรหัสใหม่.')
        else:
            try:
                new_contact.save()
            except IntegrityError:
                form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
            else:
                return redirect('settingBaseDriver')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_driver_page': 'active',
        'table_name' : 'ผู้ขับ',
        'text_mode' : 'เพิ่ม',
        'id_name' : '#id_driver_id',
        'mode' : 0,
        active :"active",
    }

    return render(request, "manage/formBase.html", context)

def editBaseDriver(request, id):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)
            
    obj = get_object_or_404(BaseDriver, driver_id = id)
 
    form = BaseDriverForm(request.POST or None, instance = obj)
    if form.is_valid():
        try:
            driver_form = form.save()

            # update weight ด้วย
            weights = Weight.objects.filter(driver_id = driver_form.pk)
            weights.update(driver_name = driver_form.driver_name)#iiiiiiiiii
        except IntegrityError:
            form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
        else:
            return redirect('settingBaseDriver')
        
    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_driver_page': 'active',
        'table_name' : 'ผู้ขับ',
        'text_mode' : 'เปลี่ยน',
        'id_name' : '#id_driver_id',
        'mode' : 1,
        active :"active",
    }

    return render(request, "manage/formBase.html", context)

################### BaseCarRegistration ####################
@login_required(login_url='login')
def settingBaseCarRegistration(request):
    #active : active คือแท็ปบริษัท active
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')

    data = BaseCarRegistration.objects.all().order_by('-car_registration_id')

    #กรองข้อมูล
    myFilter = BaseCarRegistrationFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 15)
    page = request.GET.get('page')
    base_car_registration = p.get_page(page)

    context = {'setting_page':'active', 'setting_base_car_registration_page': 'active', 'base_car_registration': base_car_registration,'filter':myFilter, 'is_edit_setting': is_edit_setting(request.user), active :"active",}
    return render(request, "manage/baseCarRegistration.html",context)

def createBaseCarRegistration(request):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    form = BaseCarRegistrationForm(request.POST or None, initial={'car_registration_id': generateCodeId('BaseCarRegistration', 1, None, None), 'user_created': request.user}) 
    if form.is_valid(): 
        new_contact = form.save(commit = False)
        duplicate = BaseCarRegistration.objects.filter(car_registration_id = new_contact.pk).exists()
        if duplicate:
            form.add_error(None, 'มีรหัสนี้อยู่แล้ว กรุณาเปลี่ยนรหัสใหม่.')
        else:
            try:
                new_contact.save()
            except IntegrityError:
                form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
            else:
                return redirect('settingBaseCarRegistration')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_car_registration_page': 'active',
        'table_name' : 'ทะเบียนรถ',
        'text_mode' : 'เพิ่ม',
        'id_name' : '#id_car_registration_id',
        'mode' : 0,
        active :"active",
    }

    return render(request, "manage/formBase.html", context)

def editBaseCarRegistration(request, id):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    obj = get_object_or_404(BaseCarRegistration, car_registration_id = id)
 
    form = BaseCarRegistrationForm(request.POST or None, instance = obj)
    if form.is_valid():
        try:
            car_registration_form = form.save()

            # update weight ด้วย
            weights = Weight.objects.filter(car_registration_id = car_registration_form.pk)
            weights.update(car_registration_name = car_registration_form.car_registration_name)#iiiiiiiiii
        except IntegrityError:
            form.add_error(None, 'มีชื่อนี้อยู่แล้ว กรุณาตั้งชื่อใหม่.')
        else:
            return redirect('settingBaseCarRegistration')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_car_registration_page': 'active',
        'table_name' : 'ทะเบียนรถ',
        'text_mode' : 'เปลี่ยน',
        'id_name' : '#id_car_registration_id',
        'mode' : 1,
        active :"active",
    }

    return render(request, "manage/formBase.html", context)

################### BaseCustomerSite ####################
@login_required(login_url='login')
def settingBaseCustomerSite(request):
    #active : active คือแท็ปบริษัท active
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')

    data = BaseCustomerSite.objects.all().order_by('id')

    #กรองข้อมูล
    myFilter = BaseCustomerSiteFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 15)
    page = request.GET.get('page')
    base_customer_site = p.get_page(page)

    context = {'setting_page':'active', 'setting_base_customer_site_page': 'active', 'base_customer_site': base_customer_site,'filter':myFilter, 'is_edit_setting': is_edit_setting(request.user), active :"active",}
    return render(request, "manage/BaseCustomerSite/baseCustomerSite.html",context)

def createBaseCustomerSite(request):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    if request.method == 'POST':
        form = BaseCustomerSiteForm(request.POST or None, request.FILES)
        if form.is_valid():
            try:
                form.save()
            except IntegrityError:
                form.add_error(None, 'This combination of field1 and field2 is not unique.')
            else:
                return redirect('settingBaseCustomerSite')
    else:
        form = BaseCustomerSiteForm(initial={'user_created': request.user})

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_customer_site_page': 'active',
        'table_name' : 'ลูกค้าและหน้างาน',
        'text_mode' : 'เพิ่ม',
        active :"active",
    }

    return render(request, "manage/BaseCustomerSite/formBaseCustomerSite.html", context)

def editBaseCustomerSite(request, id):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    data = BaseCustomerSite.objects.get(id = id)
    
    form = BaseCustomerSiteForm(instance=data)
    if request.method == 'POST':
        form = BaseCustomerSiteForm(request.POST, instance=data)
        if form.is_valid():
            try:
                customer_site_form = form.save()
            except IntegrityError:
                form.add_error(None, 'This combination of field1 and field2 is not unique.')
            else:
                return redirect('settingBaseCustomerSite')

    context = {
        'form':form,
        'setting_page':'active',
        'setting_base_customer_site_page': 'active',
        'table_name' : 'ลูกค้าและหน้างาน',
        'text_mode' : 'เปลี่ยน',
        active :"active",
    }

    return render(request, "manage/BaseCustomerSite/formBaseCustomerSite.html", context)

#################################
############# API ###############
#################################

############# Login API ###############
class LoginApiView(APIView):
    permission_classes = []

    def post(self, request: Request):
        username = request.data.get("username")
        password = request.data.get("password")

        user = authenticate(username=username, password=password)
        if user is not None:
            tokens = create_jwt_pair_for_user(user)

            response = {"message": "Login Successfull", "tokens": tokens}
            return Response(data=response, status=status.HTTP_200_OK)

        else:
            return Response(data={"message": "Invalid email or password"})

    def get(self, request: Request):
        content = {"user": str(request.user), "auth": str(request.auth)}

        return Response(data=content, status=status.HTTP_200_OK)

############# SignUp API ###############   
class SignUpApiView(generics.GenericAPIView):
    serializer_class = SignUpSerializer
    permission_classes = []

    def post(self, request: Request):
        data = request.data

        serializer = self.serializer_class(data=data)

        if serializer.is_valid():
            serializer.save()

            response = {"message": "User Created Successfully", "data": serializer.data}

            return Response(data=response, status=status.HTTP_201_CREATED)

        return Response(data=serializer.errors, status=status.HTTP_400_BAD_REQUEST)

############# Weight API ###############
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def apiWeightOverview(request):
    api_urls = {
        'List':'/weight/api/list/',
        'Detail View':'/weight/api/detail/<str:pk>/',
        'Detail By Date':'/weight/api/detail/date/<str:str_date>/',
        'Create':'/weight/api/create/',
        'Update':'/weight/api/update/<str:pk>/',
        'VStamp':'/weight/api/vStamp/<str:dt>/<str:str_lc>/',
        'Detail By Date Between and Weight Type':'/weight/api/between/<str:start_date>/<str:end_date>/<str:weight_type>/',
        'Detail By BWS':'/weight/api/between/<str:start_date>/<str:end_date>/<str:bws>/',
    }
    return Response(api_urls)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def weightList(request):
    queryset = Weight.objects.all()
    serializer = WeightSerializer(queryset, many = True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def weightDetail(request, pk):
    queryset = Weight.objects.get(weight_id = pk)
    serializer = WeightSerializer(queryset, many = False)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def weightDetailByDate(request, str_date , str_lc):
    latest_weights = WeightHistory.objects.filter(date = str_date, bws__id = str_lc).values('weight_id').distinct()
    queryset = Weight.objects.filter(weight_id__in = latest_weights)

    serializer = WeightSerializer(queryset, many = True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def weightVStamp(request, dt, str_lc):
    latest_weights = WeightHistory.objects.filter(user_update__isnull = False, v_stamp__gte = dt, bws__id = str_lc).order_by('v_stamp').values('weight_id').distinct()
    queryset = Weight.objects.filter(weight_id__in = latest_weights)
    
    serializer = WeightSerializer(queryset, many = True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def weightVStampAll(request, dt):
    latest_weights = WeightHistory.objects.filter(v_stamp__gte = dt).order_by('v_stamp').values('weight_id').distinct()
    queryset = Weight.objects.filter(weight_id__in = latest_weights)
    
    serializer = WeightSerializer(queryset, many = True)
    return Response(serializer.data)

# For Insert Report weight for SLC
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def weightDetailBetween(request, start_date, end_date , weight_type):
    #อันนี้ถ้าเป็นเครื่องขายให้รวมข้อมูลของ uni ด้วย ในโปรแกรมตาชั่งรวม (Report Center) เฉพาะเครื่องขาย
    if weight_type == 1:
        queryset = Weight.objects.filter(date__range=[start_date, end_date], bws__weight_type__id = weight_type, bws__company__code__in = ['SLC', 'UNI'])
    elif weight_type == 2:
       queryset = Weight.objects.filter(date__range=[start_date, end_date], bws__weight_type__id = weight_type, bws__company__code = 'SLC')
    
    #09-09-2024 อันเก่าไม่ครอบคลุมเคส
    #queryset = Weight.objects.filter(date__range=[start_date, end_date], bws__weight_type__id = weight_type, bws__company__code = 'SLC')

    serializer = WeightSerializer(queryset, many = True)
    return Response(serializer.data)

# For get between date by bws
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def weightDetailBetweenByBWS(request, start_date, end_date , bws):
    queryset = Weight.objects.filter(date__range=[start_date, end_date], bws = bws)

    serializer = WeightSerializer(queryset, many = True)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def weightCreate(request):
    serializer = WeightSerializer(data = request.data)
    
    if serializer.is_valid():
        try:
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except IntegrityError as e:
            return Response(serializer.data, status=status.HTTP_409_CONFLICT)
    else:
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def weightUpdate(request, pk):
    queryset = Weight.objects.get(weight_id = pk)
    serializer = WeightSerializer(instance=queryset, data = request.data)
    
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    else:
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

############# BaseScoop API ###############
class BaseScoopView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]

    queryset = BaseScoop.objects.all()
    serializer_class = BaseScoopSerializer

class BaseScoopViewById(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]

    queryset = BaseScoop.objects.all()
    def get_queryset(self):
        queryset = BaseScoop.objects.filter(scoop_id=self.kwargs["pk"])
        return queryset
    serializer_class = BaseScoopSerializer

class CreateBaseScoop(APIView):
    permission_classes = [IsAuthenticated]

    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'manage/formBase.html'

    serializer_class = BaseScoopSerializer

    def post(self, request):
        scoop_id = request.data.get("scoop_id")
        scoop_name = request.data.get("scoop_name")

        data = {'scoop_id': scoop_id, 'scoop_name': scoop_name}

        serializer = BaseScoopSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseScoopDetail(request, pk):
    queryset = BaseScoop.objects.get(scoop_id=pk)
    serializer = BaseScoopSerializer(queryset, many = False)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseScoopVStamp(request, dt):
    queryset = BaseScoop.objects.filter(v_stamp__gte = dt).order_by('v_stamp')
    serializer = BaseScoopSerializer(queryset, many = True)
    return Response(serializer.data)
        
############# BaseMill API ###############
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def apiBaseMillOverview(request):
    api_urls = {
        'List':'/baseMill/api/list/',
        'Detail View':'/baseMill/api/detail/<str:pk>/',
        'Create':'/baseMill/api/create/',
        'Update':'/baseMill/api/update/<str:pk>/',
        'Delete':'/baseMill/api/delete/<str:pk>/',
        'VStamp':'/baseMill/api/vStamp/<str:dt>/',
    }
    return Response(api_urls)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseMillList(request):
    queryset = BaseMill.objects.all()
    serializer = BaseMillSerializer(queryset, many = True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseMillDetail(request, pk):
    queryset = BaseMill.objects.get(mill_id=pk)
    serializer = BaseMillSerializer(queryset, many = False)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def baseMillCreate(request):
    serializer = BaseMillSerializer(data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def baseMillUpdate(request, pk):
    queryset = BaseMill.objects.get(mill_id=pk)
    serializer = BaseMillSerializer(instance=queryset, data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def baseMillDelete(request, pk):
    queryset = BaseMill.objects.get(mill_id=pk)
    queryset.delete()

    return Response("Item successfully delete!")

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseMillVStamp(request, dt):
    queryset = BaseMill.objects.filter(v_stamp__gte = dt).order_by('v_stamp')
    serializer = BaseMillSerializer(queryset, many = True)
    return Response(serializer.data)

############# base customer API ###############
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def apiBaseCustomerOverview(request):
    api_urls = {
        'List':'/baseCustomer/api/list/',
        'Detail View':'/baseCustomer/api/detail/<str:pk>/',
        'Create':'/baseCustomer/api/create/',
        'Update':'/baseCustomer/api/update/<str:pk>/',
        'VStamp':'/baseCustomer/api/vStamp/<str:dt>/',
    }
    return Response(api_urls)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseCustomerList(request):
    queryset = BaseCustomer.objects.all()
    serializer = BaseCustomerSerializer(queryset, many = True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseCustomerDetail(request, pk):
    queryset = BaseCustomer.objects.get(customer_id = pk)
    serializer = BaseCustomerSerializer(queryset, many = False)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def baseCustomerCreate(request):
    serializer = BaseCustomerSerializer(data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def baseCustomerUpdate(request, pk):
    try:
        base_customer = BaseCustomer.objects.get(customer_id=pk)
        weights = Weight.objects.filter(customer_id = pk)

        # Update BaseCustomer
        base_customer_serializer = BaseCustomerSerializer(instance=base_customer, data=request.data)
        if base_customer_serializer.is_valid():
            base_customer_serializer.save()

            customer_name = request.data.get("customer_name")
            # 1 Update Weight
            weights.update(customer_name = customer_name)#iiiiiiiiii
            
            # 2 Update Weight ไม่ต้องแล้ว
            '''
            data_weight = {'customer_id': pk, 'customer_name': customer_name}
            for weight in weights:
                weight_serializer = WeightSerializer(instance=weight, data=data_weight)
                if weight_serializer.is_valid():
                    weight_serializer.save()         
            '''
        return Response({'message': 'Data updated successfully'})
    except BaseCustomer.DoesNotExist or Weight.DoesNotExist:
        return Response({'error': 'Record not found'}, status=404)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseCustomerVStamp(request, dt):
    # รายการ disable ไม่ให้อัพเดท (เนื่องจากไม่ได้ใช้งานแล้ว แต่ต้องเก็บข้อมูลไว้)
    queryset = BaseCustomer.objects.filter(v_stamp__gte = dt, is_disable = False).order_by('v_stamp')
    serializer = BaseCustomerSerializer(queryset, many = True)
    return Response(serializer.data)

############# BaseStoneType API ###############
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def apiBaseStoneTypeOverview(request):
    api_urls = {
        'List':'/baseStoneType/api/list/',
        'Detail View':'/baseStoneType/api/detail/<str:pk>/',
        'Create':'/baseStoneType/api/create/',
        'Update':'/baseStoneType/api/update/<str:pk>/',
        'Delete':'/baseStoneType/api/delete/<str:pk>/',
        'VStamp':'/baseStoneType/api/vStamp/<str:dt>/',
    }
    return Response(api_urls)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseStoneTypeList(request):
    queryset = BaseStoneType.objects.all()
    serializer = BaseStoneTypeSerializer(queryset, many = True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseStoneTypeDetail(request, pk):
    queryset = BaseStoneType.objects.get(base_stone_type_id=pk)
    serializer = BaseStoneTypeSerializer(queryset, many = False)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def baseStoneTypeCreate(request):
    serializer = BaseStoneTypeSerializer(data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def baseStoneTypeUpdate(request, pk):
    queryset = BaseStoneType.objects.get(base_stone_type_id=pk)
    serializer = BaseStoneTypeSerializer(instance=queryset, data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseStoneTypeVStamp(request, dt):
    queryset = BaseStoneType.objects.filter(v_stamp__gte = dt).order_by('v_stamp')
    serializer = BaseStoneTypeSerializer(queryset, many = True)
    return Response(serializer.data)

class BaseStoneTypeList(generics.ListCreateAPIView):
    queryset = BaseStoneType.objects.all()
    serializer_class = BaseStoneTypeTestSerializer
    

############# BaseCarTeam API ###############
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def apiBaseCarTeamOverview(request):
    api_urls = {
        'List':'/baseCarTeam/api/list/',
        'Detail View':'/baseCarTeam/api/detail/<str:pk>/',
        'Create':'/baseCarTeam/api/create/',
        'Update':'/baseCarTeam/api/update/<str:pk>/',
        'Delete':'/baseCarTeam/api/delete/<str:pk>/',
        'VStamp':'/baseCarTeam/api/vStamp/<str:dt>/',
    }
    return Response(api_urls)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseCarTeamList(request):
    queryset = BaseCarTeam.objects.all()
    serializer = BaseCarTeamSerializer(queryset, many = True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseCarTeamDetail(request, pk):
    queryset = BaseCarTeam.objects.get(car_team_id=pk)
    serializer = BaseCarTeamSerializer(queryset, many = False)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def baseCarTeamCreate(request):
    serializer = BaseCarTeamSerializer(data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def baseCarTeamUpdate(request, pk):
    queryset = BaseCarTeam.objects.get(car_team_id=pk)
    serializer = BaseCarTeamSerializer(instance=queryset, data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseCarTeamVStamp(request, dt):
    queryset = BaseCarTeam.objects.filter(v_stamp__gte = dt).order_by('v_stamp')
    serializer = BaseCarTeamSerializer(queryset, many = True)
    return Response(serializer.data)

############# BaseDriver API ###############
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def apiBaseDriverOverview(request):
    api_urls = {
        'List':'/baseDriver/api/list/',
        'Detail View':'/baseDriver/api/detail/<str:pk>/',
        'Create':'/baseDriver/api/create/',
        'Update':'/baseDriver/api/update/<str:pk>/',
        'Delete':'/baseDriver/api/delete/<str:pk>/',
        'VStamp':'/baseDriver/api/vStamp/<str:dt>/',
    }
    return Response(api_urls)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseDriverList(request):
    queryset = BaseDriver.objects.all()
    serializer = BaseDriverSerializer(queryset, many = True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseDriverDetail(request, pk):
    queryset = BaseDriver.objects.get(driver_id=pk)
    serializer = BaseDriverSerializer(queryset, many = False)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def baseDriverCreate(request):
    serializer = BaseDriverSerializer(data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def baseDriverUpdate(request, pk):
    queryset = BaseDriver.objects.get(driver_id=pk)
    serializer = BaseDriverSerializer(instance=queryset, data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseDriverVStamp(request, dt):
    queryset = BaseDriver.objects.filter(v_stamp__gte = dt).order_by('v_stamp')
    serializer = BaseDriverSerializer(queryset, many = True)
    return Response(serializer.data)

############# BaseCarRegistration API ###############
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def apiBaseCarRegistrationOverview(request):
    api_urls = {
        'List':'/baseCarRegistration/api/list/',
        'Detail View':'/baseCarRegistration/api/detail/<str:pk>/',
        'Create':'/baseCarRegistration/api/create/',
        'Update':'/baseCarRegistration/api/update/<str:pk>/',
        'Delete':'/baseCarRegistration/api/delete/<str:pk>/',
        'VStamp':'/baseCarRegistration/api/vStamp/<str:dt>/',
    }
    return Response(api_urls)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseCarRegistrationList(request):
    queryset = BaseCarRegistration.objects.all()
    serializer = BaseCarRegistrationSerializer(queryset, many = True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseCarRegistrationDetail(request, pk):
    queryset = BaseCarRegistration.objects.get(car_registration_id=pk)
    serializer = BaseCarRegistrationSerializer(queryset, many = False)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def baseCarRegistrationCreate(request):
    serializer = BaseCarRegistrationSerializer(data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def baseCarRegistrationUpdate(request, pk):
    queryset = BaseCarRegistration.objects.get(car_registration_id=pk)
    serializer = BaseCarRegistrationSerializer(instance=queryset, data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseCarRegistrationVStamp(request, dt):
    queryset = BaseCarRegistration.objects.filter(v_stamp__gte = dt).order_by('v_stamp')
    serializer = BaseCarRegistrationSerializer(queryset, many = True)
    return Response(serializer.data)

############# BaseSite API ###############
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def apiBaseSiteOverview(request):
    api_urls = {
        'List':'/baseSite/api/list/',
        'Detail View':'/baseSite/api/detail/<str:pk>/',
        'Create':'/baseSite/api/create/',
        'Update':'/baseSite/api/update/<str:pk>/',
        'Delete':'/baseSite/api/delete/<str:pk>/',
        'VStamp':'/baseSite/api/vStamp/<str:dt>/',
    }
    return Response(api_urls)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseSiteList(request):
    queryset = BaseSite.objects.all()
    serializer = BaseSiteSerializer(queryset, many = True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseSiteDetail(request, pk):
    queryset = BaseSite.objects.get(base_site_id=pk)
    serializer = BaseSiteSerializer(queryset, many = False)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def baseSiteCreate(request):
    serializer = BaseSiteSerializer(data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def baseSiteUpdate(request, pk):
    queryset = BaseSite.objects.get(base_site_id=pk)
    serializer = BaseSiteSerializer(instance=queryset, data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseSiteVStamp(request, dt):
    queryset = BaseSite.objects.filter(v_stamp__gte = dt).order_by('v_stamp')
    serializer = BaseSiteSerializer(queryset, many = True)
    return Response(serializer.data)


############# Page API ###############
class SmallResultsSetPagination(PageNumberPagination):
    page_size = 100  # or any number suitable for your fro

############# BaseCar API ###############
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def apiBaseCarOverview(request):
    api_urls = {
        'List':'/baseCar/api/list/',
        'Detail View':'/baseCar/api/detail/<str:pk>/',
        'Create':'/baseCar/api/create/',
        'Update':'/baseCar/api/update/<str:pk>/',
        'Delete':'/baseCar/api/delete/<str:pk>/',
        'VStamp':'/baseCar/api/vStamp/<str:dt>/',
        'All Base Car View':'/car/partner/api/all/',
    }
    return Response(api_urls)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseCarList(request):
    queryset = BaseCar.objects.all()
    serializer = BaseCarSerializer(queryset, many = True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseCarDetail(request, pk):
    queryset = BaseCar.objects.get(car_id=pk)
    serializer = BaseCarSerializer(queryset, many = False)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def baseCarCreate(request):
    serializer = BaseCarSerializer(data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def baseCarUpdate(request, pk):
    queryset = BaseCar.objects.get(car_id=pk)
    serializer = BaseCarSerializer(instance=queryset, data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseCarVStamp(request, dt):
    queryset = BaseCar.objects.filter(v_stamp__gte = dt).order_by('v_stamp')
    serializer = BaseCarSerializer(queryset, many = True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def allCarPartner(request):
    queryset = BaseCar.objects.all()
    paginator = SmallResultsSetPagination()
    result_page = paginator.paginate_queryset(queryset, request)
    serializer = CarPartnerSerializer(result_page, many=True)
    return paginator.get_paginated_response(serializer.data)

############# BaseJobType API ###############
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def apiBaseJobTypeOverview(request):
    api_urls = {
        'List':'/baseJobType/api/list/',
        'Detail View':'/baseJobType/api/detail/<str:pk>/',
        'Create':'/baseJobType/api/create/',
        'Update':'/baseJobType/api/update/<str:pk>/',
        'Delete':'/baseJobType/api/delete/<str:pk>/',
        'VStamp':'/baseJobType/api/vStamp/<str:dt>/',
    }
    return Response(api_urls)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseJobTypeList(request):
    queryset = BaseJobType.objects.all()
    serializer = BaseJobTypeSerializer(queryset, many = True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseJobTypeDetail(request, pk):
    queryset = BaseJobType.objects.get(base_job_type_id =pk)
    serializer = BaseJobTypeSerializer(queryset, many = False)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def baseJobTypeCreate(request):
    serializer = BaseJobTypeSerializer(data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def baseJobTypeUpdate(request, pk):
    queryset = BaseJobType.objects.get(base_job_type_id=pk)
    serializer = BaseJobTypeSerializer(instance=queryset, data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseJobTypeVStamp(request, dt):
    queryset = BaseJobType.objects.filter(v_stamp__gte = dt).order_by('v_stamp')
    serializer = BaseJobTypeSerializer(queryset, many = True)
    return Response(serializer.data)


############# BaseCustomerSite API ###############
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def apiBaseCustomerSiteOverview(request):
    api_urls = {
        'List':'/baseCustomerSite/api/list/',
        'Detail View':'/baseCustomerSite/api/detail/<str:pk>/',
        'Create':'/baseCustomerSite/api/create/',
        'Update':'/baseCustomerSite/api/update/<str:pk>/',
        'Delete':'/baseCustomerSite/api/delete/<str:pk>/',
        'VStamp':'/baseCustomerSite/api/vStamp/<str:dt>/',
    }
    return Response(api_urls)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseCustomerSiteList(request):
    queryset = BaseCustomerSite.objects.all()
    serializer = BaseCustomerSiteSerializer(queryset, many = True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseCustomerSiteDetail(request, pk):
    queryset = BaseCustomerSite.objects.get(id = pk)
    serializer = BaseCustomerSiteSerializer(queryset, many = False)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def baseCustomerSiteCreate(request):
    serializer = BaseCustomerSiteSerializer(data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def baseCustomerSiteUpdate(request, pk):
    queryset = BaseCustomerSite.objects.get(id=pk)
    serializer = BaseCustomerSiteSerializer(instance=queryset, data = request.data)
    
    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def baseCustomerSiteVStamp(request, dt):
    queryset = BaseCustomerSite.objects.filter(v_stamp__gte = dt).order_by('v_stamp')
    serializer = BaseCustomerSiteSerializer(queryset, many = True)
    return Response(serializer.data)

def searchDetailMcType(request):
    #ดึงรายงานของเดือนนั้นๆ
    '''ดึงวันที่ตามหน้า dashbord
    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    end_created = previous_date_time.strftime('%Y-%m-%d')
    start_created = startDateInMonth(end_created)    
    '''

    start_created = request.session['db_start_date']
    end_created = request.session['db_end_date']

    site_id = request.GET.get('site_id', None)
    mc_id = request.GET.get('mc_id', None)
    loss = ProductionLossItem.objects.filter(production__site = site_id, mc_type = mc_id, production__created__range=[start_created, end_created]).values('loss_type__name', 'production__site__base_site_name', 'mc_type__name').annotate(sum_time = Sum('loss_time'))

    index = 1
    try:
        strName = "<table class='table'><thead class='table-info'><tr><th colspan='4'>"+ loss[0]['mc_type__name'] +"</th></thead></tr>"
        for i in loss:
            strName = ''.join([strName, "<tr>"])
            strName = ''.join([strName, "<td>" + str(index) + ")</td><td><b>"+ i['loss_type__name'] + "</td><td>"+ str(i['sum_time'])  + "</td><td> ชม./เดือน </td>"])
            strName = ''.join([strName, "</tr>"])
            index += 1
        strName = ''.join([strName, "</table>"])
    except:
        strName = ''

    data = {
        'instance': strName,
    }
    return JsonResponse(data)

def exportWeightToExpress(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None
    weight_type = request.GET.get('weight_type') or None

    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    if start_created is None and end_created is None:
        start_created = previous_date_time.strftime("%Y-%m-%d")
        end_created = previous_date_time.strftime("%Y-%m-%d")

    my_q = Q()
    if start_created is not None:
        my_q &= Q(date__gte = start_created)
    if end_created is not None:
        my_q &=Q(date__lte = end_created)
    if weight_type is not None:
        my_q &=Q(bws__weight_type = weight_type)

    my_q &=Q(bws__company__code__in = company_in)

    queryset = Weight.objects.filter(my_q)

    if queryset:
        data = {'docid': queryset.values_list('doc_id', flat=True),
                'docdat': queryset.values_list('date', flat=True),
                'datin': queryset.values_list('date_in', flat=True),
                'datout': queryset.values_list('date_out', flat=True),
                'tmin': queryset.values_list('time_in', flat=True),
                'tmout': queryset.values_list('time_out', flat=True),
                'truck': queryset.values_list('car_registration_name', flat=True),
                'cuscod': queryset.values_list('customer_id', flat=True),
                'cusname': queryset.values_list('customer_name', flat=True),
                'depcod': queryset.values_list('base_weight_station_name', flat=True),
                'stkcod': queryset.values_list('stone_type_id', flat=True),
                'stkdes': queryset.values_list('stone_type_name', flat=True),
                'trnqty': queryset.values_list('weight_total', flat=True),
                'unitpr': queryset.values_list('price_per_ton', flat=True),
                'amount': queryset.values_list('amount', flat=True),
                'vat': queryset.values_list('vat', flat=True),
                'stonenam': queryset.values_list('stone_color', flat=True),
                'transport': queryset.values_list('transport', flat=True),
                'oilcuscod': queryset.values_list('car_team__oil_customer_id', flat=True),
                'oilcusnam': queryset.values_list('car_team__car_team_name', flat=True),
                'oillt': queryset.values_list('oil_content', flat=True),
                'nillnam': queryset.values_list('mill_name', flat=True),
                'iscancle': queryset.values_list('is_cancel', flat=True),
                'sttcod': queryset.values_list('base_weight_station_name', flat=True),
                'scaleid': queryset.values_list('scale_id', flat=True),
                'scalenam': queryset.values_list('scale_name', flat=True),
                'scoopnam': queryset.values_list('scoop_name', flat=True),
                'siteid': queryset.values_list('site_id', flat=True),
                'sitenam': queryset.values_list('site_name', flat=True),
                'isvat': queryset.values_list('is_s', flat=True),
                'vattyp': queryset.values_list('vat_type', flat=True),
                'pay': queryset.values_list('pay', flat=True),
                'company': queryset.values_list('bws__company__code', flat=True),
                'bws': queryset.values_list('bws', flat=True),
                'note': queryset.values_list('note', flat=True),
                }
    else:
        data = "ไม่มีข้อมูลรายการชั่งจาก วันที่เลือก"

    df = pd.DataFrame(data)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=weight_express({active}) '+ start_created + " to "+ end_created +'.xlsx'
    return response

def setSessionCompany(request):
    name = request.GET.get('title', None)
    request.session['company_code'] = name
    try:
        company = BaseCompany.objects.get(code=request.session['company_code'])
        request.session['company'] = company.name
    except BaseCompany.DoesNotExist:
        print(f"Company with code {request.session['company_code']} not found.")

    data = {
        'instance': request.session['company_code'],
    }

    return JsonResponse(data)

def setDateInDashbord(request):
    db_start_date = request.GET.get('db_start_date', None)
    db_end_date = request.GET.get('db_end_date', None)

    request.session['db_start_date'] = db_start_date
    request.session['db_end_date'] = db_end_date

    data = {
        'db_start_date' : db_start_date,
        'db_end_date' : db_end_date,
    }
    return JsonResponse(data)

@login_required(login_url='login')
def viewStock(request):
    #active : active คือแท็ปบริษัท active
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')

    data = Stock.objects.filter(company__code__in = company_in).order_by('-created')

    #กรองข้อมูล
    myFilter = StockFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    stock = p.get_page(page)

    context = {'stock_page':'active', 'stock': stock,'filter':myFilter, active :"active",}
    return render(request, "stock/viewStock.html",context)

@login_required(login_url='login')
def removeStock(request, stock_id):
    stk = Stock.objects.get(id = stock_id)

    #ดึงข้อมูล stock ก่อนหน้ามาเพื่อ update วันถัดไป
    previous_day = Stock.objects.filter(
        created__lt = stk.created, company = stk.company
    ).aggregate(max_date=Max('created'))['max_date']
    tmp_company = stk.company

    #ลบ StockStone ใน Stock ด้วย
    ssn = StockStone.objects.filter(stk = stk)
    for sn in ssn:
        items = StockStoneItem.objects.filter(ssn = sn)
        items.delete()

    ssn.delete()
    stk.delete()

    updateTotalStockInMonthByDate(previous_day, tmp_company)#ดึงข้อมูล stock ก่อนหน้ามาเพื่อ update วันถัดไป

    return redirect('viewStock')

@login_required(login_url='login')
def removeStockStone(request, ssn_id):

    #ลบ ProductionLossItem ใน Production ด้วย
    ssn = StockStone.objects.get(id = ssn_id)
    stock_id = ssn.stk

    #ดึงข้อมูล stock ก่อนหน้ามาเพื่อ update วันถัดไป
    previous_day = Stock.objects.filter(
        created__lt = ssn.stk.created, company = ssn.stk.company
    ).aggregate(max_date=Max('created'))['max_date']
    tmp_company = ssn.stk.company

    items = StockStoneItem.objects.filter(ssn = ssn)
    items.delete()

    ssn.delete()

    updateTotalStockInMonthByDate(previous_day, tmp_company)#ดึงข้อมูล stock ก่อนหน้ามาเพื่อ update วันถัดไป
    return HttpResponseRedirect(reverse('editStep2Stock', args=(stock_id,)))

@login_required(login_url='login')
def createStock(request):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    base_stock_source = BaseStockSource.objects.all().order_by('step')
    StockStoneItemFormSet = modelformset_factory(StockStoneItem, fields=('source', 'quantity'), extra=len(base_stock_source), widgets={})
    
    if request.method == 'POST':
        form = StockForm(request.POST)
        ss_form = StockStoneForm(request.POST)
        formset = StockStoneItemFormSet(request.POST)
        if form.is_valid() and ss_form.is_valid() and formset.is_valid():
            form = form.save()

            ssn = ss_form.save()
            ssn.stk = form
            ssn.save()

            total = 0
            formset_instances = formset.save(commit=False)
            for instance in formset_instances:
                instance.ssn = ssn

                if instance.quantity:
                    if instance.source.symbol == "+":
                        total += instance.quantity
                    elif instance.source.symbol == "-":
                        total -= instance.quantity
                else:
                    instance.quantity = 0
                instance.save()

            ssn.total = total
            ssn.save()

            ss_item = StockStoneItem.objects.filter(ssn = ssn.pk)
            for i in ss_item:
                updateTotalStockInMonth(i.id)#คำนวนยอดยกมา จากวันก่อนและคำนวน total stock ใหม่

            return HttpResponseRedirect(reverse('editStep2Stock', args=(ssn.stk,)))
    else:
        form = StockForm(initial={'company': company})
        ss_form = StockStoneForm()
        formset = StockStoneItemFormSet(queryset=StockStoneItem.objects.none())

    context = {'stock_page':'active', 'form': form, 'ss_form': ss_form, 'formset' : formset, 'base_stock_source': base_stock_source, active :"active", 'disabledTab' : 'disabled', 'is_edit_stock': is_edit_stock(request.user)}
    return render(request, "stock/createStock.html",context)

@login_required(login_url='login')
def editStep2Stock(request, stock_id):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    base_stock_source = BaseStockSource.objects.all().order_by('step')
    StockStoneItemFormSet = modelformset_factory(StockStoneItem, fields=('source', 'quantity'), extra=len(base_stock_source), widgets={})
    
    try:
        stock_data = Stock.objects.get(id=stock_id)
    except Stock.DoesNotExist:
        return redirect('viewProduction')

    ssn_data = StockStone.objects.filter(stk=stock_data)

    if request.method == 'POST':
        form = StockForm(request.POST, instance=stock_data)
        ss_form = StockStoneForm(request.POST)
        formset = StockStoneItemFormSet(request.POST)
        
        if form.is_valid() and ss_form.is_valid() and formset.is_valid():
            form = form.save()

            ssn = ss_form.save(commit=False)
            if  ss_form.cleaned_data.get('stone'):
                ssn.stk = form
                ssn.save()

                total = 0
                formset_instances = formset.save(commit=False)
                for instance in formset_instances:
                    instance.ssn = ssn

                    if instance.quantity:
                        if instance.source.symbol == "+":
                            total += instance.quantity
                        elif instance.source.symbol == "-":
                            total -= instance.quantity
                    else:
                        instance.quantity = 0
                    instance.save()

                ssn.total = total
                ssn.save()

            ss_item = StockStoneItem.objects.filter(ssn__stk = stock_id)
            for i in ss_item:
                updateTotalStockInMonth(i.id)#คำนวนยอดยกมา จากวันก่อนและคำนวน total stock ใหม่

            return HttpResponseRedirect(reverse('editStep2Stock', args=(stock_id,)))
    else:
        form = StockForm(instance=stock_data)
        ss_form = StockStoneForm()
        formset = StockStoneItemFormSet(queryset=StockStoneItem.objects.none())

    context = {'stock_page':'active', 'form': form, 'ss_form': ss_form, 'formset' : formset, 'base_stock_source': base_stock_source, 'ssn_data': ssn_data,'stock_data':stock_data, active :"active", 'disabledTab' : 'disabled', 'is_edit_stock': is_edit_stock(request.user)}
    return render(request, "stock/editStep2Stock.html",context)

@login_required(login_url='login')
def editStockStoneItem(request, stock_id, ssn_id):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    base_stock_source = BaseStockSource.objects.all().order_by('step')
    
    try:
        stock_data = Stock.objects.get(id=stock_id)
    except Stock.DoesNotExist:
        return redirect('viewProduction')

    ssn_data = StockStone.objects.filter(stk = stock_id)#ssn all in stock id
    data = StockStone.objects.get(id = ssn_id)#id edit

    if request.method == 'POST':
        form = StockForm(request.POST, instance=stock_data)
        ss_form = StockStoneForm(request.POST, instance=data)
        formset = StockStoneItemInlineFormset(request.POST, instance=data)
        
        if form.is_valid() and ss_form.is_valid() and formset.is_valid():
            form = form.save()

            ssn = ss_form.save(commit=False)
            if  ss_form.cleaned_data.get('stone'):
                ssn.stk = form
                ssn.save()

                formset_instances = formset.save(commit=False)
                for instance in formset_instances: #อันนี้ไม่มี deleted_objects นะ
                    if instance.quantity is None:
                        instance.quantity = 0
                    instance.save()

                # add function calculate total stock
                ssn.total = calculateTotalStock(ssn_id)
                ssn.save()

            
            ss_item = StockStoneItem.objects.filter(ssn__stk = stock_id)
            for i in ss_item:
                updateTotalStockInMonth(i.id)#คำนวนยอดยกมา จากวันก่อนและคำนวน total stock ใหม่

            return HttpResponseRedirect(reverse('editStep2Stock', args=(stock_id,)))
    else:
        form = StockForm(instance=stock_data)
        ss_form = StockStoneForm(instance=data)
        formset = StockStoneItemInlineFormset(instance=data)

    context = {'stock_page':'active', 'form': form, 'ss_form': ss_form, 'formset' : formset, 'base_stock_source': base_stock_source, 'ssn_data': ssn_data, 'ss_id': data.id, 'ss_stone_id': data.stone.base_stone_type_id, 'stock_data':stock_data, active :"active", 'disabledTab' : 'disabled', 'is_edit_stock': is_edit_stock(request.user)}
    return render(request, "stock/editStockStoneItem.html",context)

def calculateTotalStock(ssn_id):
    total = 0
    items = StockStoneItem.objects.filter(ssn = ssn_id)
    for i in items:
        if i.quantity:
            if i.source.symbol == "+":
                total += i.quantity
            elif i.source.symbol == "-":
                total -= i.quantity
    return total

def searchStockInDay(request):
    if 'created' in request.GET and 'company' in request.GET and 'stock_id' in request.GET:
        created =  request.GET.get('created')
        company =  request.GET.get('company')
        stock_id =  request.GET.get('stock_id')

        if stock_id == '':
            have_stock = Stock.objects.filter(company = company, created = created).exists()
        else:
            have_stock = Stock.objects.filter(~Q(id = stock_id), company = company, created = created).exists()
    data = {
        'have_stock' :have_stock,
    }
    return JsonResponse(data)

def searchPortStockInDay(request):
    if 'created' in request.GET and 'company' in request.GET and 'stock_id' in request.GET:
        created =  request.GET.get('created')
        company =  request.GET.get('company')
        stock_id =  request.GET.get('stock_id')

        if stock_id == '':
            have_stock = PortStock.objects.filter(company = company, created = created).exists()
        else:
            have_stock = PortStock.objects.filter(~Q(id = stock_id), company = company, created = created).exists()
    data = {
        'have_stock' :have_stock,
    }
    return JsonResponse(data)

def searchDataWeightToStock(request):
    if 'created' in request.GET and 'company' in request.GET and 'stone' in request.GET:
        created =  request.GET.get('created')
        company =  request.GET.get('company')
        stone =  request.GET.get('stone')

        if stone:
            stone_name = BaseStoneType.objects.get(base_stone_type_id = stone).base_stone_type_name

        sell = 0
        prod = 0
        alert = ""
        #ยกมา
        try:
            latest_date = StockStone.objects.filter(
                stk__created__lt=created, stk__company=company, stone=stone
            ).aggregate(max_date=Max('stk__created'))['max_date']

            # Get the records with that latest date
            quot = StockStone.objects.filter(
                stk__created=latest_date, stk__company=company, stone=stone
            ).values('total').first()['total'] or Decimal('0.0')

        except TypeError or StockStone.DoesNotExist:
            quot = Decimal('0.0')

        #ผลิต
        prod = StoneEstimateItem.objects.filter(se__created = created, stone_type = stone, se__company = company).aggregate(s=Sum("total"))["s"] or Decimal('0.0')

        #ขาย
        sell = Weight.objects.filter(~Q(site = '200PL') & ~Q(site = '300PL'), bws__company = company, bws__weight_type = 1, stone_type = stone, date = created).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')

        #อนุเคราะห์ (ปลายทาง 300PL)
        aid = Weight.objects.filter(bws__company = company, bws__weight_type = 1, stone_type = stone, date = created, site = '300PL').aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')

        if stone:
            if quot == 0:
                alert += "ยกมา : ไม่มียอดยกมาของ "+ str(stone_name) +" วันก่อนหน้านี้<br>"
            if prod == 0:
                alert += "ผลิต : ไม่มีการคีย์ประมาณการณ์หิน หรือไม่มีรายการชั่งผลิตของ "+ str(stone_name) +" ในวันนี้<br>"
            if sell == 0:
                alert += "ขาย : ไม่มีรายการชั่งขายของ "+ str(stone_name) +" ในวันนี้<br>"

    data = {'sell' : sell, 'prod' : prod, 'aid' : aid, 'quot': quot, 'alert' : alert}
    return JsonResponse(data)

def exportExcelStockStoneInDashboard(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    company = BaseCompany.objects.get(code = active)

    end_created = request.session['db_end_date']
    start_created = request.session['db_start_date']

    my_q = Q()
    if company.biz.id == 1: #ธุรกิจเหมือง
        if start_created is not None:
            my_q &= Q(ssn__stk__created__gte = start_created)
        if end_created is not None:
            my_q &=Q(ssn__stk__created__lte = end_created)
        my_q &= Q(ssn__stk__company__code__in = company_in)
    elif company.biz.id == 2: #ธุรกิจท่าเรือ
        if start_created is not None:
            my_q &= Q(pss__ps__created__gte = start_created)
        if end_created is not None:
            my_q &=Q(pss__ps__created__lte = end_created)
        my_q &= Q(pss__ps__company__code__in = company_in)

    #เปลี่ยนออกเป็น ดึงรายงานของเดือนนั้นๆเท่านั้น
    startDate = datetime.strptime(start_created, "%Y-%m-%d").date()
    endDate = datetime.strptime(end_created, "%Y-%m-%d").date()

    #สร้าง list ระหว่าง start_date และ end_date
    list_date = [startDate+timedelta(days=x) for x in range((endDate-startDate).days + 1)]

    response = excelStockStone(request, my_q, list_date)
    return response

def exportExcelStockStone(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    company = BaseCompany.objects.get(code = active)

    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None

    my_q = Q()
    if company.biz.id == 1: #ธุรกิจเหมือง
        if start_created is not None:
            my_q &= Q(ssn__stk__created__gte = start_created)
        if end_created is not None:
            my_q &=Q(ssn__stk__created__lte = end_created)
        my_q &= Q(ssn__stk__company__code__in = company_in)
    elif company.biz.id == 2: #ธุรกิจท่าเรือ
        if start_created is not None:
            my_q &= Q(pss__ps__created__gte = start_created)
        if end_created is not None:
            my_q &=Q(pss__ps__created__lte = end_created)
        my_q &= Q(pss__ps__company__code__in = company_in)
   
    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    startDate = datetime.strptime(start_created or startDateInMonth(previous_date_time.strftime('%Y-%m-%d')), "%Y-%m-%d").date()
    endDate = datetime.strptime(end_created or previous_date_time.strftime('%Y-%m-%d'), "%Y-%m-%d").date()

    #สร้าง list ระหว่าง start_date และ end_date
    list_date = [startDate+timedelta(days=x) for x in range((endDate-startDate).days + 1)]

    response = excelStockStone(request, my_q, list_date)
    return response


def excelStockStone(request, my_q, list_date):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    company = BaseCompany.objects.get(code = active)

    if company.biz.id == 1: #ธุรกิจเหมือง
        data = StockStoneItem.objects.filter(my_q).order_by('ssn__stk__created', 'source__id', 'ssn__stone__base_stone_type_id').values_list('ssn__stk__created', 'ssn__stone__base_stone_type_name', 'source__name', 'quantity', 'ssn__total')
    elif company.biz.id == 2: #ธุรกิจท่าเรือ
        data = PortStockStoneItem.objects.filter(my_q).order_by('pss__ps__created', 'cus__customer_id', 'pss__stone__base_stone_type_id').values_list('pss__ps__created', 'pss__stone__base_stone_type_name', 'cus__customer_name', 'total', 'pss__total')

    # Create a new workbook and get the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    if data:
        worksheet.cell(row=1, column=1, value='Date')
        worksheet.merge_cells(start_row=1, start_column = 1, end_row=2, end_column=1)

        date_style = NamedStyle(name='custom_datetime', number_format='DD/MM/YYYY')
        
        # Create a set of all unique mill and stone values
        stones = set()
        sources = set()
        for item in data:
            stones.add(item[1])
            sources.add(item[2]) 

        stone_col_list = []
        
        # Create a list of colors for each line_type
        stone_colors = [generate_pastel_color() for i  in range(len(stones) + 1)]

        column_index = 2
        for st in stones:
            worksheet.cell(row=1, column=column_index, value=f'Stock {st}')
            worksheet.merge_cells(start_row=1, start_column = column_index, end_row=1, end_column=(column_index + len(sources) + 1) -1 )
            
            cell = worksheet.cell(row=1, column=column_index)
            cell.alignment = Alignment(horizontal='center')

            info = {}
            info['st'] = st
            info['strat_col'] = column_index
            info['end_col'] = column_index + len(sources) + 1
            stone_col_list.append(info)

            #อัพเดทจำนวน col ตามที่มา
            column_index += len(sources) + 1

        #set color in header in row 1-2
        for row in worksheet.iter_rows(min_row=1, max_row=2):
            # Set the background color for each cell in the column
            for cell in row:
                #cell.border = Border(top=side, bottom=side, left=side, right=side)
                cell.alignment = Alignment(horizontal='center')
                line_index = (cell.column - 2) // (len(sources) + 1 )
                fill_color = stone_colors[line_index % len(stone_colors)]
                fill = PatternFill(start_color=fill_color, fill_type="solid")
                cell.fill = fill

        # Write headers row 2 to the worksheet
        column_index = 2
        for st in stones:
            for sou in sources:
                worksheet.cell(row=2, column=column_index, value=sou).alignment = Alignment(horizontal='center')
                column_index += 1
                
            worksheet.cell(row=2, column=column_index, value= 'Total').alignment = Alignment(horizontal='center')
            worksheet.cell(row=2, column=column_index).font = Font(bold=True, color="FF0000")
            column_index += 1


        # Create a dictionary to store data by date, mill, and stone
        date_data = {}

        # Loop through the data and populate the dictionary  
        for item in data:
            date = item[0]
            stone = item[1]
            source = item[2]
            quantity = item[3]
            total = item[4]  # Assuming the 5th column is ssn__total

            if date not in date_data:
                date_data[date] = {}

            if stone not in date_data[date]:
                date_data[date][stone] = {'sources': {}, 'total': 0}

            date_data[date][stone]['sources'][source] = quantity
            date_data[date][stone]['total'] = total  # Store the total for this stone

        row_index = 3
        for idl, ldate in enumerate(list_date):
            #เขียนวันที่ใน worksheet column 1
            worksheet.cell(row=idl+3, column=1, value=ldate).style = date_style
            worksheet.cell(row=idl+3, column=1).alignment = Alignment(horizontal='center')

            for date, stone_data in date_data.items():
                #เขียน weight total ของแต่ละหินใน worksheet
                if worksheet.cell(row=idl+3, column = 1).value == date:
                    column_index = 2
                    for st in stones:
                        source_data = stone_data.get(st, {}).get('sources', {})
                        total_value = stone_data.get(st, {}).get('total', '')

                        # Write quantities by source
                        for sou in sources:
                            value = source_data.get(sou, '')
                            worksheet.cell(row=idl + 3, column=column_index, value=value).number_format = '#,##0.00'
                            column_index += 1

                        # Write the ssn__total value for the stone
                        worksheet.cell(row=idl + 3, column=column_index, value=total_value).number_format = '#,##0.00'
                        worksheet.cell(row=idl + 3, column=column_index).font = Font(bold=True, color="FF0000")
                        column_index += 1
            row_index += 1

        worksheet.cell(row=row_index, column=1, value='รวมทั้งสิ้น')
        sum_by_col = Decimal('0.00')
        for col in range(2, column_index):
            for row in range(3, row_index):
                sum_by_col = sum_by_col + Decimal( worksheet.cell(row=row, column=col).value or '0.00' )
            col_header = worksheet.cell(row=2, column=col).value
            if col_header == 'ยกมา' or col_header == 'Total':#ยกมา และ Total ไม่ต้อง show ใน รวมทั้งสิ้น
                worksheet.cell(row=row_index, column=col, value="")
            else:
                worksheet.cell(row=row_index, column=col, value=sum_by_col).number_format = '#,##0.00'
                worksheet.cell(row=row_index, column=col).font = Font(bold=True, color="FF0000")
            sum_by_col = Decimal('0.00')

        # Set the column widths
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[2].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width
            worksheet.column_dimensions[column].height = 20

        side = Side(border_style='thin', color='000000')
        set_border(worksheet, side)
        worksheet.freeze_panes = "B3" #freeze
    else:
        worksheet.cell(row = 1, column = 1, value = f'ไม่มีข้อมูล Stock หินของเดือนนี้')

    # Save workbook into memory
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    size = output.getbuffer().nbytes

    # Generator to stream file in chunks
    def file_iterator(buffer, chunk_size=8192):
        while True:
            data = buffer.read(chunk_size)
            if not data:
                break
            yield data

    response = StreamingHttpResponse(
        file_iterator(output),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="stock_stone_({active}).xlsx"'
    response["Content-Length"] = str(size)
    return response

@login_required(login_url='login')
def viewGasPrice(request):
    #active : active คือแท็ปบริษัท active
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')
    
    data = GasPrice.objects.filter(company__code__in = company_in).order_by('-created')

    #กรองข้อมูล
    myFilter = GasPriceFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    gas_price = p.get_page(page)

    context = {'ts_page':'active', 'gas_price': gas_price,'filter':myFilter, active :"active",}
    return render(request, "transport/viewGasPrice.html", context)

def createGasPrice(request):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    if request.method == 'POST':
        form = GasPriceForm(request.POST)
        if form.is_valid():
            gp = form.save()

            calculateTotalGasPriceById(gp.pk) #คำนวณราคาต้นทุน และราคาขายรวมของวันนั้นๆ
            calculateGasPriceInWeight(gp.pk) #คำนวณราคาต้นทุน และราคาขายรวมของแต่ละ weight id
            return redirect('viewGasPrice')
    else:
        form = GasPriceForm(initial={'company': company})

    context = {'ts_page':'active', 'form': form, active :"active", 'disabledTab' : 'disabled'}
    return render(request, "transport/createGasPrice.html",context)

def editGasPrice(request, gp_id):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    data = GasPrice.objects.get(id = gp_id)#id edit
    #เช็คข้อมูล หากเปลี่ยนวันที่
    created_old = data.created
    if request.method == 'POST':
        form = GasPriceForm(request.POST, instance=data)
        if form.is_valid():
            gp = form.save()

            calculateTotalGasPriceById(gp.pk) #คำนวณราคาต้นทุน และราคาขายรวมของวันนั้นๆ
            calculateGasPriceInWeight(gp.pk) #คำนวณราคาต้นทุน และราคาขายรวมของแต่ละ weight id

            checkDateChangeGasPrice(created_old, gp.pk)#ถ้าเปลี่ยนวันที่ oil_cost, oil_sell ของวันก่อนต้องเป็น 0
            return redirect('viewGasPrice')
    else:
        form = GasPriceForm(instance=data)

    context = {'ts_page':'active', 'form': form, 'gp': data , active :"active", 'disabledTab' : 'disabled'}
    return render(request, "transport/editGasPrice.html",context)

def checkDateChangeGasPrice(created_old, gp_id):
    gp = GasPrice.objects.get(id = gp_id)
    if created_old != gp.created:
        try:
            weight = Weight.objects.filter(date = created_old, bws__weight_type = 1, bws__company = gp.company, oil_content__gt = 0)
            #weight.update(oil_cost = 0, oil_sell = 0) เอาราคาทุนน้ำมัน ออกก่อน 13/02/2025
            weight.update(oil_sell = 0)
        except Weight.DoesNotExist:
            pass

def calculateTotalGasPriceById(gp_id):
    try:
        gp = GasPrice.objects.get(id = gp_id)
        sum_oil = Weight.objects.filter(date = gp.created, bws__weight_type = 1, bws__company = gp.company).aggregate(s=Sum("oil_content"))["s"] or Decimal('0.0')
        #gp.total_cost = gp.cost * sum_oil เอาราคาทุนน้ำมัน ออกก่อน 13/02/2025
        gp.total_sell = gp.sell * sum_oil
        gp.save()
    except GasPrice.DoesNotExist or Weight.DoesNotExist:
        pass

def calculateGasPriceInWeight(gp_id):
    try:
        gp = GasPrice.objects.get(id = gp_id)
        oil = Weight.objects.filter(date = gp.created, bws__weight_type = 1, bws__company = gp.company, oil_content__gt = 0)
        for ol in oil:
            #ol.oil_cost = ol.oil_content * gp.cost เอาราคาทุนน้ำมัน ออกก่อน 13/02/2025
            ol.oil_sell = ol.oil_content * gp.sell
            ol.save()
    except GasPrice.DoesNotExist or Weight.DoesNotExist:
        pass

def removeGasPrice(request, gp_id):
    gp = GasPrice.objects.get(id = gp_id)

    #set oil_cost และ oil_sell = 0 กรณีลบ GasPrice
    try:
        weight = Weight.objects.filter(date = gp.created, bws__weight_type = 1, bws__company = gp.company, oil_content__gt = 0)
        #weight.update(oil_cost = 0, oil_sell = 0) เอาราคาทุนน้ำมัน ออกก่อน 13/02/2025
        weight.update(oil_sell = 0)
    except Weight.DoesNotExist:
        pass

    gp.delete()
    return redirect('viewGasPrice')

def searchGasPrice(request):
    if 'created' in request.GET and 'gp_id' in request.GET and 'company' in request.GET:
        created =  request.GET.get('created')
        gp_id =  request.GET.get('gp_id')
        company =  request.GET.get('company')

        #if gp_id == '' create mode , else edit mode
        if gp_id == '':
            have_gas_price = GasPrice.objects.filter(company__code = company, created = created).exists()
        else:
            have_gas_price = GasPrice.objects.filter(~Q(id = gp_id),company__code = company, created = created).exists()
        #ดึง cost และ sell
        gp = GasPrice.objects.filter(company__code = company).order_by('-created').values('cost', 'sell').first()
        
    data = {
        'have_gas_price' :have_gas_price,
        'cost': gp['cost'],
        'sell': gp['sell'],
    }
    
    return JsonResponse(data)

def exportExcelGasPriceTransport(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None

    date_object = datetime.today()
    previous_date_time = date_object - timedelta(days=1)

    if end_created is None:
        end_created = previous_date_time.strftime('%Y-%m-%d')
    if start_created is None:
        start_created = startDateInMonth(end_created)

    my_q = Q()
    if start_created is not None:
        my_q &= Q(date__gte=start_created)
    if end_created is not None:
        my_q &= Q(date__lte=end_created)

    my_q &= Q(oil_content__gt=0, bws__weight_type=1, bws__company__code__in=company_in)

    queryset = Weight.objects.filter(my_q).values(
        'car_team__car_team_name', 'customer__customer_name', 'car_registration_name',
        'mill__mill_name', 'site__base_site_name', 'stone_type__base_stone_type_name'
    ).annotate(
        num_rows=Count('weight_id'), 
        sum_weight=Sum('weight_total'),
        sum_oil=Sum('oil_content'),
        price_per_unit=ExpressionWrapper(
            Sum('oil_sell') / Sum('oil_content'),
            output_field = models.DecimalField()
        ), sum_oil_sell=Sum('oil_sell')
    ).order_by('car_team__car_team_name')
    
    df = pd.DataFrame(list(queryset))

    df.columns = [
        'ทีม', 'ลูกค้า', 'ทะเบียน', 'ต้นทาง', 'ปลายทาง', 'ชนิดหิน', 
        'จำนวนเที่ยว', 'น้ำหนักรวม (ตัน)', 'น้ำมันรวม (ลิตร)',
        'ราคาขายต่อหน่วย', 'ยอดขาย'
    ]
    
    df.fillna({'ทีม': '(ไม่มีทีม)'}, inplace=True)

    grouped = df.groupby('ทีม', dropna=False)
    result = []

    for name, group in grouped:
        result.append(group)
        subtotal = pd.DataFrame({
            'ทีม': [f'รวมทีม {name}'],
            'ลูกค้า': [''], 'ทะเบียน': [''], 'ต้นทาง': [''], 'ปลายทาง': [''], 'ชนิดหิน': [''],
            'จำนวนเที่ยว': [group['จำนวนเที่ยว'].sum()],
            'น้ำหนักรวม (ตัน)': [group['น้ำหนักรวม (ตัน)'].sum()],
            'น้ำมันรวม (ลิตร)': [group['น้ำมันรวม (ลิตร)'].sum()],
            'ราคาขายต่อหน่วย': [group['ราคาขายต่อหน่วย'].sum()],
            'ยอดขาย': [group['ยอดขาย'].sum()],
        })
        result.append(subtotal)

    df = pd.concat(result, ignore_index=True)

    total_row = pd.DataFrame({
        'ทีม': ['รวมทั้งหมด'], 'ลูกค้า': [''], 'ทะเบียน': [''], 'ต้นทาง': [''], 'ปลายทาง': [''], 'ชนิดหิน': [''],
        'จำนวนเที่ยว': [df.loc[df['ทีม'].str.contains('รวมทีม', na=False) == False, 'จำนวนเที่ยว'].sum()],
        'น้ำหนักรวม (ตัน)': [df.loc[df['ทีม'].str.contains('รวมทีม', na=False) == False, 'น้ำหนักรวม (ตัน)'].sum()],
        'น้ำมันรวม (ลิตร)': [df.loc[df['ทีม'].str.contains('รวมทีม', na=False) == False, 'น้ำมันรวม (ลิตร)'].sum()],
        'ราคาขายต่อหน่วย': [df.loc[df['ทีม'].str.contains('รวมทีม', na=False) == False, 'ราคาขายต่อหน่วย'].sum()],
        'ยอดขาย': [df.loc[df['ทีม'].str.contains('รวมทีม', na=False) == False, 'ยอดขาย'].sum()],
    })

    df = pd.concat([df, total_row], ignore_index=True)

    df[['น้ำหนักรวม (ตัน)']] = df[[
        'น้ำหนักรวม (ตัน)'
    ]].applymap(lambda x: f"{x:,.3f}" if pd.notna(x) else "")

    df[['น้ำมันรวม (ลิตร)', 'ยอดขาย']] = df[[
        'น้ำมันรวม (ลิตร)', 'ยอดขาย'
    ]].applymap(lambda x: f"{x:,.2f}" if pd.notna(x) else "")

    df[['ราคาขายต่อหน่วย']] = df[[
        'ราคาขายต่อหน่วย'
    ]].applymap(lambda x: f"{x:,.4f}" if pd.notna(x) else "")

    # Create an Excel response with openpyxl
    output = BytesIO()
    output.seek(0)
	
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=GasPriceTransport({active}) {start_created} to {end_created}.xlsx'

    # Write to Excel with openpyxl engine
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

        # Get the workbook and sheet
        workbook = writer.book
        sheet = workbook.active

        # Bold subtotal and total row
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                if 'รวมทีม' in str(cell.value) or 'รวมทั้งหมด' in str(cell.value):
                    row_number = cell.row  # Get the row number
                    # Bold and red font
                    bold_red_font = Font(bold=True, color="FF0000")

                    # Apply to columns A, G, H, I, J, K in the same row
                    for col in ['A', 'G', 'H', 'I', 'J', 'K']:
                        sheet[f"{col}{row_number}"].font = bold_red_font


        right_align = Alignment(horizontal="right")
        for col in ['H', 'I', 'J', 'K']:  # Columns for numbers
            for cell in sheet[col]:  # Iterate through all cells in that column
                cell.alignment = right_align

    return response


def exportWeightFixBug(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None
    weight_type = request.GET.get('weight_type') or None

    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    if start_created is None and end_created is None:
        start_created = previous_date_time.strftime("%Y-%m-%d")
        end_created = previous_date_time.strftime("%Y-%m-%d")

    my_q = Q()
    if start_created is not None:
        my_q &= Q(date__gte = start_created)
    if end_created is not None:
        my_q &=Q(date__lte = end_created)
    if weight_type is not None:
        my_q &=Q(bws__weight_type = weight_type)

    my_q &=Q(bws__company__code__in = company_in)

    queryset = Weight.objects.filter(Q(mill_id__isnull = False) & my_q).exclude(mill_name=F('mill__mill_name'))
    if not queryset.exists():
        return HttpResponse("No data to export.")

    data = {'weight_id': queryset.values_list('weight_id', flat=True),
            'docid': queryset.values_list('doc_id', flat=True),
            'docdat': queryset.values_list('date', flat=True),
            'local_mill_name': queryset.values_list('mill_name', flat=True),
            'wrong_mill_id': queryset.values_list('mill_id', flat=True),
            'center_mill_name': queryset.values_list('mill__mill_name', flat=True),
            'bws': queryset.values_list('bws', flat=True),
            }

    df = pd.DataFrame(data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=weight_fix_bug({active}) '+ start_created + " to "+ end_created +'.xlsx'

    df.to_excel(response, index=False, engine='openpyxl')

    return response


def exportWeightHistoryFixBug(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None
    weight_type = request.GET.get('weight_type') or None

    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    if start_created is None and end_created is None:
        start_created = previous_date_time.strftime("%Y-%m-%d")
        end_created = previous_date_time.strftime("%Y-%m-%d")

    my_q = Q()
    if start_created is not None:
        my_q &= Q(date__gte = start_created)
    if end_created is not None:
        my_q &=Q(date__lte = end_created)
    if weight_type is not None:
        my_q &=Q(bws__weight_type = weight_type)

    my_q &=Q(bws__company__code__in = company_in)

    queryset = WeightHistory.objects.filter(Q(mill_id__isnull = False) & my_q).exclude(mill_name=F('mill__mill_name'))
    if not queryset.exists():
        return HttpResponse("No data to export.")

    data = {'weight_id': queryset.values_list('weight_id', flat=True),
            'docid': queryset.values_list('doc_id', flat=True),
            'docdat': queryset.values_list('date', flat=True),
            'local_mill_name': queryset.values_list('mill_name', flat=True),
            'wrong_mill_id': queryset.values_list('mill_id', flat=True),
            'center_mill_name': queryset.values_list('mill__mill_name', flat=True),
            'bws': queryset.values_list('bws', flat=True),
            }

    df = pd.DataFrame(data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=weight_history_fix_bug({active}) '+ start_created + " to "+ end_created +'.xlsx'

    df.to_excel(response, index=False, engine='openpyxl')

    return response

def searchDataWeightInDay(request):
    alert = ""
    if 'created' in request.GET and 'company' in request.GET:
        created = request.GET.get('created')
        company = request.GET.get('company')
        mode = request.GET.get('mode')

        if mode == '1':
            have_weight = Weight.objects.filter(date = created, bws__weight_type = 1, bws__company__code = company).exists()
            if not have_weight:
                alert = "ไม่มีรายการชั่งขายของวันนี้ กรุณา uploade รายการชั่งของวันที่ "+ str(created) + " มาก่อน"
        if mode == '2':
            have_weight = Weight.objects.filter(date = created, bws__weight_type = 2, bws__company__code = company).exists()
            if not have_weight:
                alert = "ไม่มีรายการชั่งผลิตวันนี้ กรุณา uploade รายการชั่งของวันที่ "+ str(created) + " มาก่อน"

    data = {'alert' : alert, 'have_weight': have_weight,}
    return JsonResponse(data)

#ค้นหา น้ำหนักจากตาชั่งเข้าโรงโม่ และ จากโรงโม่อืน เพื่อ Estimate
def searchWeightBySite(request):
    scale = Decimal('0.0')
    other = Decimal('0.0')

    if 'created' in request.GET and 'company' in request.GET and 'site' in request.GET:
        created = request.GET.get('created')
        company = request.GET.get('company')
        site = request.GET.get('site')

        #จากตาชั่งเข้าโรงโม่
        scale = Weight.objects.filter(date = created, bws__weight_type = 2, bws__company__code = company, site = site).aggregate(s=Sum("weight_total"))["s"] or Decimal('0.0')
        #จากโรงโม่อืน
        other =  StoneEstimateItem.objects.filter(se__created = created, se__company__code = company, site_id = site).aggregate(s=Sum("qty_site"))["s"] or Decimal('0.0')
        if other == 0.0:
            other =  StoneEstimateItem.objects.filter(se__created = created, se__company__code = company, nd_site_id = site).aggregate(s=Sum("nd_qty_site"))["s"] or Decimal('0.0')

    data = {'scale': scale, 'other': other}
    return JsonResponse(data)

def exportExcelPercentEstimate(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None
    site = request.GET.get('site') or None

    my_q = Q()
    if start_created is not None:
        my_q &= Q(se__created__gte = start_created)
    if end_created is not None:
        my_q &=Q(se__created__lte = end_created)
    if site is not None:
        my_q &=Q(se__site = site)

    my_q &= Q(se__company__code__in = company_in)
   
    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    startDate = datetime.strptime(start_created or startDateInMonth(previous_date_time.strftime('%Y-%m-%d')), "%Y-%m-%d").date()
    endDate = datetime.strptime(end_created or previous_date_time.strftime('%Y-%m-%d'), "%Y-%m-%d").date()

    #สร้าง list ระหว่าง start_date และ end_date
    list_date = [startDate+timedelta(days=x) for x in range((endDate-startDate).days + 1)]

    response = excelPercentEstimate(request, my_q, list_date)
    return response

def excelPercentEstimate(request, my_q, list_date):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    stone_id = StoneEstimateItem.objects.filter(Q(percent__gt = 0) & my_q).values_list('stone_type__base_stone_type_id', flat=True).order_by('stone_type__base_stone_type_id').distinct() #ดึงเฉพาะ stone_id ที่มีการคีย์ percent > 0
    data = StoneEstimateItem.objects.filter(Q(stone_type__in = stone_id) & my_q).order_by('se__created', 'se__site', 'stone_type').values_list('se__created', 'se__site__base_site_name', 'stone_type__base_stone_type_name', 'percent')

    # Create a new workbook and get the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    if data:
        worksheet.cell(row=1, column=1, value='Date')
        worksheet.merge_cells(start_row=1, start_column = 1, end_row=2, end_column=1)

        date_style = NamedStyle(name='custom_datetime', number_format='DD/MM/YYYY')
        
        # Create a set of all unique mill and stone values
        sites = set()
        stones = set()
        for item in data:
            sites.add(item[1])
            stones.add(item[2]) 

        site_col_list = []
        
        # Create a list of colors for each line_type
        site_colors = [generate_pastel_color() for i  in range(len(sites) + 1)]

        column_index = 2
        for st in sites:
            worksheet.cell(row=1, column=column_index, value=f'เปอร์เซ็นต์การผลิตหิน {st}')
            worksheet.merge_cells(start_row=1, start_column = column_index, end_row=1, end_column=(column_index + len(stones) + 1) -1 )
            
            cell = worksheet.cell(row=1, column=column_index)
            cell.alignment = Alignment(horizontal='center')

            info = {}
            info['st'] = st
            info['strat_col'] = column_index
            info['end_col'] = column_index + len(stones) + 1 
            site_col_list.append(info)

            #อัพเดทจำนวน col ตามที่มา
            column_index += len(stones) + 1

        #set color in header in row 1-2
        for row in worksheet.iter_rows(min_row=1, max_row=2):
            # Set the background color for each cell in the column
            for cell in row:
                #cell.border = Border(top=side, bottom=side, left=side, right=side)
                cell.alignment = Alignment(horizontal='center')
                line_index = (cell.column - 2) // (len(stones) + 1)
                fill_color = site_colors[line_index % len(site_colors)]
                fill = PatternFill(start_color=fill_color, fill_type="solid")
                cell.fill = fill

        # Write headers row 2 to the worksheet
        column_index = 2
        for st in sites:
            for sou in stones:
                worksheet.cell(row=2, column=column_index, value=sou).alignment = Alignment(horizontal='center')
                column_index += 1
            
            worksheet.cell(row=2, column=column_index, value= 'Total').alignment = Alignment(horizontal='center')
            worksheet.cell(row=2, column=column_index).font = Font(bold=True, color="FF0000")
            column_index += 1
                
        # Create a dictionary to store data by date, mill, and stone
        date_data = {}

        # Loop through the data and populate the dictionary  
        for item in data:
            date = item[0]
            site = item[1]
            stone = item[2]
            value = item[3]

            if date not in date_data:
                date_data[date] = {}

            if site not in date_data[date]:
                date_data[date][site] = {}

            date_data[date][site][stone] = value

        row_index = 3
        for idl, ldate in enumerate(list_date):
            #เขียนวันที่ใน worksheet column 1
            worksheet.cell(row=idl+3, column=1, value=ldate).style = date_style
            worksheet.cell(row=idl+3, column=1).alignment = Alignment(horizontal='center')

            for date, site_data in date_data.items():
                #เขียน weight total ของแต่ละหินใน worksheet
                if worksheet.cell(row=idl+3, column = 1).value == date:
                    column_index = 2
                    for site in sites:
                        sum_value = 0
                        stone_data = site_data.get(site, {})
                        for stone in stones:
                            value = stone_data.get(stone, '')
                            try:
                                value = Decimal(value) / 100
                                sum_value += value
                                cell = worksheet.cell(row=idl+3, column=column_index, value=value)
                                cell.number_format = '0%'
                            except (InvalidOperation, TypeError):
                                worksheet.cell(row=idl+3, column=column_index, value='')
                            column_index += 1
                        
                        worksheet.cell(row=idl + 3, column=column_index, value = sum_value).number_format = '0%'
                        worksheet.cell(row=idl + 3, column=column_index).font = Font(bold=True, color="FF0000")
                        column_index += 1
                    #row_index += 1
            row_index += 1

        # Set the column widths
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[2].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            if column == 'A':
                worksheet.column_dimensions[column].width = 15
            else:
                worksheet.column_dimensions[column].width = adjusted_width
            worksheet.column_dimensions[column].height = 20

        side = Side(border_style='thin', color='000000')
        set_border(worksheet, side)
        worksheet.freeze_panes = "B3" #freeze
    else:
        worksheet.cell(row = 1, column = 1, value = f'ไม่มีข้อมูลผลิตหินประจำวันของเดือนนี้')

    # Save workbook into memory
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    size = output.getbuffer().nbytes

    # Generator to stream file in chunks
    def file_iterator(buffer, chunk_size=8192):
        while True:
            data = buffer.read(chunk_size)
            if not data:
                break
            yield data

    response = StreamingHttpResponse(
        file_iterator(output),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="percent_estimate({active}).xlsx"'
    response["Content-Length"] = str(size)

    # Save the workbook to the response
    #workbook.save(response)
    return response

@login_required(login_url='login')
def viewPortStock(request):
    #active : active คือแท็ปบริษัท active
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')

    data = PortStock.objects.filter(company__code__in = company_in).order_by('-created')

    #กรองข้อมูล
    myFilter = PortStockFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    stock = p.get_page(page)

    context = {'port_stock_page':'active', 'stock': stock,'filter':myFilter, active :"active",}
    return render(request, "portStock/viewPortStock.html",context)

@login_required(login_url='login')
def createPortStock(request):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    cus_qs = BaseCustomer.objects.filter(is_port_stock = True).values('customer_id')

    PortStockStoneItemFormSet = modelformset_factory(PortStockStoneItem, fields=('cus', 'quoted', 'receive', 'pay', 'loss', 'other', 'sell_cus', 'total'), extra=len(cus_qs),)
    
    if request.method == 'POST':
        form = PortStockForm(request.POST)
        ss_form = PortStockStoneForm(request.POST)
        formset = PortStockStoneItemFormSet(request.POST)
        if form.is_valid() and ss_form.is_valid() and formset.is_valid():
            form = form.save()

            pss = ss_form.save()
            pss.ps = form
            pss.save()

            formset_instances = formset.save(commit=False)
            for instance in formset_instances:
                instance.pss = pss
                instance.save()

            psi = PortStockStoneItem.objects.filter(pss = pss.pk)
            for i in psi:
                updateTotalPortStockInMonth(i.id)#คำนวนยอดยกมา จากวันก่อนและคำนวน total stock ใหม่

            return HttpResponseRedirect(reverse('editStep2PortStock', args=(pss.ps,)))
    else:
        form = PortStockForm(initial={'company': company})
        ss_form = PortStockStoneForm()
        formset = PortStockStoneItemFormSet(queryset=PortStockStoneItem.objects.none())

    context = {'port_stock_page':'active', 'form': form, 'ss_form': ss_form, 'formset' : formset, 'cus_qs': cus_qs, active :"active", 'disabledTab' : 'disabled', 'is_edit_stock': is_edit_stock(request.user)}
    return render(request, "portStock/createPortStock.html", context)

@login_required(login_url='login')
def editStep2PortStock(request, stock_id):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    cus_qs = BaseCustomer.objects.filter(is_port_stock = True).values('customer_id')

    PortStockStoneItemFormSet = modelformset_factory(PortStockStoneItem, fields=('cus', 'quoted', 'receive', 'pay', 'loss', 'other', 'sell_cus', 'total'), extra=len(cus_qs),)
    
    try:
        stock_data = PortStock.objects.get(id=stock_id)
    except PortStock.DoesNotExist:
        return redirect('viewPortStock')

    ssn_data = PortStockStone.objects.filter(ps=stock_data)

    if request.method == 'POST':
        form = PortStockForm(request.POST, instance=stock_data)
        ss_form = PortStockStoneForm(request.POST)
        formset = PortStockStoneItemFormSet(request.POST)
        
        if form.is_valid() and ss_form.is_valid() and formset.is_valid():
            form = form.save()

            pss = ss_form.save(commit=False)
            if  ss_form.cleaned_data.get('stone'):
                pss.ps = form
                pss.save()

                formset_instances = formset.save(commit=False)
                for instance in formset_instances:
                    instance.pss = pss
                    instance.save()

            psi = PortStockStoneItem.objects.filter(pss__ps = stock_id)
            for i in psi:
                updateTotalPortStockInMonth(i.id)#คำนวนยอดยกมา จากวันก่อนและคำนวน total stock ใหม่

            return HttpResponseRedirect(reverse('editStep2PortStock', args=(stock_id,)))
    else:
        form = PortStockForm(instance=stock_data)
        ss_form = PortStockStoneForm()
        formset = PortStockStoneItemFormSet(queryset=PortStockStoneItem.objects.none())

    context = {'port_stock_page':'active', 'form': form, 'ss_form': ss_form, 'formset' : formset, 'cus_qs': cus_qs, 'ssn_data': ssn_data,'stock_data':stock_data, active :"active", 'disabledTab' : 'disabled', 'is_edit_stock': is_edit_stock(request.user)}
    return render(request, "portStock/editStep2PortStock.html",context)

@login_required(login_url='login')
def editPortStockStoneItem(request, stock_id, pss_id):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    cus_qs = BaseCustomer.objects.filter(is_port_stock = True).values('customer_id')
    
    try:
        stock_data = PortStock.objects.get(id=stock_id)
    except PortStock.DoesNotExist:
        return redirect('viewPortStock')

    ssn_data = PortStockStone.objects.filter(ps = stock_id)#ssn all in stock id
    data = PortStockStone.objects.get(id = pss_id)#id edit

    old_pay = Weight.objects.filter(stone_type = data.stone, site__store = 3, bws__company = company, bws__weight_type = 1, date = data.ps.created).aggregate(total=Sum("weight_total"))['total'] or Decimal('0.00')

    if request.method == 'POST':
        form = PortStockForm(request.POST, instance=stock_data)
        ss_form = PortStockStoneForm(request.POST, instance=data)
        formset = PortStockStoneItemInlineFormset(request.POST, instance=data)
        
        if form.is_valid() and ss_form.is_valid() and formset.is_valid():
            form = form.save()

            ssn = ss_form.save(commit=False)
            if  ss_form.cleaned_data.get('stone'):
                ssn.ps = form
                ssn.save()

                formset_instances = formset.save(commit=False)
                for instance in formset_instances: #อันนี้ไม่มี deleted_objects นะ
                    if instance.quoted is None:
                        instance.quoted = 0
                    instance.save()

            psi = PortStockStoneItem.objects.filter(pss__ps = stock_id)
            for i in psi:
                updateTotalPortStockInMonth(i.id)#คำนวนยอดยกมา จากวันก่อนและคำนวน total stock ใหม่

            return HttpResponseRedirect(reverse('editStep2PortStock', args=(stock_id,)))
    else:
        form = PortStockForm(instance=stock_data)
        ss_form = PortStockStoneForm(instance=data)
        formset = PortStockStoneItemInlineFormset(instance=data)

    context = {'stock_page':'active', 'form': form, 'ss_form': ss_form, 'formset' : formset, 'base_stock_source': cus_qs, 'ssn_data': ssn_data, 'ss_id': data.id, 'ss_stone_id': data.stone.base_stone_type_id, 'stock_data':stock_data, active :"active", 'disabledTab' : 'disabled', 'is_edit_stock': is_edit_stock(request.user), 'old_pay': old_pay}
    return render(request, "portStock/editPortStockStoneItem.html",context)

@login_required(login_url='login')
def removePortStock(request, stock_id):
    ps = PortStock.objects.get(id = stock_id)

    #ดึงข้อมูล stock ก่อนหน้ามาเพื่อ update วันถัดไป
    previous_day = PortStock.objects.filter(
        created__lt = ps.created, company = ps.company
    ).aggregate(max_date=Max('created'))['max_date']
    tmp_company = ps.company

    #ลบ StockStone ใน Stock ด้วย
    pss = PortStockStone.objects.filter(ps = ps)
    for i in pss:
        items = PortStockStoneItem.objects.filter(pss = i)
        items.delete()

    pss.delete()
    ps.delete()

    updateTotalPortStockInMonthByDate(previous_day, tmp_company)#ดึงข้อมูล stock ก่อนหน้ามาเพื่อ update วันถัดไป

    return redirect('viewPortStock')

@login_required(login_url='login')
def removePortStockStone(request, pss_id):

    #ลบ ProductionLossItem ใน Production ด้วย
    pss = PortStockStone.objects.get(id = pss_id)
    stock_id = pss.ps

    #ดึงข้อมูล stock ก่อนหน้ามาเพื่อ update วันถัดไป
    previous_day = PortStock.objects.filter(
        created__lt = pss.ps.created, company = pss.ps.company
    ).aggregate(max_date=Max('created'))['max_date']
    tmp_company = pss.ps.company

    items = PortStockStoneItem.objects.filter(pss = pss)
    items.delete()

    pss.delete()

    updateTotalPortStockInMonthByDate(previous_day, tmp_company)#ดึงข้อมูล stock ก่อนหน้ามาเพื่อ update วันถัดไป
    return HttpResponseRedirect(reverse('editStep2PortStock', args=(stock_id,)))


def searchDataWeightToPortStock(request):
    if 'created' in request.GET and 'company' in request.GET and 'stone' in request.GET:
        created =  request.GET.get('created')
        company =  request.GET.get('company')
        stone =  request.GET.get('stone')

        if stone:
            stone_name = BaseStoneType.objects.get(base_stone_type_id = stone).base_stone_type_name

        cus_id = BaseCustomer.objects.filter(is_port_stock = True).values_list('customer_id', flat=True)

        alert = ""

        #ยกมา
        try:
            latest_date = PortStockStone.objects.filter(
                ps__created__lt=created, ps__company=company, stone=stone
            ).aggregate(max_date=Max('ps__created'))['max_date']

            # Get the records with that latest date
            quot = PortStockStoneItem.objects.filter(
               pss__ps__created=latest_date, pss__ps__company=company, pss__stone=stone
            ).values('cus__customer_id', 'total')

        except TypeError or StockStone.DoesNotExist:
            quot = PortStockStoneItem.objects.none()

        #รับเข้า
        receive = Weight.objects.filter(mill__isnull = True, stone_type = stone, customer__in = cus_id, line_type = "สายยาว", bws__company = company, bws__weight_type = 1, date = created).values('customer__customer_id').annotate(total=Sum("weight_total"))

        #จ่ายภายลงเรือ ดึงจำนวนที่ลงเรือทั้งหมดตามชนิดหิน
        pay = Weight.objects.filter(stone_type = stone, site__store = 3, bws__company = company, bws__weight_type = 1, date = created).aggregate(total=Sum("weight_total"))['total'] or Decimal('0.00')

        if stone:
            if not quot:
                alert += "ยกมา : ไม่มียอดยกมาของ "+ str(stone_name) +" วันก่อนหน้านี้<br>"
            if not receive:
                alert += "รับเข้า : ไม่มีการ uplode รายการชั่ง หรือไม่มีการรับเข้าของ "+ str(stone_name) +" ในวันนี้<br>"

    data = {'list_quot': list(quot), 'list_receive': list(receive), 'pay': pay, 'alert' : alert}
    return JsonResponse(data)

def exportExcelTransport(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    company = BaseCompany.objects.get(code = active)

    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None

    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    if start_created is None and end_created is None:
        start_created = previous_date_time.strftime("%Y-%m-%d")
        end_created = previous_date_time.strftime("%Y-%m-%d")

    my_q = Q()
    if start_created is not None:
        my_q &= Q(date__gte=start_created)
    if end_created is not None:
        my_q &= Q(date__lte=end_created)

    my_q &= Q(car_team__isnull = False, bws__weight_type=1, bws__company__code__in=company_in)

    queryset = Weight.objects.filter(my_q).annotate(
        weight_range=Case(
            When(weight_total__lt=35, then=Value('น้อยกว่า 35')),
            When(weight_total__gte=35, weight_total__lt=40, then=Value('แบก 35 - 40')),
            When(weight_total__gte=40, then=Value('แบก 40 up')),
            output_field=models.CharField(),
        )
    ).values(
        'car_team__car_team_name',
        'customer__customer_name',
        'bws__company__name',
        'weight_range',
        'stone_type__base_stone_type_name',
    ).annotate(
        num_rows=Count('weight_id'),
        ori_sum_weight=Sum('origin_weight'),
        sum_weight=Sum('weight_total'),
    ).annotate(
        pay_weight=Sum(
            Case(
                When(origin_weight__lt=F('weight_total'), then=F('origin_weight')),
                default = F('weight_total'),
                output_field = models.DecimalField(),
            )
        ),
        diff_weight = ExpressionWrapper(Sum('weight_total') - Sum('origin_weight'), output_field= models.DecimalField()) 
    ).order_by('car_team__car_team_name')
    

    if not queryset.exists():
        return HttpResponse("No data to export.")

    # Convert to DataFrame
    df = pd.DataFrame.from_records(queryset)

    # Rename fields
    df.rename(columns={
        'car_team__car_team_name': 'ทีม',
        'customer__customer_name': 'ลูกค้า',
        'bws__company__name': 'บริษัท',
        'stone_type__base_stone_type_name': 'ชนิดหิน',
        'num_rows': 'จำนวนเที่ยว',
        'ori_sum_weight': 'น้ำหนักต้นทางรวม (ตัน)',
        'sum_weight': 'น้ำหนักปลายทางรวม (ตัน)',
        'pay_weight': 'น้ำหนักที่จ่าย (ตัน)',
        'diff_weight': 'ส่วนต่าง (ตัน)',
        'weight_range': 'ช่วงน้ำหนัก'
    }, inplace=True)

    # Add column "ต้นทาง - ปลายทาง"
    if company.biz.id == 1: #ธุรกิจเหมือง
        df['ต้นทาง - ปลายทาง'] = df['บริษัท'] + ' - ' + df['ลูกค้า']
    elif company.biz.id == 2: #ธุรกิจท่าเรือ
        df['ต้นทาง - ปลายทาง'] = df['ลูกค้า'] + ' - ' + df['บริษัท']

    # Keep only needed columns
    df = df[['ทีม', 'ต้นทาง - ปลายทาง', 'ช่วงน้ำหนัก', 'ชนิดหิน', 'จำนวนเที่ยว', 'น้ำหนักต้นทางรวม (ตัน)', 'น้ำหนักปลายทางรวม (ตัน)', 'น้ำหนักที่จ่าย (ตัน)', 'ส่วนต่าง (ตัน)']]
    df['ทีม'].fillna('(ไม่มีทีม)', inplace=True)

    # Group by ทีม
    grouped = df.groupby('ทีม', dropna=False)
    result = []

    for name, group in grouped:
        team_header = pd.DataFrame({
            'ทีม': [name],
            'ต้นทาง - ปลายทาง': [''],
            'ช่วงน้ำหนัก': [''],
            'ชนิดหิน': [''],
            'จำนวนเที่ยว': [''],
            'น้ำหนักต้นทางรวม (ตัน)': [''],
            'น้ำหนักปลายทางรวม (ตัน)': [''],
            'น้ำหนักที่จ่าย (ตัน)': [''],
            'ส่วนต่าง (ตัน)': [''],
        })
        result.append(team_header)
        result.append(group.assign(ทีม=''))

        subtotal_row = pd.DataFrame({
            'ทีม': [f'รวมทีม {name}'],
            'ต้นทาง - ปลายทาง': [''],
            'ช่วงน้ำหนัก': [''],
            'ชนิดหิน': [''],
            'จำนวนเที่ยว': [group['จำนวนเที่ยว'].sum()],
            'น้ำหนักต้นทางรวม (ตัน)': [group['น้ำหนักต้นทางรวม (ตัน)'].sum()],
            'น้ำหนักปลายทางรวม (ตัน)': [group['น้ำหนักปลายทางรวม (ตัน)'].sum()],
            'น้ำหนักที่จ่าย (ตัน)': [group['น้ำหนักที่จ่าย (ตัน)'].sum()],
            'ส่วนต่าง (ตัน)': [group['ส่วนต่าง (ตัน)'].sum()],
        })
        result.append(subtotal_row)

    df_final = pd.concat(result, ignore_index=True)

    # Convert numeric columns (in case they are str)
    df_final['จำนวนเที่ยว'] = pd.to_numeric(df_final['จำนวนเที่ยว'], errors='coerce')
    df_final['น้ำหนักต้นทางรวม (ตัน)'] = pd.to_numeric(df_final['น้ำหนักต้นทางรวม (ตัน)'], errors='coerce')
    df_final['น้ำหนักปลายทางรวม (ตัน)'] = pd.to_numeric(df_final['น้ำหนักปลายทางรวม (ตัน)'], errors='coerce')
    df_final['น้ำหนักที่จ่าย (ตัน)'] = pd.to_numeric(df_final['น้ำหนักที่จ่าย (ตัน)'], errors='coerce')
    df_final['ส่วนต่าง (ตัน)'] = pd.to_numeric(df_final['ส่วนต่าง (ตัน)'], errors='coerce')

    # Grand total row
    data_rows = df_final[~df_final['ทีม'].str.contains('รวมทีม', na=False)]
    grand_total_row = pd.DataFrame({
        'ทีม': ['รวมทั้งหมด'],
        'ต้นทาง - ปลายทาง': [''],
        'ช่วงน้ำหนัก': [''],
        'ชนิดหิน': [''],
        'จำนวนเที่ยว': [data_rows['จำนวนเที่ยว'].sum()],
        'น้ำหนักต้นทางรวม (ตัน)': [data_rows['น้ำหนักต้นทางรวม (ตัน)'].sum()],
        'น้ำหนักปลายทางรวม (ตัน)': [data_rows['น้ำหนักปลายทางรวม (ตัน)'].sum()],
        'น้ำหนักที่จ่าย (ตัน)': [data_rows['น้ำหนักที่จ่าย (ตัน)'].sum()],
        'ส่วนต่าง (ตัน)': [data_rows['ส่วนต่าง (ตัน)'].sum()],
    })

    # Append grand total to bottom
    df_final = pd.concat([df_final, grand_total_row], ignore_index=True)

    output = BytesIO()
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Transport({active}) {start_created} to {end_created}.xlsx'

    # Write Excel to memory buffer
    with BytesIO() as buffer:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_final.to_excel(writer, index=False, sheet_name='รายงาน', startrow=1)  # ขยับข้อมูลลงไปเริ่มแถวที่ 1

            workbook = writer.book
            sheet = writer.sheets['รายงาน']
            sheet.freeze_panes = "A3" #freeze

            str_start = datetime.strptime(start_created, '%Y-%m-%d').strftime('%d/%m/%Y')
            str_end = datetime.strptime(end_created, '%Y-%m-%d').strftime('%d/%m/%Y')
            
            # เพิ่มชื่อรายงานในแถวแรก
            sheet.merge_cells('A1:I1')  # ปรับช่วงตามจำนวนคอลัมน์ของคุณ
            title_cell = sheet['A1']
            title_cell.value = f'รายงานค่าบรรทุก {company.name} วันที่ {str_start} - {str_end}'
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Font styling
            bold_red_font = Font(bold=True, color="FF0000")
            right_align = Alignment(horizontal="right")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):

                team_value = str(row[0].value)
                row_num = row[0].row

                #แถวที่เป็นชื่อทีม (ไม่ใช่ subtotal หรือ รวมทั้งหมด)
                if team_value and not team_value.startswith('รวม'):
                    # merge A:E
                    sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
                    for col in range(1, 10):  # A=1 ถึง E=10
                        cell = sheet.cell(row=row_num, column=col)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        cell.border = thin_border

                #แถวรวมทีม / รวมทั้งหมด = สีแดงตัวหนา
                elif 'รวมทีม' in team_value or 'รวมทั้งหมด' in team_value:
                    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                        sheet.column_dimensions[col_letter].width = 25
                        cell = sheet[f'{col_letter}{row_num}']
                        cell.font = bold_red_font

            #ชิดขวาเฉพาะตัวเลข
            for col_letter in ['E', 'F', 'G', 'H', 'I']:
                for cell in sheet[col_letter][1:]:  # [1:] เพื่อข้าม header
                    cell.alignment = right_align

        # Save to response
        response.write(buffer.getvalue())

    return response

def exportExcelTransportByCompanyInDashboard(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    end_created = request.session['db_end_date']
    start_created = request.session['db_start_date']

    my_q = Q()
    if start_created is not None:
        my_q &= Q(date__gte = start_created)
    if end_created is not None:
        my_q &=Q(date__lte = end_created)

    my_q &= Q(car_team__isnull = False, bws__weight_type=1)
    my_q &= ~Q(customer_name ='ยกเลิก')

    #เปลี่ยนออกเป็น ดึงรายงานของเดือนนั้นๆเท่านั้น
    startDate = datetime.strptime(start_created, "%Y-%m-%d").date()
    endDate = datetime.strptime(end_created, "%Y-%m-%d").date()

    response = excelTransportByCompany(request, my_q, start_created, end_created)
    return response

def excelTransportByCompany(request, my_q, start_created, end_created):
    str_start = datetime.strptime(start_created, '%Y-%m-%d').strftime('%d/%m/%Y')
    str_end = datetime.strptime(end_created, '%Y-%m-%d').strftime('%d/%m/%Y')

    transport_comp = ['SLC', 'SLT', 'CTM', 'KT', 'UNI', 'STPS', 'TYM']
    all_comp = BaseCompany.objects.filter(code__in=transport_comp).values('code', 'name')

    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    has_data = False

    for comp in all_comp:
        sheet_name = comp['code'] + "-" + comp['name']
        my_q_with_comp = my_q & Q(bws__company__code=comp['code'], carry_type_name='ส่งให้')

        queryset = Weight.objects.filter(my_q_with_comp).values(
            'car_team__car_team_name', 'date', 'customer__customer_name',
            'site__base_site_name', 'stone_type__base_stone_type_name'
        ).annotate(
            num_rows=Count('weight_id'), 
            sum_weight=Sum('weight_total'),
            sum_amount_vat=Sum('amount_vat'),
        ).order_by('car_team__car_team_name')

        if queryset.exists():
            has_data = True

            df = pd.DataFrame(list(queryset))

            df.columns = [
                'ทีม','วันที่', 'ลูกค้า', 'ปลายทาง', 'ชนิดหิน', 
                'จำนวนเที่ยว', 'น้ำหนักรวม (ตัน)' , 'จำนวนเงินสุทธิ'
            ]

            if not df.empty:
                df['วันที่'] = pd.to_datetime(df['วันที่']).dt.strftime('%d/%m/%Y')
            
            df.fillna({'ทีม': '(ไม่มีทีม)'}, inplace=True)

            grouped = df.groupby('ทีม', dropna=False)
            result = []

            for name, group in grouped:
                result.append(group)
                subtotal = pd.DataFrame({
                    'ทีม': [f'รวมทีม {name}'],
                    'วันที่': [''],
                    'ลูกค้า': [''], 'ปลายทาง': [''], 'ชนิดหิน': [''],
                    'จำนวนเที่ยว': [group['จำนวนเที่ยว'].sum()],
                    'น้ำหนักรวม (ตัน)': [group['น้ำหนักรวม (ตัน)'].sum()],
                    'จำนวนเงินสุทธิ': [group['จำนวนเงินสุทธิ'].sum()],
                })
                result.append(subtotal)

            df = pd.concat(result, ignore_index=True)

            total_row = pd.DataFrame({
                'ทีม': ['รวมทั้งหมด'], 'วันที่': [''], 'ลูกค้า': [''], 'ปลายทาง': [''], 'ชนิดหิน': [''],
                'จำนวนเที่ยว': [df.loc[df['ทีม'].str.contains('รวมทีม', na=False) == False, 'จำนวนเที่ยว'].sum()],
                'น้ำหนักรวม (ตัน)': [df.loc[df['ทีม'].str.contains('รวมทีม', na=False) == False, 'น้ำหนักรวม (ตัน)'].sum()],
                'จำนวนเงินสุทธิ': [df.loc[df['ทีม'].str.contains('รวมทีม', na=False) == False, 'จำนวนเงินสุทธิ'].sum()],
            })

            df = pd.concat([df, total_row], ignore_index=True)

            df[['น้ำหนักรวม (ตัน)']] = df[[
                'น้ำหนักรวม (ตัน)'
            ]].applymap(lambda x: f"{x:,.2f}" if pd.notna(x) else "")

            df[['จำนวนเงินสุทธิ']] = df[[
                'จำนวนเงินสุทธิ'
            ]].applymap(lambda x: f"{x:,.2f}" if pd.notna(x) else "")

            safe_sheet_name = sheet_name[:31]  # Excel จำกัดชื่อ sheet 31 ตัวอักษร
            df.to_excel(writer, index=False, sheet_name=safe_sheet_name, startrow=1)  # เขียนเริ่มแถวที่ 2

    writer.close()

    # กลับไปอ่าน workbook จาก BytesIO เพื่อแก้ไขหัวตาราง
    output.seek(0)
    workbook = openpyxl.load_workbook(output)
    for comp in all_comp:
        sheet_name = (comp['code'] + "-" + comp['name'])[:31]
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet.insert_rows(1)  # แทรกแถวก่อนหน้า
            sheet.merge_cells('A1:H1')  # ปรับช่วงตามจำนวนคอลัมน์
            sheet.freeze_panes = "A4" #freeze
            title_cell = sheet['A1']
            title_cell.value = f"รายงานขนส่งตามบริษัท {comp['name']}  วันที่ {str_start} - {str_end}"
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Bold subtotal and total row
            for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
                for cell in row:
                    if 'รวมทีม' in str(cell.value) or 'รวมทั้งหมด' in str(cell.value):
                        row_number = cell.row  # Get the row number
                        # Bold and red font
                        bold_red_font = Font(bold=True, color="FF0000")

                        # Apply to columns A, G, H, I, J, K in the same row
                        for col in ['A', 'F', 'G', 'H']:
                            sheet[f"{col}{row_number}"].font = bold_red_font

            right_align = Alignment(horizontal="right")
            for col in ['H', 'F', 'G', 'H']:  # Columns for numbers
                for cell in sheet[col]:  # Iterate through all cells in that column
                    cell.alignment = right_align

            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                sheet.column_dimensions[col_letter].width = 25

    final_output = BytesIO()
    workbook.save(final_output)
    final_output.seek(0)

    response = HttpResponse(final_output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="transport_by_company.xlsx"'
    return response

def exportExcelTranToSellInDashboard(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)

    end_created = request.session['db_end_date']
    start_created = request.session['db_start_date']

    my_q = Q()
    if start_created is not None:
        my_q &= Q(date__gte = start_created)
    if end_created is not None:
        my_q &=Q(date__lte = end_created)

    my_q &= Q(bws__company__code__in = company_in)
    my_q &= ~Q(customer_name ='ยกเลิก')

    #เปลี่ยนออกเป็น ดึงรายงานของเดือนนั้นๆเท่านั้น
    startDate = datetime.strptime(start_created, "%Y-%m-%d").date()
    endDate = datetime.strptime(end_created, "%Y-%m-%d").date()

    #สร้าง list ระหว่าง start_date และ end_date
    list_date = [startDate+timedelta(days=x) for x in range((endDate-startDate).days + 1)]

    response = excelTranToSell(request, my_q, list_date)
    return response

def excelTranToSell(request, my_q, list_date):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    company = BaseCompany.objects.get(code = active)

    if company.biz.id == 1: #ธุรกิจเหมือง
        data1 = Weight.objects.filter(my_q, ~Q(site = '200PL') & ~Q(site = '300PL'), Q(customer__customer_name__contains='พอร์ท') | Q(customer__customer_name__contains='พอร์ต') | Q(customer__customer_name__contains='ท่าเรือ') , bws__weight_type = 1).order_by(
                                'date','customer','stone_type').values_list(
                                'date','customer_name', 'stone_type_name').annotate(
                                sum_weight_total = Sum('weight_total'))
        
        data2 = Weight.objects.filter(my_q, Q(site = '200PL') | Q(site = '300PL'), bws__weight_type = 1).order_by(
                                'date','site','stone_type').values_list(
                                'date','site_name', 'stone_type_name').annotate(
                                sum_weight_total = Sum('weight_total'))
        data = list(data1) + list(data2)

    # Create a new workbook and get the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    if data:
        worksheet.cell(row=1, column=1, value='Date')
        worksheet.merge_cells(start_row=1, start_column = 1, end_row=2, end_column=1)

        date_style = NamedStyle(name='custom_datetime', number_format='DD/MM/YYYY')
        
        # Create a set of all unique mill and cus values
        customer = set()
        stones = set()
        customer_data1 = [item[1] for item in data1]
        customer_data2 = [item[1] for item in data2]
        # ลบซ้ำภายในแต่ละกลุ่มก่อน
        customer_data1_unique = list(dict.fromkeys(customer_data1))
        customer_data2_unique = list(dict.fromkeys(customer_data2))

        # รวม โดยให้ data1 มาก่อน data2
        customer = customer_data1_unique + [st for st in customer_data2_unique if st not in customer_data1_unique]

        stones = list(dict.fromkeys([item[2] for item in data]))

        cus_col_list = []
        cus_colors = [generate_pastel_color() for _ in range(len(customer) + 1)]

        column_index = 2
        for st in customer:
            worksheet.cell(row=1, column=column_index, value=f'ยอดขนไป {st}')
            worksheet.merge_cells(start_row=1, start_column = column_index, end_row=1, end_column=(column_index + len(stones) + 1) -1 )
            
            cell = worksheet.cell(row=1, column=column_index)
            cell.alignment = Alignment(horizontal='center')

            info = {}
            info['st'] = st
            info['strat_col'] = column_index
            info['end_col'] = column_index + len(stones) + 1
            cus_col_list.append(info)

            #อัพเดทจำนวน col ตามที่มา
            column_index += len(stones) + 1

        #set color in header in row 1-2
        for row in worksheet.iter_rows(min_row=1, max_row=2):
            # Set the background color for each cell in the column
            for cell in row:
                #cell.border = Border(top=side, bottom=side, left=side, right=side)
                cell.alignment = Alignment(horizontal='center')
                line_index = (cell.column - 2) // (len(stones) + 1 )
                fill_color = cus_colors[line_index % len(cus_colors)]
                fill = PatternFill(start_color=fill_color, fill_type="solid")
                cell.fill = fill

        # Write headers row 2 to the worksheet
        column_index = 2
        for st in customer:
            for sou in stones:
                worksheet.cell(row=2, column=column_index, value=sou).alignment = Alignment(horizontal='center')
                column_index += 1
                
            worksheet.cell(row=2, column=column_index, value= 'Total').alignment = Alignment(horizontal='center')
            worksheet.cell(row=2, column=column_index).font = Font(bold=True, color="FF0000")
            column_index += 1


        # Create a dictionary to store data by date, mill, and cus
        date_data = {}

        # Loop through the data and populate the dictionary  
        for item in data:
            date = item[0]
            cus = item[1]
            stone = item[2]
            quantity = item[3]

            if date not in date_data:
                date_data[date] = {}

            if cus not in date_data[date]:
                date_data[date][cus] = {'stones': {}, 'total': 0}

            date_data[date][cus]['stones'][stone] = quantity
            date_data[date][cus]['total'] += quantity  # Store the total for this cus

        row_index = 3
        for idl, ldate in enumerate(list_date):
            #เขียนวันที่ใน worksheet column 1
            worksheet.cell(row=idl+3, column=1, value=ldate).style = date_style
            worksheet.cell(row=idl+3, column=1).alignment = Alignment(horizontal='center')

            for date, cus_data in date_data.items():
                #เขียน weight total ของแต่ละหินใน worksheet
                if worksheet.cell(row=idl+3, column = 1).value == date:
                    column_index = 2
                    for st in customer:
                        stone_data = cus_data.get(st, {}).get('stones', {})
                        total_value = cus_data.get(st, {}).get('total', '')

                        # Write quantities by stone
                        for sou in stones:
                            value = stone_data.get(sou, '')
                            worksheet.cell(row=idl + 3, column=column_index, value=value).number_format = '#,##0.00'
                            column_index += 1

                        # Write the ssn__total value for the cus
                        worksheet.cell(row=idl + 3, column=column_index, value=total_value).number_format = '#,##0.00'
                        worksheet.cell(row=idl + 3, column=column_index).font = Font(bold=True, color="FF0000")
                        column_index += 1
            row_index += 1

        worksheet.cell(row=row_index, column=1, value='รวมทั้งสิ้น')
        sum_by_col = Decimal('0.00')
        for col in range(2, column_index):
            for row in range(3, row_index):
                sum_by_col = sum_by_col + Decimal( worksheet.cell(row=row, column=col).value or '0.00' )
            worksheet.cell(row=row_index, column=col, value=sum_by_col).number_format = '#,##0.00'
            worksheet.cell(row=row_index, column=col).font = Font(bold=True, color="FF0000")
            sum_by_col = Decimal('0.00')

        # Set the column widths
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[2].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width
            worksheet.column_dimensions[column].height = 20

        side = Side(border_style='thin', color='000000')
        set_border(worksheet, side)

    else:
        worksheet.cell(row = 1, column = 1, value = f'ไม่มีข้อมูลขนหินไปจำหน่ายหินของเดือนนี้')

    # Save workbook into memory
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    size = output.getbuffer().nbytes

    # Generator to stream file in chunks
    def file_iterator(buffer, chunk_size=8192):
        while True:
            data = buffer.read(chunk_size)
            if not data:
                break
            yield data

    response = StreamingHttpResponse(
        file_iterator(output),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="tran_to_sell({active}).xlsx"'
    response["Content-Length"] = str(size)
    return response

@login_required(login_url='login')
def viewLoadingRate(request):
    #active : active คือแท็ปบริษัท active
    try:
        active = request.session['company_code']
        company_in = findCompanyIn(request)
    except:
        return redirect('logout')

    data = LoadingRate.objects.filter(company__code__in = company_in).order_by('-created')

    #กรองข้อมูล
    myFilter = LoadingRateFilter(request.GET, queryset = data)
    data = myFilter.qs

    #สร้าง page
    p = Paginator(data, 10)
    page = request.GET.get('page')
    lr = p.get_page(page)

    context = {'ldr_page':'active', 'lr': lr,'filter':myFilter, active :"active",}
    return render(request, "loadingRate/viewLoadingRate.html",context)

@login_required(login_url='login')
def createLoadingRate(request):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    base_weight_rang = BaseWeightRange.objects.filter(company = company).order_by('rate_min', 'rate_max')
    LoadingRateItemFormSet = modelformset_factory(LoadingRateItem, fields=('wt_range', 'tru_scoop', 'tru_shipp', 'chi_scoop', 'chi_shipp', 'bh_tru_scoop', 'bh_chi_scoop'), extra=len(base_weight_rang), widgets={})
    
    if request.method == 'POST':
        form = LoadingRateForm(request.POST)
        lrl_form = LoadingRateLocForm(request, request.POST or None)
        formset = LoadingRateItemFormSet(request.POST)
        if form.is_valid() and lrl_form.is_valid() and formset.is_valid():
            form = form.save()

            if  lrl_form.cleaned_data.get('site') or lrl_form.cleaned_data.get('mill'):
                lrl = lrl_form.save()
                lrl.Lr = form
                lrl.save()

                formset_instances = formset.save(commit=False)
                for instance in formset_instances:
                    instance.Lrl = lrl
                    instance.Lr = lrl.Lr
                    instance.save()

            return HttpResponseRedirect(reverse('editStep2LoadingRate', args=(lrl.Lr,)))

    else:
        form = LoadingRateForm(initial={'company': company})
        lrl_form = LoadingRateLocForm(request)
        formset = LoadingRateItemFormSet(queryset=LoadingRateItem.objects.none())

    context = {'ldr_page':'active', 'form': form, 'lrl_form': lrl_form, 'formset' : formset, 'base_stock_source': base_weight_rang, active :"active", 'disabledTab' : 'disabled', 'is_edit_stock': is_edit_stock(request.user)}
    return render(request, "loadingRate/createLoadingRate.html",context)


@login_required(login_url='login')
def editStep2LoadingRate(request, lr_id):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    base_weight_rang = BaseWeightRange.objects.filter(company = company).order_by('rate_min', 'rate_max')
    LoadingRateItemFormSet = modelformset_factory(LoadingRateItem, fields=('wt_range', 'tru_scoop', 'tru_shipp', 'chi_scoop', 'chi_shipp', 'bh_tru_scoop', 'bh_chi_scoop'), extra=len(base_weight_rang), widgets={})
    
    try:
        lr_data = LoadingRate.objects.get(id=lr_id)
    except LoadingRate.DoesNotExist:
        return redirect('viewLoadingRate')

    lrl_data = LoadingRateLoc.objects.filter(Lr=lr_data)

    if request.method == 'POST':
        form = LoadingRateForm(request.POST, instance=lr_data)
        lrl_form = LoadingRateLocForm(request, request.POST or None)
        formset = LoadingRateItemFormSet(request.POST)
        
        if form.is_valid() and lrl_form.is_valid() and formset.is_valid():
            form = form.save()

            if  lrl_form.cleaned_data.get('site') or lrl_form.cleaned_data.get('mill'):
                lrl = lrl_form.save()
                lrl.Lr = form
                lrl.save()

                formset_instances = formset.save(commit=False)
                for instance in formset_instances:
                    instance.Lrl = lrl
                    instance.Lr = lrl.Lr
                    instance.save()

            return HttpResponseRedirect(reverse('editStep2LoadingRate', args=(lr_id,)))
    else:
        form = LoadingRateForm(instance=lr_data)
        lrl_form = LoadingRateLocForm(request)
        formset = LoadingRateItemFormSet(queryset=LoadingRateItem.objects.none())

    context = {'ldr_page':'active', 'form': form, 'lrl_form': lrl_form, 'formset' : formset, 'base_weight_rang': base_weight_rang, 'lrl_data': lrl_data,'lr_data':lr_data, active :"active", 'disabledTab' : 'disabled',}
    return render(request, "loadingRate/editStep2LoadingRate.html",context)

@login_required(login_url='login')
def editLoadingRateItem(request, lr_id, lrl_id):
    active = request.session['company_code']
    company = BaseCompany.objects.get(code = active)

    base_weight_rang = BaseWeightRange.objects.filter(company = company).order_by('rate_min', 'rate_max')
    
    try:
        lr_data = LoadingRate.objects.get(id=lr_id)
    except LoadingRate.DoesNotExist:
        return redirect('viewLoadingRate')

    lrl_data = LoadingRateLoc.objects.filter(Lr = lr_id)#ssn all in stock id
    data = LoadingRateLoc.objects.get(id = lrl_id)#id edit

    if data.mill and data.site:
        lrl_ms = f"{data.mill.mill_id}{data.site.base_site_id}"
    elif data.mill:
        lrl_ms = f"{data.mill.mill_id}"
    elif data.site:
        lrl_ms = f"{data.site.base_site_id}"

    if request.method == 'POST':
        form = LoadingRateForm(request.POST, instance=lr_data)
        lrl_form = LoadingRateLocForm(request, request.POST, instance=data)
        formset = LoadingRateItemInlineFormset(request.POST, instance=data)
        
        if form.is_valid() and lrl_form.is_valid() and formset.is_valid():
            form = form.save()

            lrl = lrl_form.save(commit=False)
            if  lrl_form.cleaned_data.get('site') or lrl_form.cleaned_data.get('mill'):
                lrl.Lr = form
                lrl.save()

                formset_instances = formset.save(commit=False)
                for instance in formset_instances: #อันนี้ไม่มี deleted_objects นะ
                    instance.save()

            return HttpResponseRedirect(reverse('editStep2LoadingRate', args=(lr_id,)))
    else:
        form = LoadingRateForm(instance=lr_data)
        lrl_form = LoadingRateLocForm(request,instance=data)
        formset = LoadingRateItemInlineFormset(instance=data)

    context = {'ldr_page':'active', 'form': form, 'lrl_form': lrl_form, 'formset' : formset, 'base_weight_rang': base_weight_rang, 'lrl_data': lrl_data, 'lr_data':lr_data, 'lrl_id': data.id, 'lrl_ms' : lrl_ms, active :"active", 'disabledTab' : 'disabled',}
    return render(request, "loadingRate/editLoadingRateItem.html",context)


@login_required(login_url='login')
def removeLoadingRate(request, lr_id):
    lr = LoadingRate.objects.get(id = lr_id)

    #ลบ LoadingRateLoc ใน LoadingRate ด้วย
    all_lrl = LoadingRateLoc.objects.filter(Lr = lr)
    for lrl in all_lrl:
        items = LoadingRateItem.objects.filter(Lrl = lrl)
        items.delete()

    all_lrl.delete()
    lr.delete()

    #updateTotalStockInMonthByDate(previous_day, tmp_company)#ดึงข้อมูล stock ก่อนหน้ามาเพื่อ update วันถัดไป
    return redirect('viewLoadingRate')

@login_required(login_url='login')
def removeLoadingRateLoc(request, lrl_id):
    #ลบ LoadingRateItem ใน LoadingRateLoc ด้วย
    lrl = LoadingRateLoc.objects.get(id = lrl_id)
    lr_id = lrl.Lr

    items = LoadingRateItem.objects.filter(Lrl = lrl)
    items.delete()

    lrl.delete()

    #updateTotalStockInMonthByDate(previous_day, tmp_company)#ดึงข้อมูล stock ก่อนหน้ามาเพื่อ update วันถัดไป
    return HttpResponseRedirect(reverse('editStep2LoadingRate', args=(lr_id,)))

def searchLRInDay(request):
    if 'date_start_rate' in request.GET and 'company' in request.GET and 'lr_id' in request.GET:
        date_start_rate =  request.GET.get('date_start_rate')
        company =  request.GET.get('company')
        lr_id =  request.GET.get('lr_id')

        if lr_id == '':
            have_lr = LoadingRate.objects.filter(company = company, date_start_rate = date_start_rate).exists()
        else:
            have_lr = LoadingRate.objects.filter(~Q(id = lr_id), company = company, date_start_rate = date_start_rate).exists()
    data = {
        'have_lr' :have_lr,
    }
    return JsonResponse(data)

def rate_subquery(value_field, ignore_mill=False, ignore_site=False):
    filters = {
        'Lr__date_start_rate__lte': OuterRef('date'),
        'wt_range__rate_min__lte': OuterRef('weight_total'),
        'wt_range__rate_max__gt': OuterRef('weight_total'),
    }

    if not ignore_site:
        filters['Lrl__site'] = OuterRef('site')

    if not ignore_mill:
        filters['Lrl__mill'] = OuterRef('mill')

    return (
        LoadingRateItem.objects
        .filter(**filters)
        .order_by('-Lr__date_start_rate')
        .values(value_field)[:1]
    )
    
def exportWeightLoadingRate(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    comp = BaseCompany.objects.get(code = active)

    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None

    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    if start_created is None and end_created is None:
        start_created = previous_date_time.strftime("%Y-%m-%d")
        end_created = previous_date_time.strftime("%Y-%m-%d")

    str_start = datetime.strptime(start_created, '%Y-%m-%d').strftime('%d/%m/%Y')
    str_end = datetime.strptime(end_created, '%Y-%m-%d').strftime('%d/%m/%Y')

    my_q = Q()
    if start_created is not None:
        my_q &= Q(date__gte = start_created)
    if end_created is not None:
        my_q &=Q(date__lte = end_created)

    my_q &=Q(bws__company__code__in = company_in, is_cancel = False)

    # 🔹 LR มีทั้ง mill + site
    lr_mill_site = LoadingRateLoc.objects.filter(
        Lr__company__code__in=company_in,
        Lr__date_start_rate__gte=start_created,
        Lr__date_start_rate__lte=end_created,
        mill=OuterRef('mill'),
        site=OuterRef('site'),
    )

    # 🔹 LR มีเฉพาะ site (mill ต้อง NULL)
    lr_site_only = LoadingRateLoc.objects.filter(
        Lr__company__code__in=company_in,
        Lr__date_start_rate__gte=start_created,
        Lr__date_start_rate__lte=end_created,
        site=OuterRef('site'),
        mill__isnull=True,
    )

    # 🔹 LR มีเฉพาะ mill (site ต้อง NULL)
    lr_mill_only = LoadingRateLoc.objects.filter(
        Lr__company__code__in=company_in,
        Lr__date_start_rate__gte=start_created,
        Lr__date_start_rate__lte=end_created,
        mill=OuterRef('mill'),
        site__isnull=True,
    )

    queryset = (
        Weight.objects
        .filter(my_q)
        .annotate(
            has_lr=Case(
                # ✅ CASE 1: MILL + SITE (เฉพาะที่สุด ต้องมาก่อน)
                When(
                    Exists(lr_mill_site),
                    then=True,
                ),

                # ✅ CASE 2: SITE ONLY
                When(
                    Q(site__in=['200PL','300PL']) & Exists(lr_site_only),
                    then=True,
                ),

                # ✅ CASE 3: MILL ONLY
                When(
                    Exists(lr_mill_only),
                    then=True,
                ),

                default=False,
                output_field=models.BooleanField(),
            )
        )
        .filter(has_lr=True)
    )

    if queryset:
        ########## อัตราค่าขน/บรรทุก ##########
        shipping_tru = Case(
            # 🟢 CASE 1: WT1 + SITE (LR มีเฉพาะ site)
            When(
                bws__weight_type__id=1,
                site__in=['200PL', '300PL'],
                then=Subquery(
                    rate_subquery('tru_shipp', ignore_mill=True),
                    output_field=models.DecimalField()
                ),
            ),

            # 🟢 CASE 2: WT1 + MILL (LR มีเฉพาะ mill)
            When(
                bws__weight_type__id=1,
                then=Subquery(
                    rate_subquery('tru_shipp', ignore_site=True),
                    output_field=models.DecimalField()
                ),
            ),

            default=Subquery(
                rate_subquery('tru_shipp'),
                output_field=models.DecimalField()
            ),
            output_field=models.DecimalField(),
        )
        
        ########## อัตราค่าขน/บรรทุก ##########
        shipping_chi_sub = (
            LoadingRateItem.objects
            .filter(
                Lr__date_start_rate__lte=OuterRef('date'),
                wt_range__rate_min__lte=OuterRef('weight_total'),
                wt_range__rate_max__gt=OuterRef('weight_total'),
                Lrl__mill=OuterRef('mill'),
                Lrl__site=OuterRef('site'),
            )
            .order_by('-Lr__date_start_rate')
            .values('chi_shipp')[:1]
        )

        ########## อัตราค่าตักจากรถแบ็คโฮ ##########
        bh_scoop_tru = Case(
            # 🟢 CASE 1: WT1 + SITE (LR มีเฉพาะ site)
            When(
                bws__weight_type__id=1,
                site__in=['200PL', '300PL'],
                then=Subquery(
                    rate_subquery('bh_tru_scoop', ignore_mill=True),
                    output_field=models.DecimalField()
                ),
            ),

            # 🟢 CASE 2: WT1 + MILL (LR มีเฉพาะ mill)
            When(
                bws__weight_type__id=1,
                then=Subquery(
                    rate_subquery('bh_tru_scoop', ignore_site=True),
                    output_field=models.DecimalField()
                ),
            ),
            default=Subquery(
                rate_subquery('bh_tru_scoop'),
                output_field=models.DecimalField()
            ),
            output_field=models.DecimalField(),
        )
        
        ########## อัตราค่าตักจากรถแบ็คโฮ ##########
        bh_scoop_chi_sub = (
            LoadingRateItem.objects
            .filter(
                Lr__date_start_rate__lte=OuterRef('date'),
                wt_range__rate_min__lte=OuterRef('weight_total'),
                wt_range__rate_max__gt=OuterRef('weight_total'),
                Lrl__mill=OuterRef('mill'),
                Lrl__site=OuterRef('site'),
            )
            .order_by('-Lr__date_start_rate')
            .values('bh_chi_scoop')[:1]
        )

        ########## อัตราค่าตักจากรถตัก ##########
        scoop_tru = Case(
            # 🟢 CASE 1: WT1 + SITE (LR มีเฉพาะ site)
            When(
                bws__weight_type__id=1,
                site__in=['200PL', '300PL'],
                then=Subquery(
                    rate_subquery('tru_scoop', ignore_mill=True),
                    output_field=models.DecimalField()
                ),
            ),

            # 🟢 CASE 2: WT1 + MILL (LR มีเฉพาะ mill)
            When(
                bws__weight_type__id=1,
                then=Subquery(
                    rate_subquery('tru_scoop', ignore_site=True),
                    output_field=models.DecimalField()
                ),
            ),
            # 🔹 default: mill + site ปกติ
            default=Subquery(
                rate_subquery('tru_scoop'),
                output_field=models.DecimalField()
            ),
            output_field=models.DecimalField(),
        )
        
        ########## อัตราค่าตักจากรถตัก ##########
        scoop_chi_sub = (
            LoadingRateItem.objects
            .filter(
                Lr__date_start_rate__lte=OuterRef('date'),
                wt_range__rate_min__lte=OuterRef('weight_total'),
                wt_range__rate_max__gt=OuterRef('weight_total'),
                Lrl__mill=OuterRef('mill'),
                Lrl__site=OuterRef('site'),
            )
            .order_by('-Lr__date_start_rate')
            .values('chi_scoop')[:1]
        )
        
        qs = (
            queryset
            .annotate(
                scoop_tru=scoop_tru,
                scoop_chi=Subquery(scoop_chi_sub, output_field=models.DecimalField()),

                bh_scoop_tru=bh_scoop_tru,
                bh_scoop_chi=Subquery(bh_scoop_chi_sub, output_field=models.DecimalField()),

                shipping_tru=shipping_tru,
                shipping_chi=Subquery(shipping_chi_sub, output_field=models.DecimalField()),
            )
            .annotate(
                shipping_rate=Case(
                    When(car_registration__car_type='สิบล้อ', then=F('shipping_tru')),
                    When(car_registration__car_type='จีน', then=F('shipping_chi')),
                    default=F('shipping_tru'),
                    output_field=models.DecimalField(),
                ),
                bh_scoop_rate=Case(
                    When(car_registration__car_type='สิบล้อ', then=F('bh_scoop_tru')),
                    When(car_registration__car_type='จีน', then=F('bh_scoop_chi')),
                    default=F('bh_scoop_tru'),
                    output_field=models.DecimalField(),
                ),
                scoop_rate=Case(
                    When(car_registration__car_type='สิบล้อ', then=F('scoop_tru')),
                    When(car_registration__car_type='จีน', then=F('scoop_chi')),
                    default=F('scoop_tru'),
                    output_field=models.DecimalField(),
                ),
            )
        )

        data = {
                'วันที่': queryset.values_list('date', flat=True),
                'เลขที่ใบชั่ง': queryset.values_list('doc_id', flat=True),
                'ทะเบียนรถ': queryset.values_list('car_registration_name', flat=True),
                'ชื่อคนขับ': queryset.values_list('driver_name', flat=True),
                'ต้นทาง': queryset.values_list('mill_name', flat=True),
                'ปลายทาง': queryset.values_list('site_name', flat=True),
                'น้ำหนักสุทธิ': queryset.values_list('weight_total', flat=True),
                'เครื่องชั่ง': queryset.values_list('bws__weight_type__name', flat=True),
        }
        
        data2 = {
            'อัตราค่าขน/บรรทุก': [],
            'ค่าขน/บรรทุก': [],
            'อัตราค่าตักจากรถแบ็คโฮ': [],
            'ค่าตักจากรถแบ็คโฮ': [],
            'อัตราค่าตักจากรถตัก': [],
            'ค่าตักจากรถตัก': [],
        }

        for row in qs.values(
            'shipping_rate',
            'bh_scoop_rate',
            'scoop_rate',
            'weight_total'
        ):
            w = row['weight_total']

            data2['อัตราค่าขน/บรรทุก'].append(row['shipping_rate'])
            data2['ค่าขน/บรรทุก'].append(cal_ld_rate(row['shipping_rate'], w))

            data2['อัตราค่าตักจากรถแบ็คโฮ'].append(row['bh_scoop_rate'])
            data2['ค่าตักจากรถแบ็คโฮ'].append(cal_ld_rate(row['bh_scoop_rate'], w))

            data2['อัตราค่าตักจากรถตัก'].append(row['scoop_rate'])
            data2['ค่าตักจากรถตัก'].append(cal_ld_rate(row['scoop_rate'], w))
    else:
        data = {'ข้อความ': ['ไม่มีข้อมูลรายการชั่งจาก วันที่เลือก']}
        data2 = {'' : []}

    data.update(data2)
    df = pd.DataFrame(data)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet_name = 'Sheet1'

        # Write DataFrame (start from row 2)
        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=1)

        ws = writer.book[sheet_name]

        # Report Header
        ws['A1'] = f"รายงานค่าขน/ค่าตัก {comp.name}  วันที่ {str_start} - {str_end}"
        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=len(df.columns)
        )

        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.freeze_panes = ws["A3"]

        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col_letter].width = 13
        
        for col_letter in ['I', 'J', 'K', 'L', 'M', 'N']:
            ws.column_dimensions[col_letter].width = 23

    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=weight_loading_rate({active}) '+ start_created + " to "+ end_created +'.xlsx'
    return response

def cal_ld_rate(rate, weight_total):
    result = None
    if rate:
        result = Decimal(rate) * Decimal(weight_total)
        result = round(result, 2)
    return result

def exportLoadingRate(request):
    active = request.session['company_code']
    company_in = findCompanyIn(request)
    comp = BaseCompany.objects.get(code = active)

    start_created = request.GET.get('start_created') or None
    end_created = request.GET.get('end_created') or None

    current_date_time = datetime.today()
    previous_date_time = current_date_time - timedelta(days=1)

    if start_created is None and end_created is None:
        start_created = previous_date_time.strftime("%Y-%m-%d")
        end_created = previous_date_time.strftime("%Y-%m-%d")

    str_start = datetime.strptime(start_created, '%Y-%m-%d').strftime('%d/%m/%Y')
    str_end = datetime.strptime(end_created, '%Y-%m-%d').strftime('%d/%m/%Y')

    my_q = Q()
    if start_created is not None:
        my_q &= Q(date_start_rate__gte = start_created)
    if end_created is not None:
        my_q &=Q(date_start_rate__lte = end_created)

    my_q &=Q(company__code = active)

    base_wt = BaseWeightRange.objects.filter(company__code=active)
    lr_data = LoadingRate.objects.filter(my_q)

    workbook = openpyxl.Workbook()

    # =========================
    # Border style
    # =========================
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    if lr_data.exists():
        for lr in lr_data:
            sheet = workbook.create_sheet(title=lr.date_start_rate.strftime('%d-%m-%Y'))
            # Freeze row 1–2
            sheet.freeze_panes = "B4"

            # Header ซ้าย
            sheet.cell(row=1, column=1, value="น้ำหนักรถ")
            sheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
            sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center',vertical='center')

            # น้ำหนักรถ (แนวตั้ง)
            data_start_row = 4
            for i, wt in enumerate(base_wt):
                sheet.cell(row=data_start_row + i, column=1, value=wt.name)

            # Location rate
            lr_loc = LoadingRateLoc.objects.filter(Lr=lr)

            hd_colors = [generate_pastel_color() for i  in range(len(lr_loc) + 1)]
            col1_fill = PatternFill(start_color=hd_colors[0],end_color=hd_colors[0],fill_type="solid")

            column_index = 2
            color_index = 1
            for loc in lr_loc:
                fill = PatternFill(start_color=hd_colors[color_index],end_color=hd_colors[color_index],fill_type="solid")

                start_col = column_index
                end_col = column_index + 5
                
                # header
                if loc.mill and loc.site:
                    header = f"{loc.mill.mill_name} - {loc.site.base_site_name}"
                elif loc.mill:
                    header = f"{loc.mill.mill_name}"
                elif loc.site:
                    header = f"- {loc.site.base_site_name}"

                sheet.cell(row=1, column=start_col, value=header)
                sheet.merge_cells(start_row=1,start_column=start_col,end_row=1,end_column=end_col)

                # color header
                for col in range(start_col, end_col + 1):
                    cell = sheet.cell(row=1, column=col)
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                sub_col = start_col
                # sub header
                sub_headers1 = ["อัตราค่าขน/บรรทุก","อัตราค่าตักจากแบ็คโฮ","อัตราค่าตักจากรถตัก",]
                for i, text in enumerate(sub_headers1):
                    cell1 = sheet.cell(row=2, column=sub_col + i, value=text)
                    cell2 = sheet.cell(row=2, column=sub_col + i + 1, value=text)
                    sheet.merge_cells(start_row=2, start_column=sub_col + i, end_row=2 ,end_column=sub_col + i + 1)

                    cell1.fill = fill
                    cell1.alignment = Alignment(horizontal="center", vertical="center")

                    cell2.fill = fill
                    cell2.alignment = Alignment(horizontal="center", vertical="center")

                    sub_col += 1

                # sub header
                sub_headers2 = ["สิบล้อ","รถจีน","สิบล้อ","รถจีน","สิบล้อ","รถจีน",]
                for i, text in enumerate(sub_headers2):
                    cell = sheet.cell(row=3, column=start_col + i, value=text)
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # -------- Data --------
                for wt_index, wt in enumerate(base_wt):
                    try:
                        item = LoadingRateItem.objects.get(Lrl=loc,wt_range=wt)
                    except LoadingRateItem.DoesNotExist:
                        item = None

                    current_row = data_start_row + wt_index

                    sheet.cell(current_row, start_col, item.tru_shipp if item else None)
                    sheet.cell(current_row, start_col + 1, item.chi_shipp if item else None)

                    sheet.cell(current_row, start_col + 2, item.bh_tru_scoop if item else None)
                    sheet.cell(current_row, start_col + 3, item.bh_chi_scoop if item else None)

                    sheet.cell(current_row, start_col + 4,     item.tru_scoop if item else None)
                    sheet.cell(current_row, start_col + 5, item.chi_scoop if item else None)

                column_index += 6
                color_index += 1

            # border (ทั้ง sheet)
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row,min_col=1,max_col=sheet.max_column):
                for cell in row:
                    cell.border = thin_border

            # Auto width column
            for col in range(1, sheet.max_column + 1):
                col_letter = get_column_letter(col)
                max_length = 0

                for cell in sheet[col_letter]:
                    if cell.value:
                        max_length = max(max_length, 16)

                sheet.column_dimensions[col_letter].width = max_length + 3

            # add color colunm 1 
            for r in (1, 2):
                cell = sheet.cell(row=r, column=1)
                cell.fill = col1_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True)

            for r in range(4, sheet.max_row + 1):
                cell = sheet.cell(row=r, column=1)
                cell.fill = col1_fill

        workbook.remove(workbook['Sheet'])
    else:
        worksheet = workbook.active
        worksheet.cell(row = 1, column = 1, value = f'ไม่มีข้อมูลบันทึกเรทราคาค่าขน/ตัก')

    # Save workbook into memory
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    size = output.getbuffer().nbytes

    # Generator to stream file in chunks
    def file_iterator(buffer, chunk_size=8192):
        while True:
            data = buffer.read(chunk_size)
            if not data:
                break
            yield data

    response = StreamingHttpResponse(
        file_iterator(output),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="loading_rate({active}).xlsx"'
    response["Content-Length"] = str(size)
    return response